"""
Property Tax Record Scraper  v2.8.2
====================================
Automated property data extraction for Snohomish County AND King County.
County is auto-detected from the address city name.

v2.8.2 Changes (Snohomish building detail: beds / baths / sqft / construction):
    - Snohomish now also pulls the structure detail that lives one click deeper,
      on each parcel's Building Detail page (linked from the summary's Buildings
      table). snoho_fetch_detail() follows that link (carried in the DataDisplay
      row's hyperlink) and reads bedrooms, full + half baths, finished sqft,
      grade/quality, condition, floors, exterior, roof, foundation, heat, and
      fireplace from the building page's DataDisplay tables. This closes the last
      gap: Snohomish now returns MORE structure detail than King County.
    - _dd_rows() refactored to take a generic context-param dict ({p,a,m} for the
      summary, {b,a,p,d} for a building page) and to surface cell hyperlinks.

v2.8.1 Changes (Snohomish rich detail restored):
    - Snohomish leads now get assessed value, market total, last sale date + price,
      full sales history, and year built again - not just owner/parcel/address.
      The data comes from the parcel's Property Account Summary page, where each
      fact is a "DataDisplay" JSON table fetched from /API/DataDisplay/DataSources/
      GetData keyed by the table's ModuleId (read off the page; tables identified
      by column name so it survives ModuleId changes). snoho_fetch_detail() pulls
      it; scrape_property() chains it onto the search result (which supplies the
      parcel keys). Best-effort: a detail hiccup never drops the lead.
    - Beds/baths/finished sqft are NOT published on this portal, so they stay blank
      (the one remaining gap vs the old snoco.org scrape and vs King County).

v2.8.0 Changes (Snohomish County moved to the Public Access portal):
    - REBUILT the Snohomish path. The county retired the old snoco.org/proptax
      ASP.NET site (every lookup 404'd, the whole Snohomish pipeline was down).
      Property data now comes from the hosted Public Access portal's QuickSearch
      JSON API: get_session() primes the page for the DNN auth context (TabId +
      anti-forgery token); scrape_property() calls
      /DesktopModules/QuickSearch/API/Module/GetData?keywords=<house# + street
      name>&page=1 and maps the result to the master schema. Keywords drop street
      types/directions (the portal matches better that way); results are filtered
      to the input street and disambiguated by city, skipping ambiguous matches
      rather than guessing. get_session()/scrape_property() keep their old
      signatures, so main(), reenrich, and lead_intake call them unchanged.
    - The public search result carries parcel + owner + situs (enough for
      ownership verification). Rich assessor detail (assessed value, beds/baths,
      year, sale history) is NOT exposed by the public search API and is left
      blank, the same as a structure-less parcel. (A future enhancement could
      pull it from the Aumentum detail module.) King County is unchanged.

v2.7.4 Changes (King County direction-in-street-name resolve):
    - FIXED: a King County address whose STREET NAME contains a direction word
      failed to resolve. kc_normalize_address abbreviates EAST->E (etc.), but for
      streets like "East Ames Lake Drive" the county stores the full word
      ('3120 EAST AMES LAKE DR NE'), so the exact/prefix ADDR_FULL match missed.
      kc_resolve_address_live now adds a relaxed fallback: house number + the core
      street-name words (directionals and street types dropped), accepted ONLY
      when exactly one parcel matches, so it recovers these without ever silently
      grabbing a wrong parcel. (3120 E Ames Lake Dr -> parcel 0203101295.)

v2.7.3 Changes (King County condominium / townhome leads):
    - FIXED: condo/townhome leads landed as UNENRICHED. King County's address
      layer maps a unit's street address to the condominium COMPLEX MASTER parcel
      (MINOR 0000), which is common area carrying no owner and no value. The
      scraper fetched that empty master, got no owner, and could not match. It now
      detects the owner-less master, confirms via GIS land-use that it is a
      condominium, enumerates the per-unit parcels (MAJOR-00X0), and matches the
      lead to their unit by surname (disambiguated by first name when a surname
      repeats). Recovers e.g. '26810 NE Big Rock Rd' -> unit 729910-0090 (LESTER
      DEAN T, $818K, "Ridge at Big Rock Duv Condo").
    - A lead with no name, a non-condo owner-less parcel, or an absent/ambiguous
      unit match is left as-is (UNENRICHED) rather than guessing a wrong unit -
      consistent with the project's "never silently match the wrong property" rule.
      (e.g. '14715 1st Lane NE' / Michael Hoar, who owns no unit in that complex.)
    - Condo unit pages carry owner + assessed value but usually no beds/baths/year
      and no street address; the complex street from the GIS resolve is preserved
      for display, so enrichment is verified-ownership grade for these.

v2.7.2 Changes (address-matching robustness):
    - FIXED: clean_address_for_search truncated streets whose name starts with a
      state-code prefix (case-insensitive 'WA','OR','CA','ID','MT'): 'Wagner Rd'
      and 'Callow Rd' collapsed to just the house number. The strip regex now
      requires a word boundary + end anchor so it only removes a trailing state+zip.
    - FIXED: detect_county misrouted 'street, City ST zip' (city/state/zip in one
      comma segment) to Snohomish - _extract_city_segment now strips the trailing
      state+zip so King cities (Duvall, Snoqualmie, Redmond...) route correctly.
    - FIXED: the INITIAL search took results[0] without a street check, grabbing a
      wrong-street parcel (e.g. '13333 Wagner Rd' -> '13333 11th Ave W'). Initial
      results are now street-filtered (like the wildcard retry), with city used to
      disambiguate same-street-different-city collisions ('703 10th St' in
      Snohomish vs Mukilteo); ambiguous multi-city matches skip rather than guess.

v2.7.1 Changes:
    - FIXED (Bug 3): slash-format co-owner names. The assessor packs two owners
      with different surnames using a slash, e.g.
      'GUZMAN NOE BERNARDINO/BENAVIDES-ARAIZA CARINA'. split_owner_names() only
      split on ' & ', so these collapsed into one garbled owner and produced
      FALSE ownership non-matches (e.g. Noe Guzman and Naresh Chivukula, both
      actual owners, were flagged REVIEW/MISMATCH). '/' is now a recognized
      co-owner separator alongside ' & '. A 'C/O' care-of marker is excluded.
      Existing rows scraped before this fix need a re-scrape to correct them.

v2.7.0 Changes:
    - KING COUNTY IS NOW LIVE. The bulk ZIP data path (v2.3.0-v2.6.0) is
      RETIRED. KC data comes from two public HTTP GETs, no session, no
      viewstate, no 700MB cache, no five ZIP files:
        1. address -> PIN (+ city + zip) via the King County ArcGIS
           AddressPoints feature layer (kc_resolve_address_live).
        2. PIN -> live assessor detail via Dashboard.aspx?ParcelNbr=<PIN>,
           parsed by the existing parse_kc_dashboard (kc_fetch_dashboard_live).
      lookup_property_kc_live() orchestrates the two and replaces the bulk
      lookup_property_kc() in the routing path.
    - The old eReal search-form scrape was already dead (King County redesigned
      the site to a single global-search box; the txtAddress/btn_SearchAddress
      fields no longer exist). The new path sidesteps the form entirely.
    - KC rows now carry REAL owner names from the live assessor record (the
      public bulk file stripped them), so both counties use the same full
      ownership-match logic. compute_ownership_match_kc() is no longer called.
    - City and zip come from GIS natively, fixing the absentee blank-city gap
      that DistrictName could not (e.g. Fall City 98024, Duvall 98019).
    - The five KC ZIPs and kc_indexes_cache.pkl are no longer read. They can be
      deleted to reclaim ~1.5GB. The bulk loader functions remain in the file
      for now (unused) and will be removed in a follow-up cleanup.
    - No change to Snohomish, lead parsing, the database write path (v2.6.0),
      or the output schema.

v2.6.0 Changes:
    - NEW: the SQLite database propintel.db is now the system of record. Each
      scraped record is written through propintel_db_v0_1_0.ingest_record(),
      which dedups contacts (by email/name) and properties (by county+parcel),
      records a lead event per inbound lead, and flags ambiguous ownership
      matches (PARTIAL/MISMATCH) for review. Re-running a lead updates the
      property in place instead of appending a duplicate row.
    - The CSV-append output model is retired. After each run a spreadsheet VIEW
      is regenerated from the DB to propintel_export.xlsx (export, not store).
      The old snoco_property_data.csv/.xlsx are left untouched as historical
      files and can be deleted once you are comfortable with the DB.
    - The Excel-open lock failure is no longer fatal: the DB is always saved;
      only the optional spreadsheet view is skipped if the file is open.
    - Requires propintel_db_v0_1_0.py in the same folder. No change to any
      county lookup, parsing, or matching logic; this release only changes
      where results are written.

  - Snohomish County: live web scrape of the assessor site (unchanged).
  - King County: lookup from the King County bulk data ZIP files on disk.
    KC web scraping was removed from the routing path in v2.3.0 because it
    never held a stable connection. The bulk-data approach (formerly the
    standalone kc_lookup tool) is now built in. As of v2.5.0 the parsed
    indexes are cached to disk so subsequent runs skip the slow CSV rebuild.

Accepts leads from Market Leader CRM, tab-separated name+address lists,
or plain street addresses. Outputs to a persistent Excel/CSV spreadsheet
with photos saved to a property_photos folder.

v2.5.3 Changes:
    - FIXED: Krahl-style parcels (152407-9196) wrote blank Property Zip
      because the Residential Building.ZipCode field is empty AND the
      Address field has no trailing zip to strip. The May 20 v2.5.2 run
      surfaced this on the same parcel that originally exposed the city
      bug in v2.5.1.
    - kc_build_data() now extends the v2.5.2 owner-occupancy fallback to
      recover the zip too. When BOTH KC zip sources are blank and the
      billing street matches the property street, the zip is parsed out
      of the billing address.
    - New helper kc_parse_zip_from_billing(): pulls the trailing 5-digit
      zip from a billing string. Mirrors the city helper's gate semantics.
    - The fallback's billing-match check is now computed once and shared
      between the city and zip recovery paths.
    - No cache schema change. KC_CACHE_VERSION stays at 1. Existing
      v2.5.0/v2.5.1/v2.5.2 caches remain valid.

v2.5.2 Changes:
    - FIXED: King County rows whose Parcel.DistrictName was blank or
      "King County" still wrote a blank Property City even though the
      tax billing address clearly carried the city (Woodinville, Redmond,
      Fall City all seen in the May 20 v2.5.1 test run).
    - kc_build_data() now falls back to parsing the city out of the tax
      billing address when DistrictName is unusable. The fallback only
      fires for owner-occupied parcels (billing street matches property
      street) because for absentee owners the billing city is the OWNER'S
      mailing city, not the property's city.
    - New helper kc_parse_city_from_billing(): parses
      "STREET, CITY STATE ZIP" -> "City" (title-cased), handles multi-word
      cities like "Fall City", tolerates extra spaces.
    - No cache schema change. KC_CACHE_VERSION stays at 1. Existing
      v2.5.0/v2.5.1 caches remain valid; the fix is purely downstream of
      the cached data.

v2.5.1 Changes:
    - FIXED: King County output rows were missing the Property City, and
      sometimes Property Zip and Property State. The KC Residential Building
      file's Address field has no city, and one parcel in the May 20 test
      run (152407-9196, 32732 SE 44TH ST) had no zip in the address string
      either, which made parse_full_address fall through to its "no zip"
      branch and leave State and Zip blank too.
    - kc_build_data() now builds the property address from three KC fields:
      ResBldg.Address (street, with any trailing zip stripped), ResBldg.ZipCode
      (separate, more reliable zip), and Parcel.DistrictName (the KC field
      for the property's municipal jurisdiction). DistrictName = "King County"
      is the unincorporated marker; in that case the city is left blank
      rather than literally "King County" being written into the City column.
    - Property State is now always "WA" for KC rows when we have any
      address content at all (previously could be blank).
    - parse_full_address() is no longer called for KC rows; the parsed
      components are set directly inside kc_build_data() because the source
      data has the pieces separately and rebuilding them through a string
      parser was the cause of the missing-city bug.

v2.5.0 Changes:
    - NEW: King County indexes are cached to disk after the first load.
      First run after a fresh ZIP download takes 3-8 minutes (unchanged); the
      script then writes the parsed indexes to `kc_indexes_cache.pkl` next to
      the ZIPs. Every run after that loads the pickle in 5-10 seconds instead
      of re-parsing 2.6M rows of CSV.
    - Cache invalidates automatically when any KC ZIP's mtime is newer than
      the cache file's mtime, so re-downloading fresh KC data triggers a
      one-time rebuild and re-cache.
    - Cache pickle has an internal version stamp (KC_CACHE_VERSION). Future
      schema changes to the indexes will bump this and invalidate old caches
      automatically instead of producing wrong results.
    - Atomic write via `tmp` + `os.replace()`: a crash mid-save leaves the
      old cache intact, not a half-written file.
    - Corrupt or unreadable cache: warned and rebuilt from ZIPs, not fatal.
    - New functions: _kc_cache_path(), _kc_cache_is_fresh(),
      _kc_load_from_cache(), _kc_save_to_cache(). kc_load_all() rewired to
      check the cache before loading from ZIPs.

v2.4.2 Changes:
    - FIXED: Bothell addresses no longer route blindly to King County.
      Bothell straddles the King/Snohomish line and the previous version
      routed every Bothell address to KC. Snohomish-side Bothell parcels
      then failed the residential-only KC bulk index with no fallback. The
      v2.4.1 test run confirmed this (19825 30th Dr SE, Bothell, a real
      Snohomish parcel, reported as no-match).
    - New SPLIT_COUNTY_CITIES set holds cities that straddle the line.
      Currently just Bothell. Split-city addresses default to Snohomish in
      detect_county() and now retry on the other county if the primary
      lookup returns no data.
    - New helpers: _extract_city_segment() (shared parsing between
      detect_county and is_split_city) and is_split_city().
    - main() now loads both county data sources when any address is in a
      split-county city, so the fallback retry has something to fall back to.

v2.4.1 Changes:
    - FIXED (critical): King County bulk lookup could silently return the
      wrong property. kc_find_address()'s third matching tier did a bare
      substring test (`search_street in addr`), so a search for
      '19825 30TH DR SE' matched the KC record '19825 330TH AVE NE'
      ('30TH' is a substring of '330TH'). It then returned that lone match
      with no further checks: wrong street, wrong city, wrong zip, no
      warning. The KC equivalent of the v2.4.0 Snohomish wildcard bug.
      The third tier now matches by component (house number and street
      name required; directional required when the search has one; street
      type checked when both sides have one) and, when a zip is available
      on both sides, rejects candidates whose zip disagrees. If filtering
      leaves no single credible match the address is skipped and reported
      instead of guessed.
    - New functions: kc_address_components(), kc_components_match(),
      _kc_extract_zip(). kc_find_address() now takes pin_to_building (for
      the candidate-zip guard) and the raw search address (the only copy
      that still carries the search-side zip).

v2.4.0 Changes:
    - FIXED (critical): file-input county routing. Both the Market Leader
      path and the tab-separated path used to strip the city off the address
      before storing it, so detect_county() saw no city and defaulted every
      file-input address to Snohomish. King County addresses from a file
      never reached the King County bulk path. Lookups now store the RAW
      address (with city); scrape_property() and lookup_property_kc() each
      clean the address themselves before searching.
    - FIXED: detect_county() no longer misroutes on street names. It now
      matches the city in its actual position in the address (the segment
      before the state/zip, or the last segment) instead of a bare word
      search anywhere in the string. Previously '100 Kent Pl, Everett' would
      match 'kent' and misroute to King County.
    - FIXED: wildcard retry no longer grabs the wrong property. When an exact
      Snohomish search fails and the wildcard retry returns results, the
      result's street name must now match the input street name. If nothing
      matches, the address is skipped and reported instead of silently
      returning data for a different house. A truncated input with no street
      name (e.g. '19815') fails cleanly instead of matching.

v2.3.0 Changes:
    - King County now sources data from the five bulk data ZIP files on disk
      instead of web scraping. KC ZIPs load once at startup.
    - KC web-scrape functions (get_kc_session, search_address_kc,
      parse_kc_search_results, parse_kc_dashboard, scrape_property_kc) are
      retained in the file but no longer called. Kept in case KC server
      access ever becomes viable.
    - KC owner names are pulled from the Market Leader lead paste. The KC
      public bulk file has no taxpayer names. A bare KC address with no lead
      name produces blank owner fields and no Ownership Match.
    - KC Ownership Match is lead-based: CONFIRMED if the tax billing address
      matches the property address (owner-occupied), PARTIAL if it differs
      (absentee owner), NO LEAD DATA if no lead name was provided.
    - COLUMN_ORDER is unchanged. 15 columns the KC bulk data cannot fill are
      left blank for KC rows (Property Description, Taxpayer Name, Structure
      Description/Type, Neighborhood Code, Annual Tax Amount, Property
      Category, Status, Tax Code Area, Floor Details, Fireplace, Foundation,
      Exterior, Roof Type, Assessor Photo File).

v2.2.1 Changes:
    - Fixed King County search: added checkbox_acknowledge field for terms acceptance

v2.2 Changes:
    - Merged all features from both v2.1 branches
    - King County support with auto-detection (from v2.1a)
    - Property Grade, Condition, Views, Waterfront columns (from v2.1a)
    - Ownership Match column: CONFIRMED/PARTIAL/MISMATCH/NO LEAD DATA (from v2.1b)
    - Lead Type column: defaults to "Seller" for Market Leader leads (from v2.1b)
    - Parsed address columns: Property Street/City/State/Zip (from v2.1b)
    - IMPORTANT: Delete old snoco_property_data.xlsx and .csv before first run

Usage:
    python snoco_scraper_v2_5_1.py "13533 Boulder Ridge Rd"
    python snoco_scraper_v2_5_1.py "7214 204th Dr NE Redmond"
    python snoco_scraper_v2_5_1.py "123 Main St" "456 Oak Ave"
    python snoco_scraper_v2_5_1.py Seller_Leads.txt          (Market Leader paste)
    python snoco_scraper_v2_5_1.py addresses.txt             (Name+Address or plain addresses)

Requirements:
    pip install requests beautifulsoup4 openpyxl

Configuration:
    config.txt  -  Google Maps API key (one line, for Street View photos)

King County (v2.7.0): live, no downloads required. Address -> PIN via the
    King County ArcGIS AddressPoints service, then live assessor detail via
    Dashboard.aspx?ParcelNbr=. Just needs an internet connection.

Output:
    snoco_property_data.xlsx  -  Persistent spreadsheet (appended each run)
    snoco_property_data.csv   -  Persistent CSV backup
    property_photos/          -  Street View and Assessor photos
"""

import requests, re, sys, csv, time, os
from bs4 import BeautifulSoup
from urllib.parse import urljoin
from datetime import datetime


# ═══════════════════════════════════════════════════════════════════════════════
#  CONFIGURATION
# ═══════════════════════════════════════════════════════════════════════════════

# Snohomish County Public Access portal (v2.8.0). The old snoco.org/proptax
# ASP.NET site was retired; property data now comes from this hosted portal's
# QuickSearch JSON API (DNN/Aumentum). The search returns parcel + owner + situs.
SNOHO_BASE        = "https://wa-snohomish.publicaccessnow.com"
SNOHO_SEARCH_PAGE = SNOHO_BASE + "/PropertyInformation/PropertySearch.aspx?moduleId=470"
SNOHO_GETDATA     = SNOHO_BASE + "/DesktopModules/QuickSearch/API/Module/GetData"
SNOHO_MODULE_ID   = "470"
# v2.8.1: rich detail (assessed value, last sale, year built, sales history) comes
# from the parcel's Property Account Summary page. That page hosts several
# "DataDisplay" module tables, each fetched as JSON from this endpoint keyed by the
# table's own ModuleId. We read the table ModuleIds off the summary page and
# identify each table by its column names (robust to ModuleId changes).
SNOHO_SUMMARY_PAGE = SNOHO_BASE + "/PropertyInformation/PropertySearch/PropertyAccountSummary.aspx"
SNOHO_DD_GETDATA   = SNOHO_BASE + "/API/DataDisplay/DataSources/GetData"

OUTPUT_CSV  = "snoco_property_data.csv"
OUTPUT_XLSX = "snoco_property_data.xlsx"

# King County is live as of v2.7.0 (no bulk ZIP files). Its endpoints are
# defined next to the live lookup functions: KC_ARCGIS_ADDRPTS, KC_DASHBOARD_URL.

HTTP_HEADERS = {
    "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36",
    "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8",
}

COLUMN_ORDER = [
    # Metadata
    "Row #", "Date Retrieved", "County",
    # Lead Information
    "Lead First Name", "Lead Last Name", "Lead Email", "Lead Phone",
    "Lead Type", "Lead Source", "Years Since Sale", "Equity Level",
    # Property Address (full + parsed)
    "Property Address", "Property Street", "Property City", "Property State", "Property Zip",
    # Property Identification
    "Parcel Number", "Property Description", "Subdivision Name",
    # Owner Information + Match Status
    "Owner 1 First Name", "Owner 1 Last Name",
    "Owner 2 First Name", "Owner 2 Last Name",
    "Ownership Match",
    "Owner Address", "Taxpayer Name", "Tax Address",
    # Structure Details
    "Structure Description", "Structure Type", "Year Built",
    "Bedrooms", "Full or 3/4 Baths", "Half Baths", "Total Finished SF",
    "Size (gross)", "Unit of Measure",
    # Financial Data
    "Most Recent Sale Date", "Most Recent Sale Amount", "Neighborhood Code",
    "Latest Tax Year", "Assessed Value", "Market Total", "Market Land",
    "Market Improvement", "Taxable Value Regular", "Annual Tax Amount",
    # Additional Property Data
    "Use Code", "Property Category", "Status", "Tax Code Area",
    "Floor Details", "Garage SF", "Heat", "Fireplace", "Foundation",
    "Exterior", "Roof Type", "Property Grade", "Property Condition",
    "Views", "Waterfront", "Sales History",
    # Photos
    "Photo File", "Assessor Photo File",
]

KNOWN_CITIES = [
    "Snohomish", "Monroe", "Everett", "Marysville", "Lake Stevens",
    "Arlington", "Granite Falls", "Sultan", "Gold Bar", "Index",
    "Stanwood", "Darrington", "Edmonds", "Lynnwood", "Mountlake Terrace",
    "Mukilteo", "Mill Creek", "Bothell", "Woodinville", "Kenmore",
    "Brier", "Woodway", "Snoqualmie", "North Bend", "Carnation",
    "Fall City", "Redmond", "Kirkland", "Bellevue", "Duvall",
]

# Cities that are in King County (used to route to the KC scraper).
# Cities that straddle the line (e.g. Bothell) are NOT listed here; they are
# handled separately by SPLIT_COUNTY_CITIES below.
KING_COUNTY_CITIES = {
    "duvall", "fall city", "redmond", "kirkland", "bellevue",
    "woodinville", "kenmore", "sammamish", "issaquah",
    "renton", "kent", "auburn", "federal way", "burien", "tukwila",
    "seatac", "seattle", "mercer island", "newcastle", "black diamond",
    "maple valley", "covington", "enumclaw", "north bend", "snoqualmie",
    "carnation", "skykomish",
}

# Cities that straddle the King/Snohomish county line. detect_county() will
# default these to Snohomish (the majority of Aubie's leads in these cities sit
# on the Snohomish side, e.g. 98012/98021 north Bothell). The lookup loop will
# fall back to the other county if the primary lookup returns no match.
# v2.4.2: added to fix the Bothell routing bug where all Bothell addresses
# were being sent to King County and Snohomish-side Bothell parcels were
# failing the King County bulk lookup with no fallback.
SPLIT_COUNTY_CITIES = {"bothell"}

DIRECTION_MAP = {
    "Northwest": "NW", "Northeast": "NE", "Southwest": "SW", "Southeast": "SE",
    "North": "N", "South": "S", "East": "E", "West": "W",
}

STREET_TYPE_MAP = {
    "Drive": "Dr", "Street": "St", "Avenue": "Ave", "Boulevard": "Blvd",
    "Road": "Rd", "Lane": "Ln", "Court": "Ct", "Place": "Pl",
    "Circle": "Cir", "Terrace": "Ter", "Highway": "Hwy", "Parkway": "Pkwy",
    "Trail": "Trl",
}

SKIP_TITLE_CASE = {
    "Parcel Number", "Tax Code Area", "Use Code", "Neighborhood Code",
    "Row #", "Date Retrieved", "Photo File", "Assessor Photo File",
    "Sales History", "Floor Details", "Annual Tax Amount",
    "Assessed Value", "Market Total", "Market Land", "Market Improvement",
    "Taxable Value Regular", "Most Recent Sale Amount",
    "Latest Tax Year", "Year Built", "Size (gross)",
    "Bedrooms", "Full or 3/4 Baths", "Half Baths", "Total Finished SF",
    "Garage SF", "Ownership Match", "Lead Type", "County",
    "Property State", "Property Zip",
}

NAME_SUFFIXES = {"JR", "SR", "II", "III", "IV", "V", "VI", "VII", "VIII"}


# ═══════════════════════════════════════════════════════════════════════════════
#  TEXT FORMATTING
# ═══════════════════════════════════════════════════════════════════════════════

def smart_title_case(text):
    """Title Case preserving directional abbrevs, state codes, roman numerals."""
    if not text:
        return text
    if text.startswith("$") or re.match(r'^[\d/.$,%-]+$', text):
        return text
    if "/" in text or "\\" in text:
        return text

    preserve = {"SE", "NE", "NW", "SW", "N", "S", "E", "W",
                "WA", "OR", "CA", "ID", "MT", "PO"}
    roman = re.compile(r'^(I{1,3}|IV|VI{0,3}|IX|X{0,3})$')

    return " ".join(
        w if w in preserve or roman.match(w) else w.capitalize()
        for w in text.split()
    )


def apply_title_case(data):
    """Apply smart title case to all eligible text fields in a record."""
    for key, val in data.items():
        if key not in SKIP_TITLE_CASE and isinstance(val, str):
            data[key] = smart_title_case(val)
    return data


def clean_address_for_search(full_address):
    """Strip city/state/zip and abbreviate directions and street types
    so the assessor search can find the property."""
    addr = full_address.strip()
    if "," in addr:
        addr = addr.split(",")[0].strip()
    # Strip a TRAILING state code + optional zip. The \b (word boundary) and the
    # end anchor are critical: without them the case-insensitive state codes match
    # the start of ordinary street words ('WAgner Rd' -> '', 'CAllow Rd' -> ''),
    # which truncated those addresses to just the house number (v2.7.1 bug).
    addr = re.sub(r'\s+(WA|OR|CA|ID|MT)\b\s*\d{0,5}(-\d{4})?\s*$', '', addr, flags=re.I).strip()
    for city in sorted(KNOWN_CITIES, key=len, reverse=True):
        if addr.lower().endswith(city.lower()):
            addr = addr[:len(addr) - len(city)].strip()
            break
    for full, abbr in DIRECTION_MAP.items():
        addr = re.sub(r'\b' + full + r'\b', abbr, addr, flags=re.I)
    for full, abbr in STREET_TYPE_MAP.items():
        addr = re.sub(r'\b' + full + r'\b', abbr, addr, flags=re.I)
    return addr


def parse_full_address(full_address):
    """Parse 'Street, City, State Zip' into separate components.
    Handles both Snohomish format (with commas) and King County format (no commas).
    Returns dict with Property Street, Property City, Property State, Property Zip."""
    result = {
        "Property Street": "",
        "Property City": "",
        "Property State": "",
        "Property Zip": "",
    }
    if not full_address:
        return result
    
    # Check if comma-separated (Snohomish format): "9714 51ST AVE NE, MARYSVILLE, WA 98270"
    if "," in full_address:
        parts = [p.strip() for p in full_address.split(",")]
        
        if len(parts) >= 1:
            result["Property Street"] = parts[0]
        
        if len(parts) >= 2:
            result["Property City"] = parts[1]
        
        if len(parts) >= 3:
            # Last part should be "WA 98270" or "WA 98270-1234"
            state_zip = parts[2].strip()
            match = re.match(r'^([A-Z]{2})\s+(\d{5}(?:-\d{4})?)$', state_zip)
            if match:
                result["Property State"] = match.group(1)
                result["Property Zip"] = match.group(2)
            else:
                result["Property State"] = state_zip
    else:
        # King County format: "7214 204TH DR NE 98053" (no commas, zip at end)
        # Try to extract zip code from the end
        match = re.match(r'^(.+?)\s+(\d{5}(?:-\d{4})?)$', full_address.strip())
        if match:
            result["Property Street"] = match.group(1)
            result["Property Zip"] = match.group(2)
            result["Property State"] = "WA"  # Assume WA for King County
        else:
            # No zip found, just use the whole thing as street
            result["Property Street"] = full_address.strip()
    
    return result


# ═══════════════════════════════════════════════════════════════════════════════
#  OWNER NAME PARSING
# ═══════════════════════════════════════════════════════════════════════════════

def parse_owner_name(raw_name):
    """Parse county format LASTNAME FIRSTNAME [MIDDLE] [SUFFIX].
    Returns (first_name, last_name) with initials and suffixes stripped."""
    if not raw_name or not raw_name.strip():
        return "", ""
    parts = raw_name.strip().split()
    if len(parts) == 1:
        return parts[0], ""
    last_name = parts[0]
    cleaned = []
    for p in parts[1:]:
        u = p.upper().rstrip(".,")
        if len(u) == 1:
            continue
        if u in NAME_SUFFIXES:
            continue
        if len(u) == 2 and u.isalpha() and p.isupper() and len(parts) > 2:
            continue
        cleaned.append(p)
    return (" ".join(cleaned), last_name)


def split_owner_names(owner_string):
    """Split a two-owner string into Owner 1/2 First/Last.

    Co-owners are separated by either ' & ' or a slash. Both counties pack two
    owners with DIFFERENT surnames using a slash, e.g.
        'GUZMAN NOE BERNARDINO/BENAVIDES-ARAIZA CARINA'  (two surnames)
        'TRIPATHY SOMIA/CHIVUKULA NARESH'                (two surnames)
        'POTTER RONALD E/DAVINA L'                       (owner 2 shares surname)
    v2.7.1: '/' is now a recognized separator (it was the long-standing 'Bug 3'
    that mangled slash-format co-owners into one garbled name). When only a
    first name follows the separator, Owner 2 inherits Owner 1's last name.
    A 'C/O' (care-of) prefix is not treated as a separator."""
    result = {"Owner 1 First Name": "", "Owner 1 Last Name": "",
              "Owner 2 First Name": "", "Owner 2 Last Name": ""}
    if not owner_string:
        return result

    # Don't let a care-of marker masquerade as a co-owner slash.
    splittable = re.sub(r'\bC/O\b', 'C O', owner_string, flags=re.I)
    parts = re.split(r'\s*&\s*|\s*/\s*', splittable, maxsplit=1)
    if len(parts) == 2:
        o1f, o1l = parse_owner_name(parts[0].strip())
        o2f, o2l = parse_owner_name(parts[1].strip())
        if not o2f and o2l:
            o2f, o2l = o2l, o1l
        elif o2f and not o2l:
            o2l = o1l
    else:
        o1f, o1l = parse_owner_name(owner_string)
        o2f, o2l = "", ""

    result["Owner 1 First Name"] = o1f
    result["Owner 1 Last Name"]  = o1l
    result["Owner 2 First Name"] = o2f
    result["Owner 2 Last Name"]  = o2l
    return result


def fix_owner_names_with_lead(data, lead_data):
    """Use lead name to detect/fix swapped owner name order."""
    if not lead_data:
        return data
    lf = lead_data.get("Lead First Name", "").upper().strip()
    ll = lead_data.get("Lead Last Name", "").upper().strip()
    o1f = data.get("Owner 1 First Name", "").upper().strip()
    o1l = data.get("Owner 1 Last Name", "").upper().strip()

    if (lf and o1l and lf == o1l) or (ll and o1f and ll == o1f):
        data["Owner 1 First Name"], data["Owner 1 Last Name"] = \
            data["Owner 1 Last Name"], data["Owner 1 First Name"]
        if data.get("Owner 2 Last Name", "").upper() == o1l:
            data["Owner 2 Last Name"] = data["Owner 1 Last Name"]
    return data


def compute_ownership_match(data, lead_data):
    """Compare lead name to owner names and return match status.
    
    Returns:
        CONFIRMED - Lead last name matches Owner 1 or Owner 2 last name, 
                    AND first names also match (or are close enough)
        PARTIAL - Last names match but first names don't (could be spouse, etc.)
        MISMATCH - Last names don't match (property may have sold)
        NO LEAD DATA - No lead name was provided
    """
    if not lead_data:
        return "NO LEAD DATA"
    
    lead_first = lead_data.get("Lead First Name", "").upper().strip()
    lead_last = lead_data.get("Lead Last Name", "").upper().strip()
    
    if not lead_last:
        return "NO LEAD DATA"
    
    o1_first = data.get("Owner 1 First Name", "").upper().strip()
    o1_last = data.get("Owner 1 Last Name", "").upper().strip()
    o2_first = data.get("Owner 2 First Name", "").upper().strip()
    o2_last = data.get("Owner 2 Last Name", "").upper().strip()
    
    # Check if lead last name matches either owner's last name
    last_name_match = (lead_last == o1_last) or (lead_last == o2_last)
    
    if not last_name_match:
        return "MISMATCH"
    
    # Last name matched. Now check first names.
    # Match if lead first name matches owner 1 OR owner 2 first name
    # Also match if lead first is a prefix of owner first (Ed matches Edward)
    # or owner first is a prefix of lead first
    def first_name_match(lead_f, owner_f):
        if not lead_f or not owner_f:
            return False
        if lead_f == owner_f:
            return True
        # Check if one is prefix of the other (handles Ed/Edward, Mike/Michael, etc.)
        if lead_f.startswith(owner_f) or owner_f.startswith(lead_f):
            return True
        return False
    
    if first_name_match(lead_first, o1_first) or first_name_match(lead_first, o2_first):
        return "CONFIRMED"
    
    # Last name matched but first name didn't
    return "PARTIAL"


# ═══════════════════════════════════════════════════════════════════════════════
#  LEAD FILE PARSING
# ═══════════════════════════════════════════════════════════════════════════════

MARKET_LEADER_KEYWORDS = {"Source:", "Since Sale", "Address Insights", "Assigned To:"}

ML_SKIP_WORDS = {
    "Edit", "mode_edit", "Seller", "Lead", "Source:", "email", "phone",
    "note", "Add Note", "Action Center", "Send Listings", "Nurture Zone",
    "Website Activity", "favorite", "Favorites", "Views", "trending_up",
    "Address Insights", "0", "$-",
}


def parse_lead_file(filepath):
    """Parse one or more Market Leader lead pastes from a text file."""
    with open(filepath, "r", encoding="utf-8") as f:
        content = f.read()

    if re.search(r'\n\s*-{3,}\s*\n', content):
        blocks = re.split(r'\n\s*-{3,}\s*\n', content)
    else:
        lines = content.split("\n")
        blocks, current = [], []
        for i, raw in enumerate(lines):
            line = raw.strip()
            is_new = (
                current and line and line not in ML_SKIP_WORDS
                and i + 2 < len(lines)
                and lines[i+1].strip() in ("Edit", "Editmode_edit")
                and lines[i+2].strip() in ("mode_edit", "Seller")
            )
            if is_new:
                blocks.append("\n".join(current))
                current = [raw]
            else:
                current.append(raw)
        if current:
            blocks.append("\n".join(current))

    leads = []
    for b in blocks:
        b = b.strip()
        if b:
            lead = _parse_single_ml_lead(b)
            if lead.get("address"):
                leads.append(lead)
    return leads


def _parse_single_ml_lead(text):
    """Parse a single Market Leader lead block."""
    lead = dict(name="", first_name="", last_name="", email="", phone="",
                source="", address="", years_since_sale="", equity_level="",
                lead_type="Seller")  # Market Leader leads are sellers
    lines = [l.strip() for l in text.strip().split("\n") if l.strip()]
    if not lines:
        return lead

    name = re.sub(r'\s*(Edit)?mode_edit.*$', '', lines[0]).strip()
    name = re.sub(r'\s*Edit\s*$', '', name).strip()
    lead["name"] = name
    parts = name.split()
    lead["first_name"] = parts[0] if parts else ""
    lead["last_name"] = " ".join(parts[1:]) if len(parts) > 1 else ""

    for i, line in enumerate(lines):
        nxt = lines[i+1] if i+1 < len(lines) else ""

        if line == "Source:":
            lead["source"] = nxt
        if line.lower() == "email":
            m = re.search(r'[\w.\-+]+@[\w.\-]+\.\w+', nxt)
            if m: lead["email"] = m.group(0)
        elif "@" in line and not lead["email"]:
            m = re.search(r'[\w.\-+]+@[\w.\-]+\.\w+', line)
            if m: lead["email"] = m.group(0)
        if line.lower() == "phone":
            m = re.search(r'(\d{3}[-.\s]?\d{3}[-.\s]?\d{4})', nxt)
            if m: lead["phone"] = m.group(1)
        elif re.match(r'^\d{3}[-.\s]\d{3}[-.\s]\d{4}$', line) and not lead["phone"]:
            lead["phone"] = line
        if "years since sale" in line.lower():
            lead["years_since_sale"] = line
        if "equity" in line.lower():
            lead["equity_level"] = line
        if line == "Address Insights":
            street = nxt
            city = lines[i+2] if i+2 < len(lines) else ""
            if re.search(r'(WA|OR|CA|ID|MT)\s*\d{0,5}', city, re.I):
                lead["address"] = f"{street} {city}"
            else:
                lead["address"] = street

    if not lead["address"]:
        for i, line in enumerate(lines):
            if re.match(r'^\d+\s+\w', line):
                skip = ["source", "email", "phone", "note", "assigned", "hours ago"]
                if not any(kw in line.lower() for kw in skip):
                    nxt = lines[i+1] if i+1 < len(lines) else ""
                    if re.search(r'(WA|OR|CA|ID|MT)', nxt, re.I):
                        lead["address"] = f"{line} {nxt}"
                    else:
                        lead["address"] = line
                    break
    return lead


def parse_tab_separated(content):
    """Parse 'Name\\tAddress' lines into lookup tuples."""
    lookups = []
    for line in content.split("\n"):
        line = line.strip()
        if not line or "\t" not in line:
            continue
        parts = line.split("\t", 1)
        if len(parts) != 2:
            continue
        name, addr = parts[0].strip(), parts[1].strip()
        np = name.split()
        first = np[0] if np else ""
        last = " ".join(np[1:]) if len(np) > 1 else ""
        lead_data = {
            "Lead First Name": smart_title_case(first),
            "Lead Last Name": smart_title_case(last),
            "Lead Email": "", "Lead Phone": "", "Lead Source": "",
            "Lead Type": "",  # Unknown for tab-separated input
            "Years Since Sale": "", "Equity Level": "",
        }
        # Store the RAW address (with city). detect_county() must see the city
        # to route correctly. The county-specific lookup functions
        # (scrape_property / lookup_property_kc) clean it themselves.
        lookups.append((addr, lead_data))
        print(f"  Lead: {first} {last} | {addr}")
    return lookups


def build_lead_data(lead):
    """Convert a parsed lead dict into the lead_data dict for merging."""
    return {
        "Lead First Name": smart_title_case(lead.get("first_name", "")),
        "Lead Last Name":  smart_title_case(lead.get("last_name", "")),
        "Lead Email":      lead.get("email", ""),
        "Lead Phone":      lead.get("phone", ""),
        "Lead Source":     lead.get("source", ""),
        "Lead Type":       lead.get("lead_type", ""),
        "Years Since Sale": lead.get("years_since_sale", ""),
        "Equity Level":    lead.get("equity_level", ""),
    }


# ═══════════════════════════════════════════════════════════════════════════════
#  ASSESSOR WEBSITE INTERACTION (Snohomish County - Public Access portal, v2.8.0)
# ═══════════════════════════════════════════════════════════════════════════════
#  The retired snoco.org/proptax WebForms site is gone. Property data now comes
#  from the hosted Public Access portal's QuickSearch JSON API. get_session()
#  primes the search page to capture the DNN ServicesFramework auth context
#  (TabId + anti-forgery token) that the API requires, and scrape_property()
#  calls the API. Both keep their old signatures so main(), reenrich, and
#  lead_intake call them unchanged. The search result carries parcel + owner +
#  situs (enough for ownership verification); rich assessor detail (value, beds,
#  baths, year) is not exposed by the public search API and is left blank.
# ═══════════════════════════════════════════════════════════════════════════════

def get_session():
    """Prime a session against the Snohomish Public Access portal and capture the
    auth context the search API needs. Returns (session, ctx) where ctx is the
    header dict passed as `base_resp` to scrape_property(). Raises on an
    unreachable portal or missing tokens (callers already wrap this in try/except)."""
    session = requests.Session()
    session.headers.update(HTTP_HEADERS)
    resp = session.get(SNOHO_SEARCH_PAGE, timeout=20)
    resp.raise_for_status()
    html = resp.text
    rvt = re.search(r'name="__RequestVerificationToken"[^>]*value="([^"]+)"', html)
    tab = re.search(r"sf_tabId[^0-9]{0,4}(\d+)", html)
    if not rvt or not tab:
        raise RuntimeError("Snohomish portal: could not read auth tokens "
                           "(site layout changed?)")
    ctx = {"ModuleId": SNOHO_MODULE_ID, "TabId": tab.group(1),
           "RequestVerificationToken": rvt.group(1)}
    return session, ctx


# (The new scrape_property() that calls the portal API lives further down, next
#  to the King County live lookup and the shared photo/owner helpers it reuses.)


# ═══════════════════════════════════════════════════════════════════════════════
#  COUNTY DETECTION
# ═══════════════════════════════════════════════════════════════════════════════

def _extract_city_segment(address):
    """Return the lowercased city segment from a raw address, or '' if none.

    Shared by detect_county() and is_split_city() so both apply the same
    parsing rules to comma-separated and no-comma addresses.
    """
    if not address:
        return ""

    addr = address.strip()

    if "," in addr:
        segments = [s.strip() for s in addr.split(",") if s.strip()]
        city_segment = segments[-1] if segments else ""
        for idx, seg in enumerate(segments):
            if re.match(r'^[A-Z]{2}(\s+\d{5}(-\d{4})?)?$', seg, flags=re.I) \
                    or re.match(r'^\d{5}(-\d{4})?$', seg):
                if idx > 0:
                    city_segment = segments[idx - 1]
                break
    else:
        city_segment = addr

    # Strip a trailing 'ST zip' / 'zip' from the chosen segment. This is the fix
    # for 'street, City ST zip' (city+state+zip in one comma segment), where the
    # loop above can't isolate the city - e.g. 'Duvall WA 98019' -> 'Duvall',
    # 'Snoqualmie WA 98065' -> 'Snoqualmie'. Then keep the last 1-2 words as the
    # city so two-word names ('Lake Stevens', 'Fall City') survive.
    city_segment = re.sub(r'\s+[A-Z]{2}\b\s*\d{0,5}(-\d{4})?\s*$', '', city_segment, flags=re.I)
    city_segment = re.sub(r'\s+\d{5}(-\d{4})?\s*$', '', city_segment).strip()
    words = city_segment.split()
    if len(words) > 2:
        city_segment = " ".join(words[-2:])

    return city_segment.lower().strip()


def is_split_city(address):
    """Return True if the address city is in SPLIT_COUNTY_CITIES.

    Split cities straddle the King/Snohomish line, so a no-match result on
    the primary county warrants a retry on the other county.
    """
    city_lower = _extract_city_segment(address)
    for city in SPLIT_COUNTY_CITIES:
        if city_lower == city:
            return True
        if re.search(r'\b' + re.escape(city) + r'$', city_lower):
            return True
    return False


def detect_county(address):
    """Return 'king' if the address city matches a King County city, else 'snohomish'.

    As of v2.4.0 this receives the RAW address (with city/state/zip), so the
    city can be matched in its actual position instead of by a bare word
    search. The old bare-\\b search misrouted addresses whose STREET name
    happened to contain a King County city word (e.g. '100 Kent Pl, Everett').

    As of v2.4.2, cities that straddle the county line (SPLIT_COUNTY_CITIES,
    currently just Bothell) are NOT in KING_COUNTY_CITIES and so default to
    Snohomish here. Callers should use is_split_city() to decide whether to
    retry the other county on a no-match.
    """
    if not address:
        return "snohomish"

    city_lower = _extract_city_segment(address)

    # Exact match on the city segment, then a relaxed "ends with city" check
    # for the no-comma fallback (handles a one-word-city tail).
    for city in KING_COUNTY_CITIES:
        if city_lower == city:
            return "king"
    for city in KING_COUNTY_CITIES:
        if re.search(r'\b' + re.escape(city) + r'$', city_lower):
            return "king"
    return "snohomish"


# ═══════════════════════════════════════════════════════════════════════════════
#  KING COUNTY LIVE  (v2.7.0 - ArcGIS address->PIN + direct Dashboard GET)
# ═══════════════════════════════════════════════════════════════════════════════
#  King County data now comes from two public HTTP GETs - no session, no
#  viewstate, no bulk ZIP files:
#    1. Resolve address -> PIN (+ city + zip) via the King County ArcGIS
#       AddressPoints feature layer. A deterministic SQL-style ADDR_FULL query,
#       not the fussy geocoder locator.
#    2. Fetch live assessor detail via Dashboard.aspx?ParcelNbr=<PIN> and parse
#       it with the existing parse_kc_dashboard().
#
#  This REPLACES the bulk-data path (v2.3.0-v2.6.0), retired in v2.7.0. Live
#  data is current, carries real owner names (the public bulk file strips
#  them), and supplies city/zip natively (fixing the absentee blank-city gap
#  that DistrictName could not). See the spike findings for the GO decision.
# ═══════════════════════════════════════════════════════════════════════════════

KC_ARCGIS_ADDRPTS = ("https://gismaps.kingcounty.gov/arcgis/rest/services/"
                     "Address/KingCo_AddressPoints/MapServer/0/query")
KC_DASHBOARD_URL = ("https://blue.kingcounty.com/Assessor/eRealProperty/"
                    "Dashboard.aspx?ParcelNbr=")
# GIS land-use, used to confirm an owner-less master parcel is a condominium
# (layer 2 = Parcels, field PREUSE_DESC) before walking its units.
KC_PROPINFO_URL = ("https://gismaps.kingcounty.gov/arcgis/rest/services/"
                   "Property/KingCo_PropertyInfo/MapServer/2/query")


# Direction words and street-type abbreviations dropped when building the
# relaxed King County address fallback (so the match keys on the core name).
_KC_DIRECTIONS = {"N", "S", "E", "W", "NE", "NW", "SE", "SW"}
_KC_STREET_TYPES = {"ST", "AVE", "DR", "RD", "LN", "CT", "PL", "BLVD", "CIR",
                    "TER", "PKWY", "HWY", "WY", "WAY", "CV", "LOOP", "RUN",
                    "PT", "TRL", "XING", "SQ"}


def _kc_relaxed_where(norm):
    """Build a relaxed ArcGIS WHERE keyed on house number + the core street-name
    words (directionals and street types removed). Used only as a fallback for
    streets whose name carries a direction word that normalization abbreviates
    (e.g. 'East Ames Lake Dr'). Returns None when there isn't enough to be safe
    (need a leading house number plus at least one name word)."""
    parts = norm.split()
    if len(parts) < 2 or not parts[0].isdigit():
        return None
    hn = parts[0].replace("'", "''")
    core = [w for w in parts[1:]
            if w not in _KC_DIRECTIONS and w not in _KC_STREET_TYPES]
    if not core:
        return None
    likes = " AND ".join("ADDR_FULL LIKE '%{}%'".format(w.replace("'", "''"))
                         for w in core)
    return "ADDR_FULL LIKE '{} %' AND {}".format(hn, likes)


def kc_resolve_address_live(address, timeout=20):
    """Resolve a street address to a King County parcel via the public ArcGIS
    AddressPoints layer. Returns {pin, addr_full, city, zip} or None.

    Tries an exact ADDR_FULL match first, then a prefix LIKE. The layer stores
    addresses uppercased and abbreviated (e.g. '32732 SE 44TH ST') - exactly
    what kc_normalize_address() produces, so the two line up. POSTALCTYNAME is
    preferred for the city because it carries the municipal name even for
    unincorporated parcels (e.g. Fall City) where the assessor's DistrictName
    is blank.

    v2.7.4: when both anchored matches miss, a relaxed fallback keyed on house
    number + core street-name words recovers direction-in-name streets (where
    normalization abbreviated a name word, e.g. EAST->E). The fallback is
    accepted only when it returns exactly one parcel, so it never silently grabs
    a wrong parcel.

    The incoming address carries the city/state/zip (it's the RAW lead
    address, needed for county routing). clean_address_for_search() strips that
    down to the street so ADDR_FULL can match, the same prep the Snohomish path
    does before its search."""
    norm = kc_normalize_address(clean_address_for_search(address))
    if not norm:
        return None
    safe = norm.replace("'", "''")
    fields = "PIN,ADDR_FULL,ZIP5,CTYNAME,POSTALCTYNAME"

    def _query(where, record_count=5):
        try:
            resp = requests.get(KC_ARCGIS_ADDRPTS, params={
                "where": where, "outFields": fields,
                "returnGeometry": "false", "f": "json",
                "resultRecordCount": record_count,
            }, headers=HTTP_HEADERS, timeout=timeout)
            resp.raise_for_status()
            return resp.json().get("features", [])
        except Exception as e:
            print(f"  [WARNING] KC ArcGIS query failed: {e}")
            return None

    def _result(feat):
        a = feat.get("attributes", {})
        pin = str(a.get("PIN", "") or "").strip()
        if not pin:
            return None
        city = (a.get("POSTALCTYNAME") or a.get("CTYNAME") or "").strip()
        return {
            "pin": pin,
            "addr_full": (a.get("ADDR_FULL") or "").strip(),
            "city": smart_title_case(city) if city else "",
            "zip": str(a.get("ZIP5", "") or "").strip(),
        }

    # 1) exact, then 2) prefix LIKE - anchored, so the first hit is trustworthy.
    for where in (f"ADDR_FULL = '{safe}'", f"ADDR_FULL LIKE '{safe}%'"):
        feats = _query(where)
        if feats is None:
            return None
        if feats:
            r = _result(feats[0])
            if r:
                return r

    # 3) relaxed fallback - accept only an unambiguous (single) match.
    relaxed = _kc_relaxed_where(norm)
    if relaxed:
        feats = _query(relaxed, record_count=3)
        if feats and len(feats) == 1:
            r = _result(feats[0])
            if r:
                print(f"  -> KC matched via relaxed address fallback: {r['addr_full']}")
                return r
    return None


def kc_fetch_dashboard_live(pin, timeout=20):
    """GET and parse the live KC eReal Property Dashboard for a 10-digit PIN.
    Returns (data_dict, session) so the caller can reuse the session for the
    assessor photo (the MediaHandler image needs the same cookies), or
    (None, None) on failure."""
    session = requests.Session()
    session.headers.update(HTTP_HEADERS)
    try:
        r = session.get(KC_DASHBOARD_URL + pin, timeout=timeout)
        r.raise_for_status()
    except Exception as e:
        print(f"  [ERROR] KC Dashboard fetch failed for PIN {pin}: {e}")
        return None, None
    data = parse_kc_dashboard(r.text, r.url)
    if not data.get("Parcel Number"):
        return None, None
    return data, session


def kc_is_condo(pin, timeout=15):
    """True if the parcel's King County GIS land use is a condominium. Used to
    confirm an owner-less master parcel really is a condo complex before spending
    requests walking its units (vs. e.g. a vacant or exempt parcel)."""
    try:
        resp = requests.get(KC_PROPINFO_URL, params={
            "where": f"PIN='{pin}'", "outFields": "PREUSE_DESC",
            "returnGeometry": "false", "f": "json"},
            headers=HTTP_HEADERS, timeout=timeout)
        feats = resp.json().get("features", [])
    except Exception:
        return False
    if not feats:
        return False
    desc = (feats[0].get("attributes", {}).get("PREUSE_DESC") or "").upper()
    return "CONDO" in desc


def _kc_unit_owner(pin, session, timeout=12):
    """Fetch just the owner-name split for one candidate condo-unit PIN. Returns a
    split_owner_names() dict, or None when the PIN has no assessor record or no
    name (i.e. a gap past the last real unit)."""
    try:
        r = session.get(KC_DASHBOARD_URL + pin, timeout=timeout)
        r.raise_for_status()
    except Exception:
        return None
    soup = BeautifulSoup(r.text, "html.parser")
    header = soup.find("table", id="cphContent_DetailsViewDashboardHeader")
    if not header:
        return None
    name = ""
    for row in header.find_all("tr"):
        cells = row.find_all("td")
        if len(cells) == 2 and cells[0].text.strip() == "Name":
            name = cells[1].text.strip()
            break
    return split_owner_names(name) if name else None


def kc_resolve_condo_unit(master_pin, lead_data, session=None):
    """A condo/townhome street address resolves (via the KC address layer) to the
    complex MASTER parcel, MINOR 0000 - common area with no owner or value. The
    real owner + value live in per-unit parcels MAJOR-00X0. Walk the units and
    return the PIN of the one owned by THIS lead, matched by surname and
    disambiguated by first name when a surname repeats in the complex.

    Returns None - leaving the lead UNENRICHED rather than guessing - when there
    is no lead name to match on, the parcel is not a condo, or the unit match is
    absent or ambiguous. This upholds the project rule against silently attaching
    a wrong property."""
    if not lead_data:
        return None
    ll = (lead_data.get("Lead Last Name") or "").upper().strip()
    lf = (lead_data.get("Lead First Name") or "").upper().strip()
    if not ll or not kc_is_condo(master_pin):
        return None

    sess = session
    if sess is None:
        sess = requests.Session()
        sess.headers.update(HTTP_HEADERS)
    major = master_pin[:6]

    # Condo unit minors are assigned in multiples of 10 (0010, 0020, ...). Walk
    # them and stop after a run of empties once the unit block has been seen.
    units = []  # (pin, owner_dict)
    empty_run = 0
    for m in range(10, 2001, 10):
        owner = _kc_unit_owner(f"{major}{m:04d}", sess)
        if owner:
            units.append((f"{major}{m:04d}", owner))
            empty_run = 0
        elif units:
            empty_run += 1
            if empty_run >= 12:
                break
    # Fallback for the rarer complexes that number units sequentially (0001, 0002).
    if not units:
        for m in range(1, 61):
            owner = _kc_unit_owner(f"{major}{m:04d}", sess)
            if owner:
                units.append((f"{major}{m:04d}", owner))

    def lastnames(o):
        return {(o.get("Owner 1 Last Name") or "").upper().strip(),
                (o.get("Owner 2 Last Name") or "").upper().strip()}

    def firstnames(o):
        return {(o.get("Owner 1 First Name") or "").upper().strip(),
                (o.get("Owner 2 First Name") or "").upper().strip()}

    matches = [(pin, o) for pin, o in units if ll in lastnames(o)]
    if len(matches) > 1 and lf:
        narrowed = [(pin, o) for pin, o in matches if lf in firstnames(o)]
        if narrowed:
            matches = narrowed
    if len(matches) != 1:
        if matches:
            print(f"  [WARNING] Condo unit match for '{lf} {ll}' was ambiguous "
                  f"({len(matches)} candidates); left unenriched.")
        return None
    return matches[0][0]


def lookup_property_kc_live(address, lead_data=None, api_key=None,
                            photo_folder="property_photos"):
    """Live King County lookup: ArcGIS address->PIN, then a Dashboard GET.
    Returns a data dict in the master schema, or None on failure.

    Replaces the bulk lookup_property_kc(). Owner names now come from the live
    assessor record, so KC rows go through the same full ownership-match logic
    as Snohomish (handled by the caller in main())."""
    print(f"\n{'='*60}")
    print(f"  Looking up (King County, live): {address}")
    print(f"{'='*60}")

    resolved = kc_resolve_address_live(address)
    if not resolved:
        print(f"  [ERROR] No King County GIS match for: {address}")
        return None
    print(f"  -> Resolved: PIN {resolved['pin']} | {resolved['addr_full']} | "
          f"{resolved['city']} {resolved['zip']}")

    data, session = kc_fetch_dashboard_live(resolved["pin"])
    if not data:
        print(f"  [ERROR] No live assessor record for PIN {resolved['pin']}")
        return None

    # Condo / townhome complex: the address resolved to the MASTER parcel (common
    # area, MINOR 0000), which has no owner and no value. Find the lead's actual
    # unit parcel. This is a no-op for ordinary parcels, which already carry an
    # owner here, so the extra GIS/unit requests only fire for condo-shaped misses.
    if not (data.get("Owner 1 Last Name") or "").strip():
        unit_pin = kc_resolve_condo_unit(resolved["pin"], lead_data, session)
        if unit_pin:
            unit_data, unit_session = kc_fetch_dashboard_live(unit_pin)
            if unit_data and (unit_data.get("Owner 1 Last Name") or "").strip():
                print(f"  -> Condo master {resolved['pin']} is common area; "
                      f"matched lead to unit {unit_pin}")
                # Unit pages carry owner + value but usually no street address;
                # keep the complex street from the GIS resolve for display.
                if not (unit_data.get("Property Address") or "").strip():
                    unit_data["Property Address"] = resolved["addr_full"]
                    unit_data["Property Street"] = resolved["addr_full"]
                data, session = unit_data, unit_session

    # GIS supplies a more reliable city/zip than the Dashboard's Site Address
    # (which has no city). Rebuild the canonical address from the GIS pieces.
    street = data.get("Property Street") or data.get("Property Address", "")
    street = re.sub(r'\s+\d{5}(?:-\d{4})?\s*$', '', street).strip()
    city, zc = resolved["city"], resolved["zip"]
    data["Property Street"] = street
    data["Property City"]   = city
    data["Property State"]  = "WA"
    data["Property Zip"]    = zc
    if city:
        data["Property Address"] = f"{street}, {city} WA {zc}".strip()
    elif zc:
        data["Property Address"] = f"{street}, WA {zc}".strip()
    else:
        data["Property Address"] = street

    print(f"  -> Parcel: {data.get('Parcel Number', 'N/A')}")
    print(f"  -> Owner:  {data.get('Owner 1 First Name','')} {data.get('Owner 1 Last Name','')}")
    print(f"  -> Year Built: {data.get('Year Built','N/A')} | "
          f"Beds: {data.get('Bedrooms','N/A')} | "
          f"Baths: {data.get('Full or 3/4 Baths','N/A')} | "
          f"SF: {data.get('Total Finished SF','N/A')}")
    print(f"  -> Assessed: {data.get('Assessed Value','N/A')} "
          f"({data.get('Latest Tax Year','N/A')})")

    # Assessor photo (reuse the dashboard session for cookies). Non-fatal.
    kc_photo_url = data.pop("_kc_photo_url", "")
    if kc_photo_url:
        ap = download_assessor_photo(kc_photo_url, data.get("Property Address", ""),
                                     data.get("Owner 1 Last Name", ""),
                                     photo_folder, session=session)
        data["Assessor Photo File"] = ap
        if ap:
            print(f"  -> Assessor photo saved: {ap}")

    # Street View photo (Google API, same as Snohomish).
    if api_key and data.get("Property Address"):
        sv = download_street_view(data["Property Address"],
                                  data.get("Owner 1 Last Name", ""),
                                  api_key, photo_folder)
        data["Photo File"] = sv
        if sv:
            print(f"  -> Street View photo saved: {sv}")
    else:
        data["Photo File"] = ""

    data = apply_title_case(data)
    data["County"] = "King"
    return data


# ═══════════════════════════════════════════════════════════════════════════════
#  KING COUNTY SHARED HELPER
# ═══════════════════════════════════════════════════════════════════════════════

def kc_normalize_address(addr):
    """Normalize an address to King County's ADDR_FULL form for the live ArcGIS
    lookup. Returns uppercase, single-spaced, with directions and street types
    abbreviated and any trailing zip code removed."""
    if not addr:
        return ""

    # Uppercase and compress whitespace
    addr = " ".join(addr.upper().split())

    # Remove zip code from the end if present
    addr = re.sub(r'\s+\d{5}(-\d{4})?$', '', addr)

    # Standardize directional prefixes/suffixes
    addr = re.sub(r'\bNORTHWEST\b', 'NW', addr)
    addr = re.sub(r'\bNORTHEAST\b', 'NE', addr)
    addr = re.sub(r'\bSOUTHWEST\b', 'SW', addr)
    addr = re.sub(r'\bSOUTHEAST\b', 'SE', addr)
    addr = re.sub(r'\bNORTH\b', 'N', addr)
    addr = re.sub(r'\bSOUTH\b', 'S', addr)
    addr = re.sub(r'\bEAST\b', 'E', addr)
    addr = re.sub(r'\bWEST\b', 'W', addr)

    # Standardize street types
    addr = re.sub(r'\bSTREET\b', 'ST', addr)
    addr = re.sub(r'\bAVENUE\b', 'AVE', addr)
    addr = re.sub(r'\bDRIVE\b', 'DR', addr)
    addr = re.sub(r'\bROAD\b', 'RD', addr)
    addr = re.sub(r'\bLANE\b', 'LN', addr)
    addr = re.sub(r'\bCOURT\b', 'CT', addr)
    addr = re.sub(r'\bPLACE\b', 'PL', addr)
    addr = re.sub(r'\bBOULEVARD\b', 'BLVD', addr)
    addr = re.sub(r'\bCIRCLE\b', 'CIR', addr)
    addr = re.sub(r'\bTERRACE\b', 'TER', addr)
    addr = re.sub(r'\bPARKWAY\b', 'PKWY', addr)
    addr = re.sub(r'\bHIGHWAY\b', 'HWY', addr)
    addr = re.sub(r'\bWAY\b', 'WY', addr)

    return addr


# ═══════════════════════════════════════════════════════════════════════════════
#  KING COUNTY DASHBOARD PARSER  (used by the live lookup above)
# ═══════════════════════════════════════════════════════════════════════════════
#  Parses one eReal Property Dashboard page into the master schema. The live
#  path (kc_fetch_dashboard_live) GETs the page by PIN and hands it here.
# ═══════════════════════════════════════════════════════════════════════════════


def parse_kc_dashboard(html, url):
    """Parse the King County eReal Property Dashboard page.
    Returns a data dict with all available fields."""
    soup = BeautifulSoup(html, "html.parser")
    data = {"County": "King"}

    # --- Parcel header block (GridViewStyle table with Parcel Number / Name / Site Address / Legal) ---
    header_table = soup.find("table", id="cphContent_DetailsViewDashboardHeader")
    if header_table:
        for row in header_table.find_all("tr"):
            cells = row.find_all("td")
            if len(cells) == 2:
                lbl, val = cells[0].text.strip(), cells[1].text.strip()
                if lbl == "Parcel Number":
                    # KC format: "295440-0020" - store without hyphen for consistency
                    data["Parcel Number"] = val.replace("-", "")
                elif lbl == "Name":
                    data.update(split_owner_names(val.strip()))
                elif lbl == "Site Address":
                    # KC gives "7214 204TH DR NE 98053" - no city. We'll clean it up.
                    data["Property Address"] = val.strip()
                elif lbl == "Legal":
                    data["Property Description"] = val.strip()
                    # Subdivision name is the legal description directly for KC
                    data["Subdivision Name"] = val.strip()

    # --- Parse full address into components ---
    if data.get("Property Address"):
        data.update(parse_full_address(data["Property Address"]))

    # --- Building characteristics ---
    building_table = soup.find("table", id="cphContent_DetailsViewPropTypeR")
    if building_table:
        for row in building_table.find_all("tr"):
            cells = row.find_all("td")
            if len(cells) == 2:
                lbl, val = cells[0].text.strip(), cells[1].text.strip()
                if lbl == "Year Built":
                    data["Year Built"] = val
                elif lbl == "Total Square Footage":
                    data["Total Finished SF"] = val
                elif lbl == "Number Of Bedrooms":
                    data["Bedrooms"] = val
                elif lbl == "Number Of Baths":
                    # KC gives decimal baths e.g. 3.75
                    # Store in Full or 3/4 Baths as-is, leave Half Baths blank
                    data["Full or 3/4 Baths"] = val
                elif lbl == "Grade":
                    data["Property Grade"] = val
                elif lbl == "Condition":
                    data["Property Condition"] = val
                elif lbl == "Lot Size":
                    data["Size (gross)"] = val
                    data["Unit of Measure"] = "Sq Ft"
                elif lbl == "Views":
                    data["Views"] = val
                elif lbl == "Waterfront":
                    data["Waterfront"] = val.strip() if val.strip() else "No"

    # --- Tax roll history - grab most recent row ---
    tax_table = soup.find("table", id="cphContent_GridViewDBTaxRoll")
    if tax_table:
        rows = tax_table.find_all("tr")
        # First data row (index 1) is the most recent
        if len(rows) >= 2:
            cells = rows[1].find_all("td")
            if len(cells) >= 9:
                data["Latest Tax Year"]      = cells[1].text.strip()
                data["Market Land"]          = cells[2].text.strip()
                data["Market Improvement"]   = cells[3].text.strip()
                data["Market Total"]         = cells[4].text.strip()
                data["Taxable Value Regular"] = cells[8].text.strip()
                # KC doesn't separate Assessed Value from Market Total at this level
                data["Assessed Value"]       = cells[8].text.strip()

    # --- Levy code (Tax Code Area) ---
    levy_form = soup.find("table", id="cphContent_FormViewLevyDist")
    if levy_form:
        levy_label = levy_form.find("span", id="cphContent_FormViewLevyDist_Label1")
        if levy_label:
            data["Tax Code Area"] = levy_label.text.strip()

    # --- Assessor photo ---
    photo_img = soup.find("img", id="cphContent_FormViewPictCurr_CurrentImage")
    if photo_img and photo_img.get("src"):
        src = photo_img["src"]
        # Build absolute URL
        data["_kc_photo_url"] = src if src.startswith("http") else urljoin(url, src)

    return data


# ═══════════════════════════════════════════════════════════════════════════════
#  PHOTO DOWNLOADS
# ═══════════════════════════════════════════════════════════════════════════════

def _photo_filename(address, owner_last, suffix):
    """Build '9714 51st Ave NE (Potter) - Street View.jpg'."""
    street = smart_title_case(address.split(",")[0].strip())
    street = re.sub(r'[<>:"/\\|?*]', '', street).strip()
    if owner_last:
        owner = re.sub(r'[<>:"/\\|?*]', '', smart_title_case(owner_last)).strip()
        return f"{street} ({owner}) - {suffix}.jpg"
    return f"{street} - {suffix}.jpg"


def download_street_view(address, owner_last, api_key, folder):
    """Download Google Street View photo. Returns filepath or ''."""
    if not api_key:
        return ""
    os.makedirs(folder, exist_ok=True)
    path = os.path.join(folder, _photo_filename(address, owner_last, "Street View"))
    try:
        resp = requests.get("https://maps.googleapis.com/maps/api/streetview",
                            params={"size": "2048x2048", "location": address,
                                    "fov": "60", "pitch": "5", "key": api_key},
                            timeout=15)
        if resp.status_code == 200 and "image" in resp.headers.get("Content-Type", ""):
            with open(path, "wb") as f:
                f.write(resp.content)
            return path
    except Exception as e:
        print(f"  [WARNING] Street View download failed: {e}")
    return ""


def download_assessor_photo(photo_url, address, owner_last, folder, session=None):
    """Download the assessor's property photo. Returns filepath or ''.
    Pass session for King County (MediaHandler requires same session cookies)."""
    if not photo_url:
        return ""
    os.makedirs(folder, exist_ok=True)
    path = os.path.join(folder, _photo_filename(address, owner_last, "Assessor"))
    try:
        fetcher = session.get if session else requests.get
        resp = fetcher(photo_url, timeout=15)
        if resp.status_code == 200 and len(resp.content) > 1000:
            with open(path, "wb") as f:
                f.write(resp.content)
            return path
    except Exception as e:
        print(f"  [WARNING] Assessor photo download failed: {e}")
    return ""


# ═══════════════════════════════════════════════════════════════════════════════
#  SCRAPE ONE PROPERTY (Snohomish County orchestrator)
# ═══════════════════════════════════════════════════════════════════════════════

def _street_signature(addr):
    """Reduce an address to a comparable (house_number, street_core) signature
    for wildcard-result matching. Drops unit suffixes, abbreviates directions
    and street types, and ignores city/state/zip.

    Returns (number, set_of_street_words) or (None, set()) if no house number.
    """
    if not addr:
        return None, set()
    # Cut at the first comma (city/state/zip) and strip a trailing zip.
    core = addr.split(",")[0].strip()
    core = re.sub(r'\s+\d{5}(-\d{4})?$', '', core).strip()
    # Drop unit/apt suffixes: "#2", "APT 3", "UNIT B", "STE 100".
    core = re.sub(r'\s+(#|APT|UNIT|STE|SUITE)\s*\S*$', '', core, flags=re.I)
    core = re.sub(r'\s+#\S*$', '', core)
    # Normalize directions and street types to the abbreviated forms.
    for full, abbr in DIRECTION_MAP.items():
        core = re.sub(r'\b' + full + r'\b', abbr, core, flags=re.I)
    for full, abbr in STREET_TYPE_MAP.items():
        core = re.sub(r'\b' + full + r'\b', abbr, core, flags=re.I)
    parts = core.upper().split()
    if not parts or not parts[0].isdigit():
        return None, set()
    number = parts[0]
    street_words = set(parts[1:])
    if not street_words:
        # A bare house number with no street (e.g. a truncated lead '19815').
        # Treat as unparseable so callers refuse to guess a match.
        return None, set()
    return number, street_words


_SNOHO_DIRS = {"N", "S", "E", "W", "NE", "NW", "SE", "SW"}
_SNOHO_TYPES = {"ST", "AVE", "DR", "RD", "LN", "CT", "PL", "BLVD", "CIR", "TER",
                "PKWY", "HWY", "WY", "WAY", "CV", "LOOP", "RUN", "PT", "TRL"}


def _snoho_keyword(cleaned_street):
    """The portal matches best on house number + street NAME (it explicitly
    advises dropping street types and directions). Build that keyword from the
    cleaned street, e.g. '55065 E Sauk Ln' -> '55065 Sauk'."""
    parts = cleaned_street.split()
    if not parts:
        return cleaned_street
    hn = parts[0]
    core = [w for w in parts[1:]
            if w.upper() not in _SNOHO_DIRS and w.upper() not in _SNOHO_TYPES]
    return (hn + " " + " ".join(core)).strip() if core else cleaned_street


def _snoho_pick(items, cleaned_street, lead_city):
    """Choose the result matching the input street, disambiguating by city when
    several streets match. Street match = same house number and the input's
    street words are a subset of the result's (so a missing directional suffix
    like 'SE' still matches). Returns the chosen fields dict, or None when the
    match is absent or ambiguous - never guess a wrong parcel."""
    in_num, in_words = _street_signature(cleaned_street)

    def matches(f):
        n, w = _street_signature(f.get("Situs", ""))
        return in_num is not None and n == in_num and in_words and in_words.issubset(w)

    pool = [f for f in items if matches(f)] if in_num else []
    pool = pool or items
    if len(pool) == 1:
        return pool[0]
    if lead_city:
        city_matches = [f for f in pool
                        if (f.get("Situscity", "") or "").upper() == lead_city.upper()]
        if len(city_matches) == 1:
            return city_matches[0]
    return None


def _snoho_build(fields, raw_address):
    """Map a QuickSearch result to the master schema. The public search result
    carries parcel + owner + situs; rich assessor detail (value/beds/baths/year)
    is not exposed by this API, so those stay blank (like a structure-less parcel)."""
    data = {}
    data["Parcel Number"] = (fields.get("ParcelID", "") or "").strip()
    data.update(split_owner_names(fields.get("LPRDisplayName", "") or ""))
    data["Taxpayer Name"] = (fields.get("Taxpayers", "") or "").strip()

    situs = smart_title_case((fields.get("Situs", "") or "").strip())
    city = smart_title_case((fields.get("Situscity", "") or "").strip())
    # Zip from the END of the raw lead address (the portal result has no zip
    # field). Anchored to the tail so it can't grab a 5-digit house number.
    zm = re.search(r"(\d{5})(?:-\d{4})?\s*$", (raw_address or "").strip())
    zc = zm.group(1) if zm else ""
    data["Property Street"] = situs
    data["Property City"] = city
    data["Property State"] = "WA"
    data["Property Zip"] = zc
    if city:
        tail = f"{city} WA {zc}".strip()
    elif zc:
        tail = f"WA {zc}"
    else:
        tail = ""
    data["Property Address"] = ", ".join(p for p in (situs, tail) if p)
    return data


def _dd_rows(session, headers, module_id, ctx_params, referer):
    """Fetch one DataDisplay table (by its ModuleId) and return (column_set,
    list_of_row_dicts). ctx_params carries the page context the table needs:
    {p,a,m} for the Property Account Summary page, {b,a,p,d} for a Building Detail
    page. Each row dict also includes '<column>__link' for any cell that carries a
    hyperlink (used to follow the Buildings table to per-building detail)."""
    try:
        r = session.get(SNOHO_DD_GETDATA, headers={
            **headers, "ModuleId": str(module_id), "Referer": referer,
            # Force JSON: the shared session Accept lists application/xml first,
            # which makes this endpoint try (and fail) to serialize to XML (500).
            "Accept": "application/json"},
            params={**ctx_params, "itemsPerPage": "50", "page": "1"}, timeout=20)
        r.raise_for_status()
        j = r.json()
    except Exception:
        return set(), []
    rows = []
    for g in j.get("groups", []):
        for row in g.get("rows", []):
            d = {}
            for c in row.get("values", []):
                col = c.get("column")
                d[col] = c.get("value")
                if c.get("hyperlink"):
                    d[col + "__link"] = c["hyperlink"]
            rows.append(d)
    cols = {h.get("column") for h in j.get("headers", [])}
    return cols, rows


def _snoho_building_detail(session, link):
    """Follow a Buildings-table link to one building's detail page and return its
    structure facts: bedrooms, full/half baths, finished sqft, grade, condition,
    floors, exterior, roof, foundation, heat, fireplace. The page hosts DataDisplay
    tables identified by column name. Best-effort; never raises."""
    from urllib.parse import urlparse, parse_qs
    out = {}
    if not link:
        return out
    url = link if link.startswith("http") else SNOHO_BASE + link
    try:
        html = session.get(url, timeout=25).text
    except Exception:
        return out
    rvt = re.search(r'name="__RequestVerificationToken"[^>]*value="([^"]+)"', html)
    tab = re.search(r"sf_tabId[^0-9]{0,4}(\d+)", html)
    if not rvt or not tab:
        return out
    hdrs = {"TabId": tab.group(1), "RequestVerificationToken": rvt.group(1)}
    q = {k: v[0] for k, v in parse_qs(urlparse(url).query).items()}
    ctx = {k: q[k] for k in ("b", "a", "p", "d") if k in q}
    module_ids = re.findall(r'<data-display-root moduleId="(\d+)"', html)

    for mid in module_ids:
        cols, rows = _dd_rows(session, hdrs, mid, ctx, url)
        if not rows:
            continue
        # Characteristics table: Description/Units rows.
        if "Description" in cols and "Units" in cols:
            chars = {(r.get("Description") or "").strip().lower(): (r.get("Units") or "").strip()
                     for r in rows}
            if chars.get("bedroom count"):
                out["Bedrooms"] = chars["bedroom count"]
            if chars.get("full baths"):
                out["Full or 3/4 Baths"] = chars["full baths"]
            if chars.get("half baths"):
                out["Half Baths"] = chars["half baths"]
            fp = next((v for k, v in chars.items() if k.startswith("fireplace") and v), "")
            if fp:
                out["Fireplace"] = fp
        # Improvement summary: the main dwelling row carries finished area + quality.
        elif "FinishedArea" in cols:
            main = next((r for r in rows if any(t in (r.get("ImprovementModelDescr") or "").lower()
                        for t in ("detached", "single family"))), rows[0])
            for key, col in [("Total Finished SF", "FinishedArea"),
                             ("Property Grade", "QualityDescr"),
                             ("Property Condition", "ConditionDescr"),
                             ("Floor Details", "FloorDescr")]:
                v = (main.get(col) or "").strip()
                if key == "Total Finished SF":
                    v = re.sub(r"\.00$", "", v)  # 1855.00 -> 1855
                if v and v != "-":
                    out[key] = v
        # Construction elements: Foundation / Exterior / Roof / Heating by category.
        elif "ElementCategoryDescr" in cols:
            elems = {(r.get("ElementCategoryDescr") or "").strip().lower():
                     (r.get("ElementCodeDescr") or "").strip() for r in rows}
            for ecat, key in {"foundation wall": "Foundation", "ext wall cover": "Exterior",
                              "roof type": "Roof Type", "heating": "Heat"}.items():
                if elems.get(ecat):
                    out[key] = elems[ecat]
    return out


def snoho_fetch_detail(session, address_ctx, parcel, altkey, mapkey):
    """Pull rich assessor detail for a Snohomish parcel from its Property Account
    Summary page (and the linked Building Detail page). Returns a master-schema
    dict (possibly partial); never raises. The pages host each fact as a
    DataDisplay JSON table identified by column name: assessed value / market
    total / taxable, last sale date + price + sales history, year built, and (from
    the building link) bedrooms, full/half baths, finished sqft, grade, condition,
    exterior, roof, foundation, heat, fireplace."""
    out = {}
    summ = f"{SNOHO_SUMMARY_PAGE}?p={parcel}&a={altkey}&m={mapkey}"
    try:
        html = session.get(summ, timeout=25).text
    except Exception as e:
        print(f"  [WARNING] Snohomish detail page fetch failed: {e}")
        return out
    rvt = re.search(r'name="__RequestVerificationToken"[^>]*value="([^"]+)"', html)
    tab = re.search(r"sf_tabId[^0-9]{0,4}(\d+)", html)
    if not rvt or not tab:
        return out
    hdrs = {"TabId": tab.group(1), "RequestVerificationToken": rvt.group(1)}
    ctx = {"p": parcel, "a": altkey, "m": mapkey}
    module_ids = re.findall(r'<data-display-root moduleId="(\d+)"', html)

    for mid in module_ids:
        cols, rows = _dd_rows(session, hdrs, mid, ctx, summ)
        if not rows:
            continue
        # Value history: rows keyed by ValueDescription; the first value column
        # (after the label) is the most recent assessment year.
        if "ValueDescription" in cols:
            valcol = next((c for c in ("Col2", "Col3") if c in cols), None)
            for row in rows:
                desc = (row.get("ValueDescription") or "").strip().lower()
                amt = (row.get(valcol) or "").strip() if valcol else ""
                if not amt:
                    continue
                if desc == "assessed value":
                    out["Assessed Value"] = amt.lstrip("$")
                elif desc == "market total":
                    out["Market Total"] = amt.lstrip("$")
                elif desc == "taxable value regular":
                    out["Taxable Value Regular"] = amt.lstrip("$")
        # Sale / transfer history: most recent row first.
        elif "SaleDate" in cols:
            hist = []
            for row in rows:
                d = (row.get("SaleDate") or "").split(" ")[0].strip()
                price = (row.get("AdjustedSalesPrice") or "").strip()
                form = (row.get("DocSubType") or row.get("ConveyanceForm") or "").strip()
                if d:
                    hist.append(f"{d}: ${price} ({form})" if price else f"{d}: {form}")
            if rows:
                top = rows[0]
                out["Most Recent Sale Date"] = (top.get("SaleDate") or "").split(" ")[0].strip()
                out["Most Recent Sale Amount"] = (top.get("AdjustedSalesPrice") or "").strip()
            if hist:
                out["Sales History"] = " | ".join(hist)
        # Year built + the Buildings table's link to per-building detail
        # (bedrooms, baths, finished sqft, construction).
        elif "YearBuilt" in cols:
            yb = next((r.get("YearBuilt") for r in rows if (r.get("YearBuilt") or "").strip()), "")
            if yb:
                out["Year Built"] = str(yb).strip()
            if "BuildingName" in cols:
                link = next((r.get("BuildingName__link") for r in rows
                             if r.get("BuildingName__link")), "")
                if link:
                    out.update(_snoho_building_detail(session, link))
    return out


def scrape_property(session, address, base_resp, api_key=None, photo_folder="property_photos"):
    """Look up an address in the Snohomish County Public Access portal and return a
    data dict in the master schema, or None.

    `base_resp` is the auth ctx dict from get_session(). v2.8.0 replaced the
    retired snoco.org/proptax WebForms scrape with the QuickSearch JSON API; v2.8.1
    adds the rich detail (assessed value, last sale, year built, sales history) from
    the parcel's Property Account Summary page. Beds/baths/finished sqft are not
    published on this portal and stay blank. Receives the RAW address (with
    city/zip) and cleans it itself, the same split as the King County path."""
    cleaned = clean_address_for_search(address)
    lead_city = _extract_city_segment(address)
    keyword = _snoho_keyword(cleaned)

    print(f"\n{'='*60}")
    print(f"  Searching (Snohomish live): {cleaned}")
    print(f"{'='*60}")

    try:
        r = session.get(SNOHO_GETDATA, params={"keywords": keyword, "page": "1"},
                        headers=base_resp, timeout=25)
        r.raise_for_status()
        items = [it.get("fields", {}) for it in r.json().get("items", [])]
    except Exception as e:
        print(f"  [ERROR] Snohomish search failed: {e}")
        return None

    if not items:
        print(f"  [ERROR] No Snohomish parcel found for: {cleaned}")
        return None

    chosen = _snoho_pick(items, cleaned, lead_city)
    if not chosen:
        cities = ", ".join(sorted({(i.get("Situscity", "") or "").strip() for i in items}))
        print(f"  [ERROR] {len(items)} Snohomish result(s), none unambiguously matched "
              f"'{cleaned}' (results in: {cities}). Skipping rather than guessing.")
        return None

    data = _snoho_build(chosen, address)
    print(f"  -> Parcel: {data.get('Parcel Number')}")
    print(f"  -> Owner:  {data.get('Owner 1 First Name','')} {data.get('Owner 1 Last Name','')}")
    print(f"  -> Situs:  {data.get('Property Address')}")

    # Rich detail (assessed value, last sale, year built, sales history) from the
    # parcel summary page. Best-effort: a detail hiccup never drops the lead, which
    # still carries verified owner + parcel + address.
    parcel = data.get("Parcel Number", "")
    altkey = str(chosen.get("Altkey", "") or "")
    mapkey = (chosen.get("MapKey") or parcel or "")
    if parcel:
        detail = snoho_fetch_detail(session, base_resp, parcel, altkey, mapkey)
        data.update(detail)
        print(f"  -> Assessed: {data.get('Assessed Value','N/A')} | "
              f"Year built: {data.get('Year Built','N/A')} | "
              f"Last sale: {data.get('Most Recent Sale Date','N/A')} "
              f"for {data.get('Most Recent Sale Amount','N/A')}")

    # Street View photo (Google), same as the King County path.
    if api_key and data.get("Property Address"):
        sv = download_street_view(data["Property Address"],
                                  data.get("Owner 1 Last Name", ""),
                                  api_key, photo_folder)
        data["Photo File"] = sv
        if sv:
            print(f"  -> Street View photo saved: {sv}")
    else:
        data["Photo File"] = ""

    data = apply_title_case(data)
    data["County"] = "Snohomish"
    return data


# ═══════════════════════════════════════════════════════════════════════════════
#  OUTPUT: CSV + EXCEL
# ═══════════════════════════════════════════════════════════════════════════════

def get_next_row_number(script_dir):
    """Determine next Row # by reading existing CSV."""
    csv_path = os.path.join(script_dir, OUTPUT_CSV)
    if not os.path.exists(csv_path):
        return 1
    try:
        with open(csv_path, "r", encoding="utf-8") as f:
            reader = csv.DictReader(f)
            max_row = 0
            for row in reader:
                try:
                    max_row = max(max_row, int(row.get("Row #", 0)))
                except (ValueError, TypeError):
                    pass
            return max_row + 1
    except Exception:
        return 1


def write_outputs(all_data, script_dir):
    """Write/append data to CSV and rebuild Excel from CSV."""
    if not all_data:
        return

    csv_path  = os.path.join(script_dir, OUTPUT_CSV)
    xlsx_path = os.path.join(script_dir, OUTPUT_XLSX)

    # Merge fieldnames: priority order + any new fields
    existing_fields = []
    if os.path.exists(csv_path):
        try:
            with open(csv_path, "r", encoding="utf-8") as f:
                reader = csv.DictReader(f)
                existing_fields = list(reader.fieldnames or [])
        except Exception:
            pass

    all_keys = set()
    for d in all_data:
        all_keys.update(d.keys())
    if existing_fields:
        all_keys.update(existing_fields)

    fieldnames = [f for f in COLUMN_ORDER if f in all_keys]
    for f in sorted(all_keys):
        if f not in fieldnames:
            fieldnames.append(f)

    # Read existing CSV rows
    existing_rows = []
    if os.path.exists(csv_path):
        try:
            with open(csv_path, "r", encoding="utf-8") as f:
                existing_rows = list(csv.DictReader(f))
        except Exception:
            pass

    # Append new data
    all_rows = existing_rows + all_data
    new_count = len(all_data)

    # Write complete CSV
    with open(csv_path, "w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames, extrasaction="ignore")
        writer.writeheader()
        for row in all_rows:
            writer.writerow(row)

    print(f">> CSV updated: {csv_path}")
    print(f"   Total rows: {len(all_rows)} ({new_count} new)")

    # Build Excel from CSV
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Property Data"

        header_font = Font(bold=True, color="FFFFFF", size=11)
        header_fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")

        # Headers
        for col, name in enumerate(fieldnames, 1):
            cell = ws.cell(row=1, column=col, value=name)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")

        # Data rows
        for r_idx, row in enumerate(all_rows, 2):
            for c_idx, field in enumerate(fieldnames, 1):
                ws.cell(row=r_idx, column=c_idx, value=row.get(field, ""))

        # Auto-width (capped at 40)
        for col in ws.columns:
            max_len = 0
            col_letter = col[0].column_letter
            for cell in col:
                if cell.value:
                    max_len = max(max_len, len(str(cell.value)))
            ws.column_dimensions[col_letter].width = min(max_len + 2, 40)

        # Freeze header row
        ws.freeze_panes = "A2"

        wb.save(xlsx_path)
        print(f">> Excel updated: {xlsx_path}")
        print(f"   Total rows: {len(all_rows)} ({new_count} new)")

    except ImportError:
        print("  [WARNING] openpyxl not installed - Excel file not created")
    except PermissionError:
        print(f"  [ERROR] Cannot save Excel - close {OUTPUT_XLSX} and re-run")
    except Exception as e:
        print(f"  [WARNING] Excel error: {e}")


# ═══════════════════════════════════════════════════════════════════════════════
#  OUTPUT: SQLite DATABASE  (v2.6.0 - system of record)
# ═══════════════════════════════════════════════════════════════════════════════

def save_to_database(all_data, script_dir):
    """Ingest this run's records into propintel.db and refresh the spreadsheet
    view. Each record dict uses the same COLUMN_ORDER keys the CSV used, so the
    shared propintel_db.ingest_record() path handles dedup, repeat-lead
    flagging, and latest-wins property updates.

    The whole batch is one transaction. The spreadsheet (propintel_export.xlsx)
    is regenerated from the DB afterward as a convenience view; it is not the
    store, so it can be deleted and rebuilt at any time."""
    try:
        import propintel_db_v0_1_0 as pdb
    except ImportError as e:
        print(f"\n[ERROR] Could not import propintel_db_v0_1_0 ({e}).")
        print(f"        Records were NOT saved. Make sure "
              f"propintel_db_v0_1_0.py is in {script_dir}.")
        return

    conn = pdb.connect()
    pdb.init_db(conn)
    when = pdb._now()

    prop_new = prop_upd = le_new = le_collapsed = 0
    for rec in all_data:
        out = pdb.ingest_record(conn, rec, when)
        if out["property"] == "new":
            prop_new += 1
        else:
            prop_upd += 1
        if out["lead_event"] == "new":
            le_new += 1
        elif out["lead_event"] == "collapsed":
            le_collapsed += 1
    conn.commit()

    print(f"\n>> Saved to {pdb.DEFAULT_DB_FILENAME}: "
          f"{prop_new} new / {prop_upd} updated propert{'y' if (prop_new + prop_upd) == 1 else 'ies'}, "
          f"{le_new} lead event(s)"
          + (f", {le_collapsed} same-day duplicate(s) skipped" if le_collapsed else ""))

    # Regenerate the spreadsheet view from the DB (export, not the store).
    export_path = os.path.join(script_dir, "propintel_export.xlsx")
    try:
        n = pdb.export_xlsx(conn, export_path)
        print(f">> Spreadsheet view refreshed: {export_path} ({n} rows)")
    except ImportError:
        print(f"  [NOTE] openpyxl not installed; DB saved but no Excel view. "
              f"Export later with: python propintel_db_v0_1_0.py export propintel_export.xlsx")
    except PermissionError:
        print(f"  [WARNING] {export_path} is open in Excel - could not refresh "
              f"the view. The DB was saved. Close it and re-export with: "
              f"python propintel_db_v0_1_0.py export propintel_export.xlsx")
    except Exception as e:
        print(f"  [WARNING] Excel view refresh skipped: {e}")

    conn.close()


# ═══════════════════════════════════════════════════════════════════════════════
#  MAIN
# ═══════════════════════════════════════════════════════════════════════════════

def load_api_key(script_dir):
    """Load Google Maps API key from config.txt."""
    path = os.path.join(script_dir, "config.txt")
    if os.path.exists(path):
        with open(path, "r") as f:
            key = f.read().strip()
            if key:
                return key
    return None


def detect_input_format(content):
    """Detect whether file content is Market Leader, tab-separated, or plain addresses."""
    if any(kw in content for kw in MARKET_LEADER_KEYWORDS):
        return "market_leader"
    if "\t" in content and any(
        re.match(r'^[A-Za-z].*\t\d+\s', line.strip())
        for line in content.split("\n") if line.strip()
    ):
        return "tab_separated"
    return "plain_addresses"


def main():
    script_dir = os.path.dirname(os.path.abspath(__file__))
    lookups = []  # list of (search_address, lead_data_or_None)

    # --- Parse input ---
    if len(sys.argv) < 2:
        print("Usage: python snoco_scraper.py <address or file.txt>")
        sys.exit(1)

    if len(sys.argv) == 2 and sys.argv[1].endswith(".txt"):
        filepath = sys.argv[1]
        if not os.path.isabs(filepath):
            filepath = os.path.join(script_dir, filepath)

        with open(filepath, "r", encoding="utf-8") as f:
            content = f.read()

        fmt = detect_input_format(content)

        if fmt == "market_leader":
            leads = parse_lead_file(filepath)
            if not leads:
                print("[ERROR] Could not parse any leads from file")
                sys.exit(1)
            print(f"Detected Market Leader lead file: {len(leads)} lead(s) found")
            for lead in leads:
                addr = lead.get("address", "")
                if addr:
                    lead_data = build_lead_data(lead)
                    # Store the RAW address (with city). detect_county() must
                    # see the city to route correctly. scrape_property() and
                    # lookup_property_kc() clean the address themselves.
                    lookups.append((addr, lead_data))
                    print(f"  Lead: {lead.get('first_name', '')} {lead.get('last_name', '')} | {addr}")
                else:
                    print(f"  [WARNING] No address for lead: {lead.get('name', 'unknown')}")

        elif fmt == "tab_separated":
            print("Detected Name + Address format (tab-separated)")
            lookups = parse_tab_separated(content)

        else:
            addresses = [line.strip() for line in content.split("\n") if line.strip()]
            for addr in addresses:
                lookups.append((addr, None))
    else:
        # Command-line addresses
        for addr in sys.argv[1:]:
            lookups.append((addr, None))

    if not lookups:
        print("[ERROR] No addresses to look up")
        sys.exit(1)

    # --- Setup ---
    print(f"\nProperty Tax Scraper v2.8.2 (Snohomish web + King County LIVE)")
    print(f"{'='*50}")
    print(f"Properties to look up: {len(lookups)}")

    api_key = load_api_key(script_dir)
    photo_folder = os.path.join(script_dir, "property_photos")
    print(f"Google API key loaded - Street View photos enabled" if api_key
          else "No config.txt found - Street View photos disabled")

    # Which counties does this run actually need?
    counties_needed = {detect_county(addr) for addr, _ in lookups}

    # v2.4.2: if any address is in a split-county city (e.g. Bothell), we may
    # need to retry the other county on a no-match, so make sure both data
    # sources are loaded for the run.
    if any(is_split_city(addr) for addr, _ in lookups):
        counties_needed.update({"snohomish", "king"})

    # Snohomish: live JSON API on the Public Access portal (only connect if a
    # Snohomish address is present). base_resp here is the auth-context dict.
    session, base_resp = None, None
    if "snohomish" in counties_needed:
        print("\nConnecting to the Snohomish County Public Access portal...")
        try:
            session, base_resp = get_session()
            print("Connected!")
        except Exception as e:
            print(f"[WARNING] Could not connect to Snohomish County: {e}")
            session, base_resp = None, None

    # King County: live (ArcGIS address->PIN + Dashboard GET). v2.7.0 retired
    # the bulk ZIP data path. Nothing to preload; each KC address is resolved
    # on demand, and per-address failures are handled in the lookup loop.
    if "king" in counties_needed:
        print("King County: live lookups enabled (ArcGIS + assessor Dashboard).")

    if "snohomish" in counties_needed and not session:
        if "king" in counties_needed:
            print("[WARNING] Snohomish site unreachable; only King County "
                  "addresses will be processed this run.")
        else:
            print("[ERROR] Could not connect to Snohomish County. Check your "
                  "internet connection.")
            sys.exit(1)

    print()

    next_row = get_next_row_number(script_dir)
    today = datetime.now().strftime("%m/%d/%Y")

    # --- Look up each property ---
    all_data = []
    success = 0
    for i, (address, lead_data) in enumerate(lookups):
        county = detect_county(address)
        data = None

        if county == "king":
            # KC live lookup. Owner names come from the live assessor record.
            data = lookup_property_kc_live(address, lead_data, api_key, photo_folder)
        else:
            if session:
                data = scrape_property(session, address, base_resp, api_key, photo_folder)
            else:
                print(f"\n[SKIP] {address} - Snohomish County website unavailable")

        # v2.4.2: split-county-city fallback. If a primary lookup returned
        # nothing AND the city straddles the line (Bothell), try the other
        # county before giving up. This catches the v2.4.1 case where a
        # Snohomish-side Bothell parcel routed to King County and failed the
        # residential-only bulk index.
        if not data and is_split_city(address):
            other = "snohomish" if county == "king" else "king"
            print(f"  -> Split-county city, retrying as {other.title()} County...")
            if other == "king":
                data = lookup_property_kc_live(address, lead_data, api_key, photo_folder)
                if data:
                    county = "king"
            elif other == "snohomish" and session:
                data = scrape_property(session, address, base_resp, api_key, photo_folder)
                if data:
                    county = "snohomish"

        if data:
            # v2.7.0: both counties now have real assessor owner names, so they
            # share the same full ownership-match logic (lead name vs owner
            # name). The old KC lead-based match (compute_ownership_match_kc) is
            # no longer needed.
            if lead_data:
                # Lead fields don't include Owner * keys, so this never clobbers
                # the assessor owner names.
                data.update(lead_data)
                data = fix_owner_names_with_lead(data, lead_data)
                data["Ownership Match"] = compute_ownership_match(data, lead_data)
                print(f"  -> Ownership Match: {data['Ownership Match']}")
            else:
                data["Ownership Match"] = "NO LEAD DATA"
                data["Lead Type"] = ""

            data["Row #"] = str(next_row)
            data["Date Retrieved"] = today
            next_row += 1
            all_data.append(data)
            success += 1

        # Refresh the Snohomish session every 10 properties (viewstate staleness).
        # King County needs no refresh: each live lookup uses its own request.
        if (i + 1) % 10 == 0 and (i + 1) < len(lookups):
            if session:
                print(f"\n  [Refreshing Snohomish session...]")
                try:
                    session, base_resp = get_session()
                except Exception:
                    pass
                time.sleep(1)

    # --- Write output ---
    # v2.6.0: the SQLite database (propintel.db) is the system of record. Each
    # scraped record is ingested with dedup + repeat-lead flagging via the
    # shared propintel_db.ingest_record() path. The CSV append model is retired;
    # a spreadsheet VIEW is regenerated from the DB after the run (export, not
    # the store), so re-running a lead no longer creates duplicate rows and the
    # old "close the xlsx and re-run" lock failure is gone.
    if all_data:
        save_to_database(all_data, script_dir)

    print(f"\n{'='*50}")
    print(f"DONE! {success}/{len(lookups)} properties scraped successfully.")


if __name__ == "__main__":
    main()
