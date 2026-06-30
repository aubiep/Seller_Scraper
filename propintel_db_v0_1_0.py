"""
PropIntel Database  v0.1.0
==========================
The system of record for the PropIntel suite. A single local SQLite file
(propintel.db) that replaces the appending snoco_property_data.csv/.xlsx as
the place where leads, properties, and outreach live.

Why a database instead of a spreadsheet:
  - Repeat-lead detection. The same person coming back a second or third time
    is a query (count of lead_events for a contact+property), not a manual scan.
  - Filter/sort by zip, city, county for mailing lists without opening Excel.
  - No file-lock failures. The scraper's "close the xlsx and re-run" problem
    goes away because nothing holds the file open.
  - One store that intake, the scraper, Homebot, and the iAuto letter queue all
    read and write, instead of a separate spreadsheet per tool.

Excel is now an EXPORT format, not the store. `export` dumps a flat sheet that
matches the old COLUMN_ORDER whenever you want a spreadsheet to look at.

Schema (four entity tables + a version stamp):
  contacts      - people: seller leads, buyers, clients, open-house meets.
  properties    - one row per parcel (county + parcel_number unique). The
                  assessor data. Latest retrieval wins on re-scrape.
  lead_events   - each inbound lead / interaction tying a contact to a
                  property. Multiple events for the same contact+property IS
                  the repeat-lead signal.
  outreach      - downstream actions (Homebot signup, iAuto letter). status
                  draft/ready/sent/failed mirrors the dashboard letter queue.

This module is the data-access layer. Higher-level tools (scraper, the app)
import it; it also has a small CLI for setup and migration:

    python propintel_db_v0_1_0.py init
    python propintel_db_v0_1_0.py import snoco_property_data.csv
    python propintel_db_v0_1_0.py stats
    python propintel_db_v0_1_0.py export propintel_export.xlsx

Standard library only (sqlite3). openpyxl is needed only for `export`.
"""

import sqlite3
import csv
import os
import sys
from datetime import datetime

SCHEMA_VERSION = 4
DEFAULT_DB_FILENAME = "propintel.db"


# ═══════════════════════════════════════════════════════════════════════════════
#  SCHEMA
# ═══════════════════════════════════════════════════════════════════════════════

SCHEMA_SQL = """
CREATE TABLE IF NOT EXISTS contacts (
    id            INTEGER PRIMARY KEY,
    first_name    TEXT DEFAULT '',
    last_name     TEXT DEFAULT '',
    email         TEXT DEFAULT '',
    phone         TEXT DEFAULT '',
    contact_type  TEXT DEFAULT '',   -- seller_lead | buyer | client | prospect | open_house | other
    notes         TEXT DEFAULT '',
    created_at    TEXT DEFAULT '',
    updated_at    TEXT DEFAULT ''
);
CREATE INDEX IF NOT EXISTS idx_contacts_name  ON contacts(last_name, first_name);
CREATE INDEX IF NOT EXISTS idx_contacts_email ON contacts(email);

CREATE TABLE IF NOT EXISTS properties (
    id                       INTEGER PRIMARY KEY,
    county                   TEXT DEFAULT '',
    parcel_number            TEXT DEFAULT '',
    property_address         TEXT DEFAULT '',
    property_street          TEXT DEFAULT '',
    property_city            TEXT DEFAULT '',
    property_state           TEXT DEFAULT '',
    property_zip             TEXT DEFAULT '',
    owner1_first             TEXT DEFAULT '',
    owner1_last              TEXT DEFAULT '',
    owner2_first             TEXT DEFAULT '',
    owner2_last              TEXT DEFAULT '',
    owner_address            TEXT DEFAULT '',
    taxpayer_name            TEXT DEFAULT '',
    tax_address              TEXT DEFAULT '',
    property_description      TEXT DEFAULT '',
    subdivision_name         TEXT DEFAULT '',
    use_code                 TEXT DEFAULT '',
    property_category        TEXT DEFAULT '',
    status                   TEXT DEFAULT '',
    tax_code_area            TEXT DEFAULT '',
    neighborhood_code        TEXT DEFAULT '',
    structure_description    TEXT DEFAULT '',
    structure_type           TEXT DEFAULT '',
    year_built               TEXT DEFAULT '',
    bedrooms                 TEXT DEFAULT '',
    baths_full_three_quarter TEXT DEFAULT '',
    half_baths               TEXT DEFAULT '',
    total_finished_sf        TEXT DEFAULT '',
    size_gross               TEXT DEFAULT '',
    unit_of_measure          TEXT DEFAULT '',
    floor_details            TEXT DEFAULT '',
    garage_sf                TEXT DEFAULT '',
    heat                     TEXT DEFAULT '',
    fireplace                TEXT DEFAULT '',
    foundation               TEXT DEFAULT '',
    exterior                 TEXT DEFAULT '',
    roof_type                TEXT DEFAULT '',
    property_grade           TEXT DEFAULT '',
    property_condition       TEXT DEFAULT '',
    views                    TEXT DEFAULT '',
    waterfront               TEXT DEFAULT '',
    most_recent_sale_date    TEXT DEFAULT '',
    most_recent_sale_amount  TEXT DEFAULT '',
    latest_tax_year          TEXT DEFAULT '',
    assessed_value           TEXT DEFAULT '',
    market_total             TEXT DEFAULT '',
    market_land              TEXT DEFAULT '',
    market_improvement       TEXT DEFAULT '',
    taxable_value_regular    TEXT DEFAULT '',
    annual_tax_amount        TEXT DEFAULT '',
    sales_history            TEXT DEFAULT '',
    photo_file               TEXT DEFAULT '',
    assessor_photo_file      TEXT DEFAULT '',
    latitude                 REAL,
    longitude                REAL,
    geocode_source           TEXT DEFAULT '',
    geocoded_at              TEXT DEFAULT '',
    last_lookup_at           TEXT DEFAULT '',
    last_lookup_status       TEXT DEFAULT '',
    last_lookup_detail       TEXT DEFAULT '',
    first_retrieved          TEXT DEFAULT '',
    last_retrieved           TEXT DEFAULT '',
    updated_at               TEXT DEFAULT '',
    UNIQUE(county, parcel_number)
);
CREATE INDEX IF NOT EXISTS idx_properties_zip  ON properties(property_zip);
CREATE INDEX IF NOT EXISTS idx_properties_city ON properties(property_city);

CREATE TABLE IF NOT EXISTS lead_events (
    id               INTEGER PRIMARY KEY,
    contact_id       INTEGER REFERENCES contacts(id),
    property_id      INTEGER REFERENCES properties(id),
    lead_type        TEXT DEFAULT '',
    lead_source      TEXT DEFAULT '',
    ownership_match  TEXT DEFAULT '',     -- CONFIRMED | LIKELY | REVIEW | MISMATCH | NO LEAD DATA
    needs_review     INTEGER DEFAULT 0,   -- 1 when the match is unconfirmed/ambiguous
    match_confidence INTEGER DEFAULT 0,   -- 0-100, from the ownership reasoner (schema v2)
    match_relationship TEXT DEFAULT '',   -- self | spouse_or_relative | entity | ... (schema v2)
    match_reason     TEXT DEFAULT '',     -- human-readable explanation (schema v2)
    years_since_sale TEXT DEFAULT '',
    equity_level     TEXT DEFAULT '',
    received_date    TEXT DEFAULT '',
    source_row       INTEGER,             -- original CSV Row # (provenance)
    notes            TEXT DEFAULT '',
    created_at       TEXT DEFAULT ''
);
CREATE INDEX IF NOT EXISTS idx_lead_events_contact  ON lead_events(contact_id);
CREATE INDEX IF NOT EXISTS idx_lead_events_property ON lead_events(property_id);

CREATE TABLE IF NOT EXISTS outreach (
    id          INTEGER PRIMARY KEY,
    contact_id  INTEGER REFERENCES contacts(id),
    property_id INTEGER REFERENCES properties(id),
    channel     TEXT DEFAULT '',   -- homebot | iauto
    template    TEXT DEFAULT '',
    status      TEXT DEFAULT '',   -- draft | ready | sent | failed
    detail      TEXT DEFAULT '',   -- error message, queue ref, etc.
    queued_at   TEXT DEFAULT '',
    sent_at     TEXT DEFAULT '',
    created_at  TEXT DEFAULT ''
);
CREATE INDEX IF NOT EXISTS idx_outreach_contact ON outreach(contact_id);
CREATE INDEX IF NOT EXISTS idx_outreach_status  ON outreach(status);

CREATE TABLE IF NOT EXISTS schema_meta (
    key   TEXT PRIMARY KEY,
    value TEXT
);
"""


# Map old CSV header -> properties table column. Lead-only and event-only
# columns (Row #, Date Retrieved, Lead *, Ownership Match, ...) are handled
# separately in the importer and are intentionally absent here.
CSV_TO_PROPERTY = {
    "County": "county",
    "Property Address": "property_address",
    "Property Street": "property_street",
    "Property City": "property_city",
    "Property State": "property_state",
    "Property Zip": "property_zip",
    "Parcel Number": "parcel_number",
    "Property Description": "property_description",
    "Subdivision Name": "subdivision_name",
    "Owner 1 First Name": "owner1_first",
    "Owner 1 Last Name": "owner1_last",
    "Owner 2 First Name": "owner2_first",
    "Owner 2 Last Name": "owner2_last",
    "Owner Address": "owner_address",
    "Taxpayer Name": "taxpayer_name",
    "Tax Address": "tax_address",
    "Structure Description": "structure_description",
    "Structure Type": "structure_type",
    "Year Built": "year_built",
    "Bedrooms": "bedrooms",
    "Full or 3/4 Baths": "baths_full_three_quarter",
    "Half Baths": "half_baths",
    "Total Finished SF": "total_finished_sf",
    "Size (gross)": "size_gross",
    "Unit of Measure": "unit_of_measure",
    "Most Recent Sale Date": "most_recent_sale_date",
    "Most Recent Sale Amount": "most_recent_sale_amount",
    "Neighborhood Code": "neighborhood_code",
    "Latest Tax Year": "latest_tax_year",
    "Assessed Value": "assessed_value",
    "Market Total": "market_total",
    "Market Land": "market_land",
    "Market Improvement": "market_improvement",
    "Taxable Value Regular": "taxable_value_regular",
    "Annual Tax Amount": "annual_tax_amount",
    "Use Code": "use_code",
    "Property Category": "property_category",
    "Status": "status",
    "Tax Code Area": "tax_code_area",
    "Floor Details": "floor_details",
    "Garage SF": "garage_sf",
    "Heat": "heat",
    "Fireplace": "fireplace",
    "Foundation": "foundation",
    "Exterior": "exterior",
    "Roof Type": "roof_type",
    "Property Grade": "property_grade",
    "Property Condition": "property_condition",
    "Views": "views",
    "Waterfront": "waterfront",
    "Sales History": "sales_history",
    "Photo File": "photo_file",
    "Assessor Photo File": "assessor_photo_file",
}

# Reverse map (property column -> CSV header) plus the lead/event columns, used
# by export() to reproduce the original flat COLUMN_ORDER sheet.
EXPORT_COLUMN_ORDER = [
    "Row #", "Date Retrieved", "County",
    "Lead First Name", "Lead Last Name", "Lead Email", "Lead Phone",
    "Lead Type", "Lead Source", "Years Since Sale", "Equity Level",
    "Property Address", "Property Street", "Property City", "Property State", "Property Zip",
    "Parcel Number", "Property Description", "Subdivision Name",
    "Owner 1 First Name", "Owner 1 Last Name", "Owner 2 First Name", "Owner 2 Last Name",
    "Ownership Match", "Owner Address", "Taxpayer Name", "Tax Address",
    "Structure Description", "Structure Type", "Year Built",
    "Bedrooms", "Full or 3/4 Baths", "Half Baths", "Total Finished SF",
    "Size (gross)", "Unit of Measure",
    "Most Recent Sale Date", "Most Recent Sale Amount", "Neighborhood Code",
    "Latest Tax Year", "Assessed Value", "Market Total", "Market Land",
    "Market Improvement", "Taxable Value Regular", "Annual Tax Amount",
    "Use Code", "Property Category", "Status", "Tax Code Area",
    "Floor Details", "Garage SF", "Heat", "Fireplace", "Foundation",
    "Exterior", "Roof Type", "Property Grade", "Property Condition",
    "Views", "Waterfront", "Sales History",
    "Photo File", "Assessor Photo File",
]


# ═══════════════════════════════════════════════════════════════════════════════
#  CONNECTION + INIT
# ═══════════════════════════════════════════════════════════════════════════════

def connect(db_path=None):
    """Open (creating if needed) the PropIntel database. Returns a connection
    with row access by column name and foreign keys enforced."""
    if db_path is None:
        db_path = os.path.join(os.path.dirname(os.path.abspath(__file__)),
                               DEFAULT_DB_FILENAME)
    conn = sqlite3.connect(db_path)
    conn.row_factory = sqlite3.Row
    conn.execute("PRAGMA foreign_keys = ON")
    return conn


def init_db(conn):
    """Create all tables/indexes if they don't exist, run lightweight column
    migrations on pre-existing databases, and stamp the version."""
    conn.executescript(SCHEMA_SQL)
    _migrate(conn)
    conn.execute(
        "INSERT INTO schema_meta(key, value) VALUES('schema_version', ?) "
        "ON CONFLICT(key) DO UPDATE SET value=excluded.value",
        (str(SCHEMA_VERSION),),
    )
    conn.commit()


def _migrate(conn):
    """Add columns introduced by later schema versions to an older database.
    SQLite ADD COLUMN is cheap and non-destructive; existing rows get the
    default. Safe to run every startup (checks what already exists)."""
    have = {r["name"] for r in conn.execute("PRAGMA table_info(lead_events)")}
    additions = [
        ("match_confidence", "INTEGER DEFAULT 0"),
        ("match_relationship", "TEXT DEFAULT ''"),
        ("match_reason", "TEXT DEFAULT ''"),
    ]
    for col, decl in additions:
        if col not in have:
            conn.execute(f"ALTER TABLE lead_events ADD COLUMN {col} {decl}")

    # Schema v3: property coordinates for radius / map filtering (geocode module).
    # Schema v4: last re-run assessor-lookup outcome (status + human-readable
    # detail of what was found), surfaced on the lead detail page so a failed
    # re-enrichment shows what the assessor returned instead of just failing.
    have_p = {r["name"] for r in conn.execute("PRAGMA table_info(properties)")}
    prop_additions = [
        ("latitude", "REAL"),
        ("longitude", "REAL"),
        ("geocode_source", "TEXT DEFAULT ''"),
        ("geocoded_at", "TEXT DEFAULT ''"),
        ("last_lookup_at", "TEXT DEFAULT ''"),
        ("last_lookup_status", "TEXT DEFAULT ''"),
        ("last_lookup_detail", "TEXT DEFAULT ''"),
    ]
    for col, decl in prop_additions:
        if col not in have_p:
            conn.execute(f"ALTER TABLE properties ADD COLUMN {col} {decl}")


def _now():
    return datetime.now().strftime("%Y-%m-%d %H:%M:%S")


def titlecase_name(s):
    """Normalize a person's name to display case. Fixes the common all-lower or
    ALL-CAPS lead names ('mary surprenant' -> 'Mary Surprenant') and handles
    Mc/Mac/O'/hyphen/apostrophe forms. Leaves names that already carry a
    deliberate internal capital (DeAngelo, McLeod) untouched."""
    if not s:
        return s
    out = []
    for w in s.split():
        if not w:
            continue
        # Already intentionally mixed-case (and not ALLCAPS) -> trust it.
        if not w.isupper() and w[1:] != w[1:].lower():
            out.append(w)
            continue
        lw = w.lower()
        if lw.startswith("mc") and len(lw) > 2:
            out.append("Mc" + lw[2].upper() + lw[3:])
        elif lw.startswith("mac") and len(lw) > 4:
            out.append("Mac" + lw[3].upper() + lw[4:])
        elif "'" in lw:
            out.append("'".join(p[:1].upper() + p[1:] for p in lw.split("'")))
        elif "-" in lw:
            out.append("-".join(p[:1].upper() + p[1:] for p in lw.split("-")))
        else:
            out.append(lw[:1].upper() + lw[1:])
    return " ".join(out)


# ═══════════════════════════════════════════════════════════════════════════════
#  UPSERT HELPERS
# ═══════════════════════════════════════════════════════════════════════════════

def _norm(s):
    return (s or "").strip()


def _contact_key(first, last, email):
    """Identity key for dedup: name + email, case-folded. A contact with an
    email is keyed on email; otherwise on name alone."""
    email = _norm(email).lower()
    first = _norm(first).lower()
    last = _norm(last).lower()
    if email:
        return ("email", email)
    if first or last:
        return ("name", f"{first}|{last}")
    return None


def find_or_create_contact(conn, first, last, email, phone, contact_type, when):
    """Return the contact id for this person, creating the row if new.
    Dedup is by email when present, else by case-folded first+last name.
    Blank-name/blank-email people (owner-only rows with no lead) return None."""
    # Normalize incoming names to display case (lead names often arrive lower/UPPER).
    first = titlecase_name(_norm(first))
    last = titlecase_name(_norm(last))
    key = _contact_key(first, last, email)
    if key is None:
        return None

    cur = conn.cursor()
    if key[0] == "email":
        cur.execute("SELECT id FROM contacts WHERE lower(email)=? AND email<>''",
                    (key[1],))
    else:
        cur.execute(
            "SELECT id FROM contacts WHERE lower(first_name)=? AND lower(last_name)=? "
            "AND (email='' OR email IS NULL)",
            (_norm(first).lower(), _norm(last).lower()),
        )
    row = cur.fetchone()
    if row:
        cid = row["id"]
        # Backfill phone/email/type if we now have them and didn't before.
        cur.execute(
            "UPDATE contacts SET "
            "email = CASE WHEN email='' THEN ? ELSE email END, "
            "phone = CASE WHEN phone='' THEN ? ELSE phone END, "
            "contact_type = CASE WHEN contact_type='' THEN ? ELSE contact_type END, "
            "updated_at = ? WHERE id = ?",
            (_norm(email), _norm(phone), _norm(contact_type), when, cid),
        )
        return cid

    cur.execute(
        "INSERT INTO contacts(first_name,last_name,email,phone,contact_type,"
        "created_at,updated_at) VALUES(?,?,?,?,?,?,?)",
        (_norm(first), _norm(last), _norm(email), _norm(phone),
         _norm(contact_type), when, when),
    )
    return cur.lastrowid


# Auto-geocode a property at ingest so new leads carry coordinates for the
# dashboard radius filter without a separate backfill run. Best-effort and
# non-blocking; set False to skip (e.g. a huge bulk import where the per-row
# geocode latency isn't wanted - the /letters/bulk "Geocode N now" button or
# `geocode_v0_1_0.py --backfill` can fill them in afterward).
GEOCODE_ON_UPSERT = True


def _maybe_geocode(conn, pid):
    """Geocode a property that has a street address but no coordinates yet.
    Never raises and never blocks the write: the row is already saved; this only
    adds lat/long. The lazy import avoids a circular dependency (geocode_v0_1_0
    imports this module) and keeps `requests` off the import path for tools that
    never write properties."""
    if not GEOCODE_ON_UPSERT:
        return
    try:
        row = conn.execute(
            "SELECT property_address, property_street, property_city, property_state, "
            "property_zip, latitude FROM properties WHERE id=?", (pid,)).fetchone()
        if not row or row["latitude"] is not None or not _norm(row["property_street"]):
            return
        addr = _norm(row["property_address"]) or " ".join(
            _norm(row[c]) for c in
            ("property_street", "property_city", "property_state", "property_zip")).strip()
        if not addr:
            return
        import geocode_v0_1_0 as _geo
        hit = _geo.geocode_address(addr)
        if hit:
            conn.execute(
                "UPDATE properties SET latitude=?, longitude=?, geocode_source=?, "
                "geocoded_at=? WHERE id=?",
                (hit["lat"], hit["lng"], hit["source"], _now(), pid))
    except Exception:
        pass  # best-effort: ingest must never fail because geocoding did


def upsert_property(conn, prop_fields, retrieved_date, when):
    """Insert or update a property keyed on (county, parcel_number). On an
    existing parcel, non-blank incoming fields overwrite (latest retrieval
    wins) and first/last_retrieved are maintained. Returns the property id.

    Parcels with no parcel_number can't be deduped reliably, so they fall back
    to keying on (county, property_address)."""
    county = _norm(prop_fields.get("county"))
    parcel = _norm(prop_fields.get("parcel_number"))
    cur = conn.cursor()

    if parcel:
        cur.execute("SELECT id FROM properties WHERE county=? AND parcel_number=?",
                    (county, parcel))
    else:
        cur.execute(
            "SELECT id FROM properties WHERE county=? AND parcel_number='' "
            "AND property_address=?",
            (county, _norm(prop_fields.get("property_address"))),
        )
    row = cur.fetchone()

    if row:
        pid = row["id"]
        # Overwrite only with non-blank incoming values so a sparse re-scrape
        # never erases good data.
        sets, vals = [], []
        for col, val in prop_fields.items():
            if _norm(val):
                sets.append(f"{col} = ?")
                vals.append(_norm(val))
        sets.append("last_retrieved = ?"); vals.append(retrieved_date or when)
        sets.append("updated_at = ?");     vals.append(when)
        vals.append(pid)
        cur.execute(f"UPDATE properties SET {', '.join(sets)} WHERE id = ?", vals)
        _maybe_geocode(conn, pid)
        return pid

    cols = list(prop_fields.keys()) + ["first_retrieved", "last_retrieved", "updated_at"]
    vals = [_norm(prop_fields[c]) for c in prop_fields] + \
           [retrieved_date or when, retrieved_date or when, when]
    placeholders = ",".join("?" * len(cols))
    cur.execute(f"INSERT INTO properties({','.join(cols)}) VALUES({placeholders})", vals)
    pid = cur.lastrowid
    _maybe_geocode(conn, pid)
    return pid


def lead_event_exists(conn, contact_id, property_id, received_date, lead_source):
    """True if an identical lead event is already recorded. Used to collapse
    test re-runs (same person, same property, same day, same source) so they
    don't inflate the repeat-lead count."""
    cur = conn.cursor()
    cur.execute(
        "SELECT 1 FROM lead_events WHERE "
        "IFNULL(contact_id,-1)=IFNULL(?,-1) AND property_id=? "
        "AND received_date=? AND lead_source=? LIMIT 1",
        (contact_id, property_id, received_date, lead_source),
    )
    return cur.fetchone() is not None


def _analyze_match(row):
    """Run the ownership reasoner on a record dict (legacy COLUMN_ORDER keys).
    Returns (status, needs_review, confidence, relationship, reason). Falls back
    to the row's precomputed Ownership Match if the reasoner module is absent."""
    lead_first = row.get("Lead First Name", "")
    lead_last = row.get("Lead Last Name", "")
    if not (_norm(lead_first) or _norm(lead_last)):
        return ("NO LEAD DATA", 0, 0, "none", "No lead name provided.")
    try:
        import ownership_match_v0_1_0 as om
    except Exception:
        m = _norm(row.get("Ownership Match")).upper()
        return (m or "REVIEW", 1 if m in ("PARTIAL", "MISMATCH", "") else 0, 0, "", "")
    owners = [
        {"first": row.get("Owner 1 First Name", ""), "last": row.get("Owner 1 Last Name", "")},
        {"first": row.get("Owner 2 First Name", ""), "last": row.get("Owner 2 Last Name", "")},
    ]
    r = om.analyze_ownership(lead_first, lead_last, owners)
    return (r["status"], 1 if om.needs_review(r) else 0,
            r["confidence"], r["relationship"], r["reason"])


def add_lead_event(conn, contact_id, property_id, fields, when):
    """Insert a lead event with its ownership-match verdict. The match status,
    needs_review flag, confidence, relationship, and reason are supplied by the
    caller (computed via _analyze_match)."""
    cur = conn.cursor()
    cur.execute(
        "INSERT INTO lead_events(contact_id,property_id,lead_type,lead_source,"
        "ownership_match,needs_review,match_confidence,match_relationship,"
        "match_reason,years_since_sale,equity_level,received_date,source_row,"
        "created_at) VALUES(?,?,?,?,?,?,?,?,?,?,?,?,?,?)",
        (contact_id, property_id,
         _norm(fields.get("lead_type")), _norm(fields.get("lead_source")),
         _norm(fields.get("ownership_match")), int(fields.get("needs_review", 1)),
         int(fields.get("match_confidence", 0)), _norm(fields.get("match_relationship")),
         _norm(fields.get("match_reason")),
         _norm(fields.get("years_since_sale")), _norm(fields.get("equity_level")),
         _norm(fields.get("received_date")), fields.get("source_row"), when),
    )
    return cur.lastrowid


def add_addressless_lead(conn, first, last, email, phone, source, received_date,
                         seen_address="", when=None):
    """Capture a high-intent seller lead that has no usable property address - e.g.
    a Zurple 'Homeowner Asked for a CMA' alert. The person explicitly asked for a
    valuation but can't be enriched/scored without their home address. Creates the
    contact and a placeholder property keyed to the person (parcel 'NOADDR-<email|
    name>' so re-imports dedup, not collapse), with a lead_event flagged for
    follow-up. `seen_address` is any address the source listed but that we don't
    trust as their home (Zurple often lists a browsed listing) - noted, not attached.
    Returns the contact id, or None if there's no usable name/email."""
    when = when or _now()
    cid = find_or_create_contact(conn, first, last, email, phone, "seller_lead", when)
    if cid is None:
        return None
    key = (_norm(email) or f"{_norm(first)} {_norm(last)}").lower()
    pid = upsert_property(conn, {
        "county": "", "parcel_number": f"NOADDR-{key}",
        "property_address": "", "property_street": "", "property_city": "",
        "property_state": "", "property_zip": ""}, when, when)
    rdate = _norm(received_date) or when[:10]
    if _norm(seen_address):
        verdict = "UNVERIFIED"
        reason = (f"Seller lead. Source-listed address '{_norm(seen_address)}' is "
                  "unverified (may be a browsed listing, not their home). Confirm it, "
                  "set it via Edit contact, then it enriches.")
    else:
        verdict = "NO ADDRESS"
        reason = ("Seller asked for a CMA; no property address in the lead. "
                  "Add an address (Edit contact) to enrich + score.")
    if not lead_event_exists(conn, cid, pid, rdate, _norm(source)):
        add_lead_event(conn, cid, pid, {
            "lead_type": "Seller", "lead_source": source,
            "ownership_match": verdict, "needs_review": 0, "match_confidence": 0,
            "match_relationship": "", "match_reason": reason,
            "received_date": rdate, "source_row": None}, when)
    return cid


def ingest_record(conn, row, when=None):
    """Ingest one record into the database. `row` is a dict keyed by the legacy
    COLUMN_ORDER headers (e.g. "Property Address", "Lead First Name",
    "Ownership Match", "Date Retrieved", "Row #") - exactly the dict the scraper
    builds for each property and that the old CSV stored. This is the single
    write path shared by the scraper and the CSV importer.

    Upserts the property (county+parcel dedup, latest non-blank wins), finds or
    creates the contact (email/name dedup), and records a lead event unless an
    identical same-day event already exists. The caller is responsible for
    commit() so a batch can be one transaction.

    Returns a dict: {property_id, property: 'new'|'updated', contact_id,
    lead_event: 'new'|'collapsed'|'none'}.
    """
    if when is None:
        when = _now()

    # --- property ---
    prop_fields = {col: row.get(csv_h, "") for csv_h, col in CSV_TO_PROPERTY.items()}
    retrieved = _norm(row.get("Date Retrieved"))
    county = _norm(prop_fields.get("county"))
    parcel = _norm(prop_fields.get("parcel_number"))
    if parcel:
        existed = conn.execute(
            "SELECT 1 FROM properties WHERE county=? AND parcel_number=?",
            (county, parcel)).fetchone() is not None
    else:
        existed = conn.execute(
            "SELECT 1 FROM properties WHERE county=? AND parcel_number='' "
            "AND property_address=?",
            (county, _norm(prop_fields.get("property_address")))).fetchone() is not None
    pid = upsert_property(conn, prop_fields, retrieved, when)

    # --- contact ---
    first = row.get("Lead First Name", "")
    last = row.get("Lead Last Name", "")
    email = row.get("Lead Email", "")
    phone = row.get("Lead Phone", "")
    ctype = "seller_lead" if (_norm(first) or _norm(last) or _norm(email)) else ""
    cid = find_or_create_contact(conn, first, last, email, phone, ctype, when)

    # --- lead event ---
    le_outcome = "none"
    if cid is not None:
        lead_source = _norm(row.get("Lead Source"))
        if lead_event_exists(conn, cid, pid, retrieved, lead_source):
            le_outcome = "collapsed"
        else:
            try:
                source_row = int(row.get("Row #") or 0) or None
            except (ValueError, TypeError):
                source_row = None
            status, nr, conf, rel, reason = _analyze_match(row)
            add_lead_event(conn, cid, pid, {
                "lead_type": row.get("Lead Type"),
                "lead_source": lead_source,
                "ownership_match": status,
                "needs_review": nr,
                "match_confidence": conf,
                "match_relationship": rel,
                "match_reason": reason,
                "years_since_sale": row.get("Years Since Sale"),
                "equity_level": row.get("Equity Level"),
                "received_date": retrieved,
                "source_row": source_row,
            }, when)
            le_outcome = "new"

    return {"property_id": pid, "property": "updated" if existed else "new",
            "contact_id": cid, "lead_event": le_outcome}


# ═══════════════════════════════════════════════════════════════════════════════
#  CSV IMPORT (migration from snoco_property_data.csv)
# ═══════════════════════════════════════════════════════════════════════════════

def import_from_csv(conn, csv_path):
    """Migrate the legacy snoco_property_data.csv into the database.

    Each CSV row is split into:
      - a property (deduped on county+parcel; latest retrieval wins),
      - a contact (deduped on email/name; skipped for owner-only rows with no
        lead name or email),
      - a lead event (skipped when an identical same-day event already exists,
        which collapses the repeated test-run rows).

    Returns a stats dict.
    """
    stats = {
        "csv_rows": 0, "properties_new": 0, "properties_updated": 0,
        "contacts_new": 0, "lead_events_new": 0,
        "lead_events_collapsed": 0, "rows_without_lead": 0,
    }
    when = _now()

    with open(csv_path, "r", encoding="utf-8") as f:
        reader = csv.DictReader(f)
        rows = list(reader)

    # Process in Row # order so "latest retrieval wins" actually keeps the
    # newest data (the CSV is already append-ordered, but be explicit).
    def _row_key(r):
        try:
            return int(r.get("Row #", 0) or 0)
        except (ValueError, TypeError):
            return 0
    rows.sort(key=_row_key)

    contacts_before = conn.execute("SELECT COUNT(*) FROM contacts").fetchone()[0]
    props_before = conn.execute("SELECT COUNT(*) FROM properties").fetchone()[0]

    for r in rows:
        stats["csv_rows"] += 1
        out = ingest_record(conn, r, when)
        if out["property"] == "updated":
            stats["properties_updated"] += 1
        if out["contact_id"] is None:
            stats["rows_without_lead"] += 1
        if out["lead_event"] == "collapsed":
            stats["lead_events_collapsed"] += 1
        elif out["lead_event"] == "new":
            stats["lead_events_new"] += 1

    conn.commit()
    stats["contacts_new"] = conn.execute(
        "SELECT COUNT(*) FROM contacts").fetchone()[0] - contacts_before
    # properties_new counted above is per-row "first sight"; reconcile to the
    # actual unique count added for an honest number.
    stats["properties_new"] = conn.execute(
        "SELECT COUNT(*) FROM properties").fetchone()[0] - props_before
    return stats


# ═══════════════════════════════════════════════════════════════════════════════
#  REPORTING + EXPORT
# ═══════════════════════════════════════════════════════════════════════════════

def stats_summary(conn):
    """Return counts and the repeat-lead list (contacts with >1 event on the
    same property)."""
    g = lambda q, *a: conn.execute(q, a).fetchone()[0]
    out = {
        "contacts": g("SELECT COUNT(*) FROM contacts"),
        "properties": g("SELECT COUNT(*) FROM properties"),
        "lead_events": g("SELECT COUNT(*) FROM lead_events"),
        "needs_review": g("SELECT COUNT(*) FROM lead_events WHERE needs_review=1"),
        "by_county": conn.execute(
            "SELECT county, COUNT(*) c FROM properties GROUP BY county ORDER BY c DESC"
        ).fetchall(),
        "repeats": conn.execute(
            "SELECT c.first_name, c.last_name, p.property_address, "
            "COUNT(*) n FROM lead_events le "
            "JOIN contacts c ON c.id=le.contact_id "
            "JOIN properties p ON p.id=le.property_id "
            "GROUP BY le.contact_id, le.property_id HAVING n>1 ORDER BY n DESC"
        ).fetchall(),
    }
    return out


def reanalyze_matches(conn):
    """Re-run the ownership reasoner over every existing lead event and update
    its verdict in place. Use after upgrading the reasoner or importing legacy
    data whose matches predate it. Returns a {status: count} summary."""
    import ownership_match_v0_1_0 as om
    rows = conn.execute(
        "SELECT le.id, c.first_name AS lf, c.last_name AS ll, "
        "p.owner1_first, p.owner1_last, p.owner2_first, p.owner2_last "
        "FROM lead_events le "
        "JOIN contacts c   ON c.id = le.contact_id "
        "JOIN properties p ON p.id = le.property_id"
    ).fetchall()
    summary = {}
    for r in rows:
        owners = [
            {"first": r["owner1_first"], "last": r["owner1_last"]},
            {"first": r["owner2_first"], "last": r["owner2_last"]},
        ]
        res = om.analyze_ownership(r["lf"], r["ll"], owners)
        conn.execute(
            "UPDATE lead_events SET ownership_match=?, needs_review=?, "
            "match_confidence=?, match_relationship=?, match_reason=? WHERE id=?",
            (res["status"], 1 if om.needs_review(res) else 0, res["confidence"],
             res["relationship"], res["reason"], r["id"]),
        )
        summary[res["status"]] = summary.get(res["status"], 0) + 1
    conn.commit()
    return summary


def review_queue(conn, limit=100):
    """Return lead events that need human review, worst-confidence first, with
    the lead name, property, verdict, and reason. This is the work list for the
    investigative step (and the future app's review screen)."""
    return conn.execute(
        "SELECT c.first_name, c.last_name, p.property_address, p.owner1_first, "
        "p.owner1_last, le.ownership_match, le.match_confidence, "
        "le.match_relationship, le.match_reason "
        "FROM lead_events le "
        "JOIN contacts c   ON c.id = le.contact_id "
        "JOIN properties p ON p.id = le.property_id "
        "WHERE le.needs_review = 1 "
        "ORDER BY le.match_confidence DESC, c.last_name "
        "LIMIT ?", (limit,)
    ).fetchall()


def submission_counts(conn):
    """Per-contact lead-submission count and date range, for the repeat-lead
    signal (a person submitting more than once = strong motivation). Returns
    {contact_id: {"n": int, "first": date_str, "last": date_str}}. Dates are the
    raw received_date strings (formats are mixed in the data, so treat as labels,
    not for sorting)."""
    out = {}
    for r in conn.execute(
        "SELECT contact_id, COUNT(*) n, MIN(received_date) first_d, "
        "MAX(received_date) last_d FROM lead_events GROUP BY contact_id"):
        out[r["contact_id"]] = {"n": r["n"], "first": r["first_d"] or "",
                                "last": r["last_d"] or ""}
    return out


def new_unprocessed_leads(conn, limit=25):
    """Leads with no outreach yet (no iAuto letter queued/sent, not enrolled in
    Homebot): the 'just came in, not acted on' list for the daily home screen.
    One row per contact+property, most recently ingested first (MAX(le.id), which
    also pulls that latest lead_event's fields via SQLite's min/max-row rule)."""
    return conn.execute(
        "SELECT le.contact_id, le.property_id, MAX(le.id) leid, "
        "c.first_name, c.last_name, c.email, le.lead_type, le.lead_source, "
        "le.ownership_match, le.received_date, "
        "p.property_address, p.property_city, p.photo_file, p.assessor_photo_file "
        "FROM lead_events le "
        "JOIN contacts c   ON c.id = le.contact_id "
        "JOIN properties p ON p.id = le.property_id "
        "WHERE NOT EXISTS (SELECT 1 FROM outreach o "
        "  WHERE o.contact_id=le.contact_id AND o.property_id=le.property_id) "
        "GROUP BY le.contact_id, le.property_id "
        "ORDER BY leid DESC LIMIT ?", (limit,)).fetchall()


def followup_due(conn, days=14):
    """iAuto letters sent at least `days` ago with no newer outreach to that
    contact: the 'initial letter went out ~2 weeks back, time to follow up' list.
    sent_at is a 'YYYY-MM-DD HH:MM:SS' stamp so date math is reliable here."""
    return conn.execute(
        "SELECT o.id, o.contact_id, o.property_id, o.sent_at, o.template, "
        "c.first_name, c.last_name, p.property_address, p.property_city "
        "FROM outreach o "
        "JOIN contacts c   ON c.id = o.contact_id "
        "JOIN properties p ON p.id = o.property_id "
        "WHERE o.channel='iauto' AND o.status='sent' "
        "AND o.sent_at <= datetime('now', ?) "
        "AND NOT EXISTS (SELECT 1 FROM outreach o2 "
        "  WHERE o2.contact_id=o.contact_id AND o2.id > o.id) "
        "ORDER BY o.sent_at ASC", (f"-{int(days)} days",)).fetchall()


def export_xlsx(conn, xlsx_path):
    """Write a flat sheet matching the legacy COLUMN_ORDER: one row per lead
    event joined to its property (and a property-only row when a property has
    no lead events). Excel is an output, not the store."""
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment

    # property column -> CSV header, for assembling export rows
    prop_to_csv = {v: k for k, v in CSV_TO_PROPERTY.items()}

    rows_out = []
    # Lead-event rows
    le_rows = conn.execute(
        "SELECT le.*, c.first_name lf, c.last_name ll, c.email le_email, "
        "c.phone le_phone, p.* FROM lead_events le "
        "JOIN properties p ON p.id=le.property_id "
        "LEFT JOIN contacts c ON c.id=le.contact_id "
        "ORDER BY le.source_row"
    ).fetchall()

    def _emit(prop_row, le=None, lead=None):
        d = {}
        for col in prop_row.keys():
            if col in prop_to_csv:
                d[prop_to_csv[col]] = prop_row[col]
        if lead:
            d["Lead First Name"] = lead.get("first", "")
            d["Lead Last Name"] = lead.get("last", "")
            d["Lead Email"] = lead.get("email", "")
            d["Lead Phone"] = lead.get("phone", "")
        if le is not None:
            d["Lead Type"] = le["lead_type"]
            d["Lead Source"] = le["lead_source"]
            d["Ownership Match"] = le["ownership_match"]
            d["Years Since Sale"] = le["years_since_sale"]
            d["Equity Level"] = le["equity_level"]
            d["Date Retrieved"] = le["received_date"]
            d["Row #"] = le["source_row"]
        return d

    seen_props = set()
    for le in le_rows:
        seen_props.add(le["property_id"])
        rows_out.append(_emit(le, le, {
            "first": le["lf"] or "", "last": le["ll"] or "",
            "email": le["le_email"] or "", "phone": le["le_phone"] or ""}))
    # Properties with no lead events (owner-only rows like the Pouncey home)
    for p in conn.execute("SELECT * FROM properties").fetchall():
        if p["id"] not in seen_props:
            rows_out.append(_emit(p))

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Property Data"
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
    for col, name in enumerate(EXPORT_COLUMN_ORDER, 1):
        cell = ws.cell(row=1, column=col, value=name)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
    for r_idx, row in enumerate(rows_out, 2):
        for c_idx, field in enumerate(EXPORT_COLUMN_ORDER, 1):
            ws.cell(row=r_idx, column=c_idx, value=row.get(field, ""))
    for col in ws.columns:
        max_len = max((len(str(c.value)) for c in col if c.value), default=0)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 2, 40)
    ws.freeze_panes = "A2"
    wb.save(xlsx_path)
    return len(rows_out)


# ═══════════════════════════════════════════════════════════════════════════════
#  CLI
# ═══════════════════════════════════════════════════════════════════════════════

def _print_stats(s):
    print(f"  Contacts:     {s['contacts']:,}")
    print(f"  Properties:   {s['properties']:,}")
    print(f"  Lead events:  {s['lead_events']:,}")
    print(f"  Needs review: {s['needs_review']:,} (PARTIAL/MISMATCH/blank match)")
    print(f"  By county:")
    for r in s["by_county"]:
        print(f"    - {r['county'] or '(blank)'}: {r['c']:,}")
    if s["repeats"]:
        print(f"  Repeat leads (same person, same property, >1 event):")
        for r in s["repeats"]:
            print(f"    - {r['first_name']} {r['last_name']} @ "
                  f"{r['property_address'][:45]}  x{r['n']}")
    else:
        print("  Repeat leads: none")


def main():
    if len(sys.argv) < 2:
        print("Usage:")
        print("  python propintel_db_v0_1_0.py init")
        print("  python propintel_db_v0_1_0.py import <file.csv>")
        print("  python propintel_db_v0_1_0.py stats")
        print("  python propintel_db_v0_1_0.py reanalyze   (re-run ownership reasoning)")
        print("  python propintel_db_v0_1_0.py review      (list leads needing review)")
        print("  python propintel_db_v0_1_0.py export <file.xlsx>")
        sys.exit(1)

    cmd = sys.argv[1].lower()
    conn = connect()

    if cmd == "init":
        init_db(conn)
        print(f"Database initialized (schema v{SCHEMA_VERSION}) at "
              f"{os.path.join(os.path.dirname(os.path.abspath(__file__)), DEFAULT_DB_FILENAME)}")

    elif cmd == "import":
        if len(sys.argv) < 3:
            print("[ERROR] import needs a CSV path")
            sys.exit(1)
        init_db(conn)
        csv_path = sys.argv[2]
        if not os.path.isabs(csv_path):
            csv_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), csv_path)
        if not os.path.exists(csv_path):
            print(f"[ERROR] CSV not found: {csv_path}")
            sys.exit(1)
        print(f"Importing {csv_path} ...")
        s = import_from_csv(conn, csv_path)
        print(f"  CSV rows read:          {s['csv_rows']:,}")
        print(f"  Unique properties:      {s['properties_new']:,} new")
        print(f"  Unique contacts:        {s['contacts_new']:,} new")
        print(f"  Lead events recorded:   {s['lead_events_new']:,}")
        print(f"  Same-day dupes collapsed: {s['lead_events_collapsed']:,}")
        print(f"  Owner-only rows (no lead): {s['rows_without_lead']:,}")
        print()
        _print_stats(stats_summary(conn))

    elif cmd == "stats":
        _print_stats(stats_summary(conn))

    elif cmd == "reanalyze":
        init_db(conn)
        summary = reanalyze_matches(conn)
        total = sum(summary.values())
        print(f"Re-analyzed {total:,} lead event(s) with the ownership reasoner:")
        for status in ("CONFIRMED", "LIKELY", "REVIEW", "MISMATCH", "NO LEAD DATA"):
            if summary.get(status):
                print(f"  {status:13s} {summary[status]:,}")
        for status, n in summary.items():
            if status not in ("CONFIRMED", "LIKELY", "REVIEW", "MISMATCH", "NO LEAD DATA"):
                print(f"  {status:13s} {n:,}")

    elif cmd == "review":
        rows = review_queue(conn)
        if not rows:
            print("Nothing needs review. Every lead is a confident match.")
        else:
            print(f"{len(rows)} lead(s) needing review (highest confidence first):\n")
            for r in rows:
                owner = f"{r['owner1_first']} {r['owner1_last']}".strip()
                print(f"  [{r['ownership_match']:8s} {r['match_confidence']:>3}] "
                      f"{r['first_name']} {r['last_name']}  ->  owner: {owner or '(none)'}")
                print(f"        {r['property_address'][:60]}")
                print(f"        {r['match_reason']}")
                print()

    elif cmd == "export":
        if len(sys.argv) < 3:
            print("[ERROR] export needs an output .xlsx path")
            sys.exit(1)
        out = sys.argv[2]
        if not os.path.isabs(out):
            out = os.path.join(os.path.dirname(os.path.abspath(__file__)), out)
        n = export_xlsx(conn, out)
        print(f"Exported {n:,} rows to {out}")

    else:
        print(f"[ERROR] Unknown command: {cmd}")
        sys.exit(1)

    conn.close()


if __name__ == "__main__":
    main()
