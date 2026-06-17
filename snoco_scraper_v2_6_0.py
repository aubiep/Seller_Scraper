"""
Property Tax Record Scraper  v2.6.0
====================================
Automated property data extraction for Snohomish County AND King County.
County is auto-detected from the address city name.

v2.6.0 Changes:
    - NEW: the SQLite database propintel.db is now the system of record. Each
      scraped record is written through propintel_db_v0_1_0.ingest_record(),
      which dedups contacts (by email/name) and properties (by county+parcel),
      records a lead event per inbound lead, and flags ambiguous ownership
      matches (PARTIAL/MISMATCH) for review. Re-running a lead updates the
      property in place instead of appending a duplicate row.
    - The CSV-append output model is retired. After each run a spreadsheet VIEW
      is regenerated from the DB to propintel_export.xlsx (export, not store).
      The old snoco_property_data.csv/.xlsx are left untouched as historical
      files and can be deleted once you are comfortable with the DB.
    - The Excel-open lock failure is no longer fatal: the DB is always saved;
      only the optional spreadsheet view is skipped if the file is open.
    - Requires propintel_db_v0_1_0.py in the same folder. No change to any
      county lookup, parsing, or matching logic; this release only changes
      where results are written.

  - Snohomish County: live web scrape of the assessor site (unchanged).
  - King County: lookup from the King County bulk data ZIP files on disk.
    KC web scraping was removed from the routing path in v2.3.0 because it
    never held a stable connection. The bulk-data approach (formerly the
    standalone kc_lookup tool) is now built in. As of v2.5.0 the parsed
    indexes are cached to disk so subsequent runs skip the slow CSV rebuild.

Accepts leads from Market Leader CRM, tab-separated name+address lists,
or plain street addresses. Outputs to a persistent Excel/CSV spreadsheet
with photos saved to a property_photos folder.

v2.5.3 Changes:
    - FIXED: Krahl-style parcels (152407-9196) wrote blank Property Zip
      because the Residential Building.ZipCode field is empty AND the
      Address field has no trailing zip to strip. The May 20 v2.5.2 run
      surfaced this on the same parcel that originally exposed the city
      bug in v2.5.1.
    - kc_build_data() now extends the v2.5.2 owner-occupancy fallback to
      recover the zip too. When BOTH KC zip sources are blank and the
      billing street matches the property street, the zip is parsed out
      of the billing address.
    - New helper kc_parse_zip_from_billing(): pulls the trailing 5-digit
      zip from a billing string. Mirrors the city helper's gate semantics.
    - The fallback's billing-match check is now computed once and shared
      between the city and zip recovery paths.
    - No cache schema change. KC_CACHE_VERSION stays at 1. Existing
      v2.5.0/v2.5.1/v2.5.2 caches remain valid.

v2.5.2 Changes:
    - FIXED: King County rows whose Parcel.DistrictName was blank or
      "King County" still wrote a blank Property City even though the
      tax billing address clearly carried the city (Woodinville, Redmond,
      Fall City all seen in the May 20 v2.5.1 test run).
    - kc_build_data() now falls back to parsing the city out of the tax
      billing address when DistrictName is unusable. The fallback only
      fires for owner-occupied parcels (billing street matches property
      street) because for absentee owners the billing city is the OWNER'S
      mailing city, not the property's city.
    - New helper kc_parse_city_from_billing(): parses
      "STREET, CITY STATE ZIP" -> "City" (title-cased), handles multi-word
      cities like "Fall City", tolerates extra spaces.
    - No cache schema change. KC_CACHE_VERSION stays at 1. Existing
      v2.5.0/v2.5.1 caches remain valid; the fix is purely downstream of
      the cached data.

v2.5.1 Changes:
    - FIXED: King County output rows were missing the Property City, and
      sometimes Property Zip and Property State. The KC Residential Building
      file's Address field has no city, and one parcel in the May 20 test
      run (152407-9196, 32732 SE 44TH ST) had no zip in the address string
      either, which made parse_full_address fall through to its "no zip"
      branch and leave State and Zip blank too.
    - kc_build_data() now builds the property address from three KC fields:
      ResBldg.Address (street, with any trailing zip stripped), ResBldg.ZipCode
      (separate, more reliable zip), and Parcel.DistrictName (the KC field
      for the property's municipal jurisdiction). DistrictName = "King County"
      is the unincorporated marker; in that case the city is left blank
      rather than literally "King County" being written into the City column.
    - Property State is now always "WA" for KC rows when we have any
      address content at all (previously could be blank).
    - parse_full_address() is no longer called for KC rows; the parsed
      components are set directly inside kc_build_data() because the source
      data has the pieces separately and rebuilding them through a string
      parser was the cause of the missing-city bug.

v2.5.0 Changes:
    - NEW: King County indexes are cached to disk after the first load.
      First run after a fresh ZIP download takes 3-8 minutes (unchanged); the
      script then writes the parsed indexes to `kc_indexes_cache.pkl` next to
      the ZIPs. Every run after that loads the pickle in 5-10 seconds instead
      of re-parsing 2.6M rows of CSV.
    - Cache invalidates automatically when any KC ZIP's mtime is newer than
      the cache file's mtime, so re-downloading fresh KC data triggers a
      one-time rebuild and re-cache.
    - Cache pickle has an internal version stamp (KC_CACHE_VERSION). Future
      schema changes to the indexes will bump this and invalidate old caches
      automatically instead of producing wrong results.
    - Atomic write via `tmp` + `os.replace()`: a crash mid-save leaves the
      old cache intact, not a half-written file.
    - Corrupt or unreadable cache: warned and rebuilt from ZIPs, not fatal.
    - New functions: _kc_cache_path(), _kc_cache_is_fresh(),
      _kc_load_from_cache(), _kc_save_to_cache(). kc_load_all() rewired to
      check the cache before loading from ZIPs.

v2.4.2 Changes:
    - FIXED: Bothell addresses no longer route blindly to King County.
      Bothell straddles the King/Snohomish line and the previous version
      routed every Bothell address to KC. Snohomish-side Bothell parcels
      then failed the residential-only KC bulk index with no fallback. The
      v2.4.1 test run confirmed this (19825 30th Dr SE, Bothell, a real
      Snohomish parcel, reported as no-match).
    - New SPLIT_COUNTY_CITIES set holds cities that straddle the line.
      Currently just Bothell. Split-city addresses default to Snohomish in
      detect_county() and now retry on the other county if the primary
      lookup returns no data.
    - New helpers: _extract_city_segment() (shared parsing between
      detect_county and is_split_city) and is_split_city().
    - main() now loads both county data sources when any address is in a
      split-county city, so the fallback retry has something to fall back to.

v2.4.1 Changes:
    - FIXED (critical): King County bulk lookup could silently return the
      wrong property. kc_find_address()'s third matching tier did a bare
      substring test (`search_street in addr`), so a search for
      '19825 30TH DR SE' matched the KC record '19825 330TH AVE NE'
      ('30TH' is a substring of '330TH'). It then returned that lone match
      with no further checks: wrong street, wrong city, wrong zip, no
      warning. The KC equivalent of the v2.4.0 Snohomish wildcard bug.
      The third tier now matches by component (house number and street
      name required; directional required when the search has one; street
      type checked when both sides have one) and, when a zip is available
      on both sides, rejects candidates whose zip disagrees. If filtering
      leaves no single credible match the address is skipped and reported
      instead of guessed.
    - New functions: kc_address_components(), kc_components_match(),
      _kc_extract_zip(). kc_find_address() now takes pin_to_building (for
      the candidate-zip guard) and the raw search address (the only copy
      that still carries the search-side zip).

v2.4.0 Changes:
    - FIXED (critical): file-input county routing. Both the Market Leader
      path and the tab-separated path used to strip the city off the address
      before storing it, so detect_county() saw no city and defaulted every
      file-input address to Snohomish. King County addresses from a file
      never reached the King County bulk path. Lookups now store the RAW
      address (with city); scrape_property() and lookup_property_kc() each
      clean the address themselves before searching.
    - FIXED: detect_county() no longer misroutes on street names. It now
      matches the city in its actual position in the address (the segment
      before the state/zip, or the last segment) instead of a bare word
      search anywhere in the string. Previously '100 Kent Pl, Everett' would
      match 'kent' and misroute to King County.
    - FIXED: wildcard retry no longer grabs the wrong property. When an exact
      Snohomish search fails and the wildcard retry returns results, the
      result's street name must now match the input street name. If nothing
      matches, the address is skipped and reported instead of silently
      returning data for a different house. A truncated input with no street
      name (e.g. '19815') fails cleanly instead of matching.

v2.3.0 Changes:
    - King County now sources data from the five bulk data ZIP files on disk
      instead of web scraping. KC ZIPs load once at startup.
    - KC web-scrape functions (get_kc_session, search_address_kc,
      parse_kc_search_results, parse_kc_dashboard, scrape_property_kc) are
      retained in the file but no longer called. Kept in case KC server
      access ever becomes viable.
    - KC owner names are pulled from the Market Leader lead paste. The KC
      public bulk file has no taxpayer names. A bare KC address with no lead
      name produces blank owner fields and no Ownership Match.
    - KC Ownership Match is lead-based: CONFIRMED if the tax billing address
      matches the property address (owner-occupied), PARTIAL if it differs
      (absentee owner), NO LEAD DATA if no lead name was provided.
    - COLUMN_ORDER is unchanged. 15 columns the KC bulk data cannot fill are
      left blank for KC rows (Property Description, Taxpayer Name, Structure
      Description/Type, Neighborhood Code, Annual Tax Amount, Property
      Category, Status, Tax Code Area, Floor Details, Fireplace, Foundation,
      Exterior, Roof Type, Assessor Photo File).

v2.2.1 Changes:
    - Fixed King County search: added checkbox_acknowledge field for terms acceptance

v2.2 Changes:
    - Merged all features from both v2.1 branches
    - King County support with auto-detection (from v2.1a)
    - Property Grade, Condition, Views, Waterfront columns (from v2.1a)
    - Ownership Match column: CONFIRMED/PARTIAL/MISMATCH/NO LEAD DATA (from v2.1b)
    - Lead Type column: defaults to "Seller" for Market Leader leads (from v2.1b)
    - Parsed address columns: Property Street/City/State/Zip (from v2.1b)
    - IMPORTANT: Delete old snoco_property_data.xlsx and .csv before first run

Usage:
    python snoco_scraper_v2_5_1.py "13533 Boulder Ridge Rd"
    python snoco_scraper_v2_5_1.py "7214 204th Dr NE Redmond"
    python snoco_scraper_v2_5_1.py "123 Main St" "456 Oak Ave"
    python snoco_scraper_v2_5_1.py Seller_Leads.txt          (Market Leader paste)
    python snoco_scraper_v2_5_1.py addresses.txt             (Name+Address or plain addresses)

Requirements:
    pip install requests beautifulsoup4 openpyxl

Configuration:
    config.txt  -  Google Maps API key (one line, for Street View photos)

King County bulk data (place these ZIPs in the script folder):
    Residential Building.zip   -  addresses, beds, baths, sqft, year built
    Parcel.zip                 -  lot size, plat name, views, waterfront
    Real Property Account.zip  -  tax billing address (no taxpayer names)
    Real Property Sales.zip    -  sale date, price, grantor/grantee
    Value History.zip          -  assessed values by year
    Download from: https://info.kingcounty.gov/assessor/DataDownload/default.aspx

Output:
    snoco_property_data.xlsx  -  Persistent spreadsheet (appended each run)
    snoco_property_data.csv   -  Persistent CSV backup
    property_photos/          -  Street View and Assessor photos
"""

import requests, re, sys, csv, time, os, zipfile
from bs4 import BeautifulSoup
from urllib.parse import urljoin
from datetime import datetime
from io import TextIOWrapper


# ═══════════════════════════════════════════════════════════════════════════════
#  CONFIGURATION
# ═══════════════════════════════════════════════════════════════════════════════

BASE_URL    = "https://www.snoco.org/proptax/"
KC_BASE_URL = "https://blue.kingcounty.com/Assessor/eRealProperty/default.aspx"
OUTPUT_CSV  = "snoco_property_data.csv"
OUTPUT_XLSX = "snoco_property_data.xlsx"

# --- King County bulk data files ---------------------------------------------
# As of v2.3.0, King County data comes from these five bulk ZIP files on disk,
# not from web scraping. Download them from:
# https://info.kingcounty.gov/assessor/DataDownload/default.aspx
# Place all five in the same folder as this script.
KC_RESBLDG_ZIP = "Residential Building.zip"
KC_PARCEL_ZIP  = "Parcel.zip"
KC_RPACCT_ZIP  = "Real Property Account.zip"
KC_SALES_ZIP   = "Real Property Sales.zip"
KC_VALUE_ZIP   = "Value History.zip"

# CSV file names inside the ZIPs
KC_RESBLDG_CSV = "EXTR_ResBldg.csv"
KC_PARCEL_CSV  = "EXTR_Parcel.csv"
KC_RPACCT_CSV  = "EXTR_RPAcct_NoName.csv"  # public version has no taxpayer names
KC_SALES_CSV   = "EXTR_RPSale.csv"
KC_VALUE_CSV   = "EXTR_ValueHistory.csv"

# Sanity check: warn if a loader skips more than this fraction of rows.
# Catches silent schema changes (e.g. King County renames a column).
KC_ROW_LOAD_WARN_THRESHOLD = 0.90  # warn if < 90% of rows load

HTTP_HEADERS = {
    "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36",
    "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8",
}

COLUMN_ORDER = [
    # Metadata
    "Row #", "Date Retrieved", "County",
    # Lead Information
    "Lead First Name", "Lead Last Name", "Lead Email", "Lead Phone",
    "Lead Type", "Lead Source", "Years Since Sale", "Equity Level",
    # Property Address (full + parsed)
    "Property Address", "Property Street", "Property City", "Property State", "Property Zip",
    # Property Identification
    "Parcel Number", "Property Description", "Subdivision Name",
    # Owner Information + Match Status
    "Owner 1 First Name", "Owner 1 Last Name",
    "Owner 2 First Name", "Owner 2 Last Name",
    "Ownership Match",
    "Owner Address", "Taxpayer Name", "Tax Address",
    # Structure Details
    "Structure Description", "Structure Type", "Year Built",
    "Bedrooms", "Full or 3/4 Baths", "Half Baths", "Total Finished SF",
    "Size (gross)", "Unit of Measure",
    # Financial Data
    "Most Recent Sale Date", "Most Recent Sale Amount", "Neighborhood Code",
    "Latest Tax Year", "Assessed Value", "Market Total", "Market Land",
    "Market Improvement", "Taxable Value Regular", "Annual Tax Amount",
    # Additional Property Data
    "Use Code", "Property Category", "Status", "Tax Code Area",
    "Floor Details", "Garage SF", "Heat", "Fireplace", "Foundation",
    "Exterior", "Roof Type", "Property Grade", "Property Condition",
    "Views", "Waterfront", "Sales History",
    # Photos
    "Photo File", "Assessor Photo File",
]

KNOWN_CITIES = [
    "Snohomish", "Monroe", "Everett", "Marysville", "Lake Stevens",
    "Arlington", "Granite Falls", "Sultan", "Gold Bar", "Index",
    "Stanwood", "Darrington", "Edmonds", "Lynnwood", "Mountlake Terrace",
    "Mukilteo", "Mill Creek", "Bothell", "Woodinville", "Kenmore",
    "Brier", "Woodway", "Snoqualmie", "North Bend", "Carnation",
    "Fall City", "Redmond", "Kirkland", "Bellevue", "Duvall",
]

# Cities that are in King County (used to route to the KC scraper).
# Cities that straddle the line (e.g. Bothell) are NOT listed here; they are
# handled separately by SPLIT_COUNTY_CITIES below.
KING_COUNTY_CITIES = {
    "duvall", "fall city", "redmond", "kirkland", "bellevue",
    "woodinville", "kenmore", "sammamish", "issaquah",
    "renton", "kent", "auburn", "federal way", "burien", "tukwila",
    "seatac", "seattle", "mercer island", "newcastle", "black diamond",
    "maple valley", "covington", "enumclaw", "north bend", "snoqualmie",
    "carnation", "skykomish",
}

# Cities that straddle the King/Snohomish county line. detect_county() will
# default these to Snohomish (the majority of Aubie's leads in these cities sit
# on the Snohomish side, e.g. 98012/98021 north Bothell). The lookup loop will
# fall back to the other county if the primary lookup returns no match.
# v2.4.2: added to fix the Bothell routing bug where all Bothell addresses
# were being sent to King County and Snohomish-side Bothell parcels were
# failing the King County bulk lookup with no fallback.
SPLIT_COUNTY_CITIES = {"bothell"}

DIRECTION_MAP = {
    "Northwest": "NW", "Northeast": "NE", "Southwest": "SW", "Southeast": "SE",
    "North": "N", "South": "S", "East": "E", "West": "W",
}

STREET_TYPE_MAP = {
    "Drive": "Dr", "Street": "St", "Avenue": "Ave", "Boulevard": "Blvd",
    "Road": "Rd", "Lane": "Ln", "Court": "Ct", "Place": "Pl",
    "Circle": "Cir", "Terrace": "Ter", "Highway": "Hwy", "Parkway": "Pkwy",
    "Trail": "Trl",
}

SKIP_TITLE_CASE = {
    "Parcel Number", "Tax Code Area", "Use Code", "Neighborhood Code",
    "Row #", "Date Retrieved", "Photo File", "Assessor Photo File",
    "Sales History", "Floor Details", "Annual Tax Amount",
    "Assessed Value", "Market Total", "Market Land", "Market Improvement",
    "Taxable Value Regular", "Most Recent Sale Amount",
    "Latest Tax Year", "Year Built", "Size (gross)",
    "Bedrooms", "Full or 3/4 Baths", "Half Baths", "Total Finished SF",
    "Garage SF", "Ownership Match", "Lead Type", "County",
    "Property State", "Property Zip",
}

NAME_SUFFIXES = {"JR", "SR", "II", "III", "IV", "V", "VI", "VII", "VIII"}


# ═══════════════════════════════════════════════════════════════════════════════
#  TEXT FORMATTING
# ═══════════════════════════════════════════════════════════════════════════════

def smart_title_case(text):
    """Title Case preserving directional abbrevs, state codes, roman numerals."""
    if not text:
        return text
    if text.startswith("$") or re.match(r'^[\d/.$,%-]+$', text):
        return text
    if "/" in text or "\\" in text:
        return text

    preserve = {"SE", "NE", "NW", "SW", "N", "S", "E", "W",
                "WA", "OR", "CA", "ID", "MT", "PO"}
    roman = re.compile(r'^(I{1,3}|IV|VI{0,3}|IX|X{0,3})$')

    return " ".join(
        w if w in preserve or roman.match(w) else w.capitalize()
        for w in text.split()
    )


def apply_title_case(data):
    """Apply smart title case to all eligible text fields in a record."""
    for key, val in data.items():
        if key not in SKIP_TITLE_CASE and isinstance(val, str):
            data[key] = smart_title_case(val)
    return data


def clean_address_for_search(full_address):
    """Strip city/state/zip and abbreviate directions and street types
    so the assessor search can find the property."""
    addr = full_address.strip()
    if "," in addr:
        addr = addr.split(",")[0].strip()
    addr = re.sub(r'\s+(WA|OR|CA|ID|MT)\s*\d{0,5}.*$', '', addr, flags=re.I).strip()
    for city in sorted(KNOWN_CITIES, key=len, reverse=True):
        if addr.lower().endswith(city.lower()):
            addr = addr[:len(addr) - len(city)].strip()
            break
    for full, abbr in DIRECTION_MAP.items():
        addr = re.sub(r'\b' + full + r'\b', abbr, addr, flags=re.I)
    for full, abbr in STREET_TYPE_MAP.items():
        addr = re.sub(r'\b' + full + r'\b', abbr, addr, flags=re.I)
    return addr


def parse_full_address(full_address):
    """Parse 'Street, City, State Zip' into separate components.
    Handles both Snohomish format (with commas) and King County format (no commas).
    Returns dict with Property Street, Property City, Property State, Property Zip."""
    result = {
        "Property Street": "",
        "Property City": "",
        "Property State": "",
        "Property Zip": "",
    }
    if not full_address:
        return result
    
    # Check if comma-separated (Snohomish format): "9714 51ST AVE NE, MARYSVILLE, WA 98270"
    if "," in full_address:
        parts = [p.strip() for p in full_address.split(",")]
        
        if len(parts) >= 1:
            result["Property Street"] = parts[0]
        
        if len(parts) >= 2:
            result["Property City"] = parts[1]
        
        if len(parts) >= 3:
            # Last part should be "WA 98270" or "WA 98270-1234"
            state_zip = parts[2].strip()
            match = re.match(r'^([A-Z]{2})\s+(\d{5}(?:-\d{4})?)$', state_zip)
            if match:
                result["Property State"] = match.group(1)
                result["Property Zip"] = match.group(2)
            else:
                result["Property State"] = state_zip
    else:
        # King County format: "7214 204TH DR NE 98053" (no commas, zip at end)
        # Try to extract zip code from the end
        match = re.match(r'^(.+?)\s+(\d{5}(?:-\d{4})?)$', full_address.strip())
        if match:
            result["Property Street"] = match.group(1)
            result["Property Zip"] = match.group(2)
            result["Property State"] = "WA"  # Assume WA for King County
        else:
            # No zip found, just use the whole thing as street
            result["Property Street"] = full_address.strip()
    
    return result


# ═══════════════════════════════════════════════════════════════════════════════
#  OWNER NAME PARSING
# ═══════════════════════════════════════════════════════════════════════════════

def parse_owner_name(raw_name):
    """Parse county format LASTNAME FIRSTNAME [MIDDLE] [SUFFIX].
    Returns (first_name, last_name) with initials and suffixes stripped."""
    if not raw_name or not raw_name.strip():
        return "", ""
    parts = raw_name.strip().split()
    if len(parts) == 1:
        return parts[0], ""
    last_name = parts[0]
    cleaned = []
    for p in parts[1:]:
        u = p.upper().rstrip(".,")
        if len(u) == 1:
            continue
        if u in NAME_SUFFIXES:
            continue
        if len(u) == 2 and u.isalpha() and p.isupper() and len(parts) > 2:
            continue
        cleaned.append(p)
    return (" ".join(cleaned), last_name)


def split_owner_names(owner_string):
    """Split 'POUNCEY AUBIE E IV & KAREN R' into Owner 1/2 First/Last.
    Owner 2 inherits Owner 1's last name when only a first name is given."""
    result = {"Owner 1 First Name": "", "Owner 1 Last Name": "",
              "Owner 2 First Name": "", "Owner 2 Last Name": ""}
    if not owner_string:
        return result

    if " & " in owner_string:
        p = owner_string.split(" & ", 1)
        o1f, o1l = parse_owner_name(p[0].strip())
        o2f, o2l = parse_owner_name(p[1].strip())
        if not o2f and o2l:
            o2f, o2l = o2l, o1l
        elif o2f and not o2l:
            o2l = o1l
    else:
        o1f, o1l = parse_owner_name(owner_string)
        o2f, o2l = "", ""

    result["Owner 1 First Name"] = o1f
    result["Owner 1 Last Name"]  = o1l
    result["Owner 2 First Name"] = o2f
    result["Owner 2 Last Name"]  = o2l
    return result


def fix_owner_names_with_lead(data, lead_data):
    """Use lead name to detect/fix swapped owner name order."""
    if not lead_data:
        return data
    lf = lead_data.get("Lead First Name", "").upper().strip()
    ll = lead_data.get("Lead Last Name", "").upper().strip()
    o1f = data.get("Owner 1 First Name", "").upper().strip()
    o1l = data.get("Owner 1 Last Name", "").upper().strip()

    if (lf and o1l and lf == o1l) or (ll and o1f and ll == o1f):
        data["Owner 1 First Name"], data["Owner 1 Last Name"] = \
            data["Owner 1 Last Name"], data["Owner 1 First Name"]
        if data.get("Owner 2 Last Name", "").upper() == o1l:
            data["Owner 2 Last Name"] = data["Owner 1 Last Name"]
    return data


def compute_ownership_match(data, lead_data):
    """Compare lead name to owner names and return match status.
    
    Returns:
        CONFIRMED - Lead last name matches Owner 1 or Owner 2 last name, 
                    AND first names also match (or are close enough)
        PARTIAL - Last names match but first names don't (could be spouse, etc.)
        MISMATCH - Last names don't match (property may have sold)
        NO LEAD DATA - No lead name was provided
    """
    if not lead_data:
        return "NO LEAD DATA"
    
    lead_first = lead_data.get("Lead First Name", "").upper().strip()
    lead_last = lead_data.get("Lead Last Name", "").upper().strip()
    
    if not lead_last:
        return "NO LEAD DATA"
    
    o1_first = data.get("Owner 1 First Name", "").upper().strip()
    o1_last = data.get("Owner 1 Last Name", "").upper().strip()
    o2_first = data.get("Owner 2 First Name", "").upper().strip()
    o2_last = data.get("Owner 2 Last Name", "").upper().strip()
    
    # Check if lead last name matches either owner's last name
    last_name_match = (lead_last == o1_last) or (lead_last == o2_last)
    
    if not last_name_match:
        return "MISMATCH"
    
    # Last name matched. Now check first names.
    # Match if lead first name matches owner 1 OR owner 2 first name
    # Also match if lead first is a prefix of owner first (Ed matches Edward)
    # or owner first is a prefix of lead first
    def first_name_match(lead_f, owner_f):
        if not lead_f or not owner_f:
            return False
        if lead_f == owner_f:
            return True
        # Check if one is prefix of the other (handles Ed/Edward, Mike/Michael, etc.)
        if lead_f.startswith(owner_f) or owner_f.startswith(lead_f):
            return True
        return False
    
    if first_name_match(lead_first, o1_first) or first_name_match(lead_first, o2_first):
        return "CONFIRMED"
    
    # Last name matched but first name didn't
    return "PARTIAL"


# ═══════════════════════════════════════════════════════════════════════════════
#  LEAD FILE PARSING
# ═══════════════════════════════════════════════════════════════════════════════

MARKET_LEADER_KEYWORDS = {"Source:", "Since Sale", "Address Insights", "Assigned To:"}

ML_SKIP_WORDS = {
    "Edit", "mode_edit", "Seller", "Lead", "Source:", "email", "phone",
    "note", "Add Note", "Action Center", "Send Listings", "Nurture Zone",
    "Website Activity", "favorite", "Favorites", "Views", "trending_up",
    "Address Insights", "0", "$-",
}


def parse_lead_file(filepath):
    """Parse one or more Market Leader lead pastes from a text file."""
    with open(filepath, "r", encoding="utf-8") as f:
        content = f.read()

    if re.search(r'\n\s*-{3,}\s*\n', content):
        blocks = re.split(r'\n\s*-{3,}\s*\n', content)
    else:
        lines = content.split("\n")
        blocks, current = [], []
        for i, raw in enumerate(lines):
            line = raw.strip()
            is_new = (
                current and line and line not in ML_SKIP_WORDS
                and i + 2 < len(lines)
                and lines[i+1].strip() in ("Edit", "Editmode_edit")
                and lines[i+2].strip() in ("mode_edit", "Seller")
            )
            if is_new:
                blocks.append("\n".join(current))
                current = [raw]
            else:
                current.append(raw)
        if current:
            blocks.append("\n".join(current))

    leads = []
    for b in blocks:
        b = b.strip()
        if b:
            lead = _parse_single_ml_lead(b)
            if lead.get("address"):
                leads.append(lead)
    return leads


def _parse_single_ml_lead(text):
    """Parse a single Market Leader lead block."""
    lead = dict(name="", first_name="", last_name="", email="", phone="",
                source="", address="", years_since_sale="", equity_level="",
                lead_type="Seller")  # Market Leader leads are sellers
    lines = [l.strip() for l in text.strip().split("\n") if l.strip()]
    if not lines:
        return lead

    name = re.sub(r'\s*(Edit)?mode_edit.*$', '', lines[0]).strip()
    name = re.sub(r'\s*Edit\s*$', '', name).strip()
    lead["name"] = name
    parts = name.split()
    lead["first_name"] = parts[0] if parts else ""
    lead["last_name"] = " ".join(parts[1:]) if len(parts) > 1 else ""

    for i, line in enumerate(lines):
        nxt = lines[i+1] if i+1 < len(lines) else ""

        if line == "Source:":
            lead["source"] = nxt
        if line.lower() == "email":
            m = re.search(r'[\w.\-+]+@[\w.\-]+\.\w+', nxt)
            if m: lead["email"] = m.group(0)
        elif "@" in line and not lead["email"]:
            m = re.search(r'[\w.\-+]+@[\w.\-]+\.\w+', line)
            if m: lead["email"] = m.group(0)
        if line.lower() == "phone":
            m = re.search(r'(\d{3}[-.\s]?\d{3}[-.\s]?\d{4})', nxt)
            if m: lead["phone"] = m.group(1)
        elif re.match(r'^\d{3}[-.\s]\d{3}[-.\s]\d{4}$', line) and not lead["phone"]:
            lead["phone"] = line
        if "years since sale" in line.lower():
            lead["years_since_sale"] = line
        if "equity" in line.lower():
            lead["equity_level"] = line
        if line == "Address Insights":
            street = nxt
            city = lines[i+2] if i+2 < len(lines) else ""
            if re.search(r'(WA|OR|CA|ID|MT)\s*\d{0,5}', city, re.I):
                lead["address"] = f"{street} {city}"
            else:
                lead["address"] = street

    if not lead["address"]:
        for i, line in enumerate(lines):
            if re.match(r'^\d+\s+\w', line):
                skip = ["source", "email", "phone", "note", "assigned", "hours ago"]
                if not any(kw in line.lower() for kw in skip):
                    nxt = lines[i+1] if i+1 < len(lines) else ""
                    if re.search(r'(WA|OR|CA|ID|MT)', nxt, re.I):
                        lead["address"] = f"{line} {nxt}"
                    else:
                        lead["address"] = line
                    break
    return lead


def parse_tab_separated(content):
    """Parse 'Name\\tAddress' lines into lookup tuples."""
    lookups = []
    for line in content.split("\n"):
        line = line.strip()
        if not line or "\t" not in line:
            continue
        parts = line.split("\t", 1)
        if len(parts) != 2:
            continue
        name, addr = parts[0].strip(), parts[1].strip()
        np = name.split()
        first = np[0] if np else ""
        last = " ".join(np[1:]) if len(np) > 1 else ""
        lead_data = {
            "Lead First Name": smart_title_case(first),
            "Lead Last Name": smart_title_case(last),
            "Lead Email": "", "Lead Phone": "", "Lead Source": "",
            "Lead Type": "",  # Unknown for tab-separated input
            "Years Since Sale": "", "Equity Level": "",
        }
        # Store the RAW address (with city). detect_county() must see the city
        # to route correctly. The county-specific lookup functions
        # (scrape_property / lookup_property_kc) clean it themselves.
        lookups.append((addr, lead_data))
        print(f"  Lead: {first} {last} | {addr}")
    return lookups


def build_lead_data(lead):
    """Convert a parsed lead dict into the lead_data dict for merging."""
    return {
        "Lead First Name": smart_title_case(lead.get("first_name", "")),
        "Lead Last Name":  smart_title_case(lead.get("last_name", "")),
        "Lead Email":      lead.get("email", ""),
        "Lead Phone":      lead.get("phone", ""),
        "Lead Source":     lead.get("source", ""),
        "Lead Type":       lead.get("lead_type", ""),
        "Years Since Sale": lead.get("years_since_sale", ""),
        "Equity Level":    lead.get("equity_level", ""),
    }


# ═══════════════════════════════════════════════════════════════════════════════
#  ASSESSOR WEBSITE INTERACTION (Snohomish County)
# ═══════════════════════════════════════════════════════════════════════════════

def get_session():
    """Open a session to the Snohomish County assessor site."""
    session = requests.Session()
    session.headers.update(HTTP_HEADERS)
    resp = session.get(BASE_URL, allow_redirects=True)
    resp.raise_for_status()
    return session, resp


def _extract_asp_fields(html):
    """Pull ASP.NET hidden form fields for POST."""
    soup = BeautifulSoup(html, "html.parser")
    return {
        name: tag.get("value", "")
        for name in ("__VIEWSTATE", "__VIEWSTATEGENERATOR",
                      "__EVENTVALIDATION", "__EVENTTARGET", "__EVENTARGUMENT")
        if (tag := soup.find("input", {"name": name}))
    }


def search_address(session, address, base_resp):
    """Submit an address search and return the response."""
    fields = _extract_asp_fields(base_resp.text)
    form = {**fields, "mParcelID": "", "mStreetAddress": address,
            "mCity": "", "mStateProvince": "", "mPostalCode": "",
            "mSubmit": "Parcel Info"}
    resp = session.post(base_resp.url, data=form, allow_redirects=True)
    resp.raise_for_status()
    return resp


def parse_search_results(resp):
    """Extract parcel links from the search results page."""
    soup = BeautifulSoup(resp.text, "html.parser")
    if "ParcelInfo.aspx" in resp.url or "Property Account Summary" in resp.text:
        return [("DIRECT", resp)]
    results = []
    for link in soup.find_all("a"):
        href = link.get("href", "")
        if "ParcelInfo" in href and "parcel_number" in href:
            parcel = link.text.strip()
            row = link.find_parent("tr")
            addr = ""
            if row:
                cells = row.find_all("td")
                if len(cells) >= 2:
                    addr = cells[1].text.strip()
            results.append((parcel, addr, urljoin(resp.url, href)))
    return results


def _find_table(tables, *keywords):
    """Find the smallest table whose text contains ALL keywords."""
    matches = [t for t in tables if all(kw in t.text for kw in keywords)]
    return min(matches, key=lambda t: len(t.text)) if matches else None


# ═══════════════════════════════════════════════════════════════════════════════
#  PARCEL PAGE PARSING (Snohomish County)
# ═══════════════════════════════════════════════════════════════════════════════

def parse_parcel_page(html, url):
    """Extract all property data from ParcelInfo.aspx.
    Returns (data_dict, structure_detail_url)."""
    soup = BeautifulSoup(html, "html.parser")
    data = {}
    tables = soup.find_all("table")

    # --- Parcel Number & Address ---
    t = soup.find("table", class_="OutputTable")
    if t:
        cells = [c.text.strip() for c in t.find_all("td")]
        for i, c in enumerate(cells):
            if c == "Parcel Number" and i+1 < len(cells):
                data["Parcel Number"] = cells[i+1]
            if c == "Property Address" and i+1 < len(cells):
                data["Property Address"] = cells[i+1]

    # --- Parse full address into components ---
    if data.get("Property Address"):
        data.update(parse_full_address(data["Property Address"]))

    # --- General Information ---
    t = _find_table(tables, "Property Description", "Property Category", "Tax Code Area")
    if t:
        for row in t.find_all("tr"):
            cells = row.find_all("td")
            if len(cells) == 2:
                lbl, val = cells[0].text.strip(), cells[1].text.strip()
                if lbl in ("Property Description", "Property Category", "Status", "Tax Code Area"):
                    data[lbl] = val

    # --- Subdivision Name ---
    desc = data.get("Property Description", "")
    if desc and not desc.upper().startswith("SEC "):
        parts = re.split(
            r'\s+[Bb][Ll][Kk]\s+|\s+-\s+[Ll][Oo][Tt]\s+|\s+[Ll][Oo][Tt]\s+\d|\s+-\s+[Uu]nit\s+',
            desc, maxsplit=1)
        sub = parts[0].strip() if parts else ""
        data["Subdivision Name"] = re.sub(r',\s*Corrected Plat Of\s*$', '', sub, flags=re.I).strip()
    else:
        data["Subdivision Name"] = ""

    # --- Property Characteristics ---
    t = _find_table(tables, "Use Code", "Unit of Measure", "Size (gross)")
    if t:
        for row in t.find_all("tr"):
            cells = row.find_all("td")
            if len(cells) == 2:
                lbl, val = cells[0].text.strip(), cells[1].text.strip()
                if lbl in ("Use Code", "Unit of Measure", "Size (gross)"):
                    data[lbl] = val

    # --- Owners & Taxpayers ---
    t = _find_table(tables, "Role", "Percent", "Name", "Address")
    if t:
        owners, taxpayers = [], []
        for row in t.find_all("tr")[1:]:
            cells = row.find_all("td")
            if len(cells) >= 4:
                role = cells[0].text.strip()
                name, addr = cells[2].text.strip(), cells[3].text.strip()
                if role == "Owner":
                    owners.append((name, addr))
                elif role == "Taxpayer":
                    taxpayers.append((name, addr))
        if owners:
            data.update(split_owner_names(owners[0][0]))
            data["Owner Address"] = owners[0][1]
        else:
            data.update(split_owner_names(""))
            data["Owner Address"] = ""
        data["Taxpayer Name"] = taxpayers[0][0] if taxpayers else ""
        data["Tax Address"]   = taxpayers[0][1] if taxpayers else ""

    # --- Property Values ---
    t = _find_table(tables, "Value Type", "Tax Year", "Taxable Value Regular")
    if t:
        rows = t.find_all("tr")
        if rows:
            for hc in rows[0].find_all("td")[1:]:
                m = re.search(r'(\d{4})', hc.text)
                if m:
                    data["Latest Tax Year"] = m.group(1)
                    break
            for row in rows[1:]:
                cells = row.find_all("td")
                if len(cells) >= 2:
                    lbl, val = cells[0].text.strip(), cells[1].text.strip()
                    if lbl in ("Taxable Value Regular", "Market Total",
                               "Assessed Value", "Market Land", "Market Improvement"):
                        data[lbl] = val

    # --- Annual Tax ---
    t = _find_table(tables, "Tax Year", "Installment", "Due Date", "Principal")
    if t:
        total = 0
        for row in t.find_all("tr")[1:]:
            for c in row.find_all("td"):
                ct = c.text.strip()
                if ct.startswith("$"):
                    try:
                        total += float(ct.replace("$", "").replace(",", ""))
                        break
                    except ValueError:
                        pass
        data["Annual Tax Amount"] = f"${total:,.2f}" if total > 0 else ""

    # --- Structure Info Link ---
    struct_link = ""
    t = _find_table(tables, "Description", "Type", "Year Built", "More Information")
    if t:
        for row in t.find_all("tr")[1:]:
            cells = row.find_all("td")
            if len(cells) >= 3:
                data["Structure Description"] = cells[0].text.strip()
                data["Structure Type"]        = cells[1].text.strip()
                data["Year Built"]            = cells[2].text.strip()
                if len(cells) >= 4:
                    a = cells[3].find("a")
                    if a and a.get("href"):
                        h = a["href"]
                        struct_link = h if h.startswith("http") else urljoin(url, h)
                break

    # --- Sales History ---
    sales = []
    t = _find_table(tables, "Sale Date", "Sale Amount", "Grantor")
    if t:
        for row in t.find_all("tr")[1:]:
            cells = row.find_all("td")
            if len(cells) >= 8:
                sales.append({
                    "date": cells[0].text.strip(),
                    "amount": cells[3].text.strip(),
                    "deed": cells[5].text.strip(),
                    "grantor": cells[7].text.strip(),
                    "grantee": cells[8].text.strip() if len(cells) > 8 else "",
                })
    data["Most Recent Sale Date"] = ""
    data["Most Recent Sale Amount"] = ""
    for s in reversed(sales):
        try:
            if float(s["amount"].replace("$", "").replace(",", "").strip()) > 0:
                data["Most Recent Sale Date"]   = s["date"]
                data["Most Recent Sale Amount"] = s["amount"]
                break
        except ValueError:
            continue
    data["Sales History"] = "; ".join(
        f"{s['date']} - {s['amount']} ({s['deed']}) {s['grantor']} -> {s['grantee']}"
        for s in sales) if sales else ""

    # --- Neighborhood Code ---
    t = _find_table(tables, "Neighborhood Code", "Township", "Range")
    if t:
        for row in t.find_all("tr")[1:]:
            cells = row.find_all("td")
            if cells:
                data["Neighborhood Code"] = cells[0].text.strip()
                break

    return data, struct_link


# ═══════════════════════════════════════════════════════════════════════════════
#  STRUCTURE PAGE PARSING (Snohomish County)
# ═══════════════════════════════════════════════════════════════════════════════

def parse_structure_page(session, struct_url):
    """Fetch and parse the structure detail page for beds, baths, sqft, etc."""
    if not struct_url:
        return {}
    try:
        resp = session.get(struct_url, timeout=15)
        resp.raise_for_status()
    except Exception as e:
        print(f"  [WARNING] Could not fetch structure page: {e}")
        return {}

    soup = BeautifulSoup(resp.text, "html.parser")

    def field(label):
        for td in soup.find_all("td"):
            if td.text.strip() == label:
                nxt = td.find_next_sibling("td")
                return nxt.text.strip() if nxt else ""
        return ""

    d = {
        "Bedrooms": field("Bedrooms"),
        "Full or 3/4 Baths": field("Full or 3/4 Baths"),
        "Half Baths": "",
        "Heat": field("Heat"),
        "Fireplace": field("Fireplace"),
        "Foundation": field("Foundation"),
        "Exterior": field("Exterior"),
        "Roof Type": field("ROOF\u00a0\u00a0Type:"),
        "Garage SF": field("Attached Garage SF"),
    }

    for td in soup.find_all("td"):
        if td.text.strip() == "1/2 Baths":
            nxt = td.find_next_sibling("td")
            if nxt:
                d["Half Baths"] = nxt.text.strip()
            break

    # Floor areas
    total_sf, details = 0, []
    all_tds = soup.find_all("td")
    for i, td in enumerate(all_tds):
        if td.text.strip() == "Floor":
            ahead = [t.text.strip() for t in all_tds[i+1:i+7]]
            floor_num = ahead[0] if ahead else ""
            finished = ""
            for j, txt in enumerate(ahead):
                if txt == "Finished SF" and j+1 < len(ahead):
                    finished = ahead[j+1]
                    break
            if floor_num and finished:
                details.append(f"Floor {floor_num}: {finished} SF")
                try:
                    total_sf += int(finished)
                except ValueError:
                    pass
    d["Total Finished SF"] = str(total_sf) if total_sf > 0 else ""
    d["Floor Details"] = "; ".join(details) if details else ""

    # Assessor photo URL
    d["_assessor_photo_url"] = ""
    for img in soup.find_all("img"):
        src = img.get("src", "")
        if "photos" in src and ".jpg" in src.lower():
            d["_assessor_photo_url"] = src
            break

    return d


# ═══════════════════════════════════════════════════════════════════════════════
#  COUNTY DETECTION
# ═══════════════════════════════════════════════════════════════════════════════

def _extract_city_segment(address):
    """Return the lowercased city segment from a raw address, or '' if none.

    Shared by detect_county() and is_split_city() so both apply the same
    parsing rules to comma-separated and no-comma addresses.
    """
    if not address:
        return ""

    addr = address.strip()
    city_segment = ""

    if "," in addr:
        segments = [s.strip() for s in addr.split(",") if s.strip()]
        city_segment = segments[-1] if segments else ""
        for idx, seg in enumerate(segments):
            if re.match(r'^[A-Z]{2}(\s+\d{5}(-\d{4})?)?$', seg, flags=re.I) \
                    or re.match(r'^\d{5}(-\d{4})?$', seg):
                if idx > 0:
                    city_segment = segments[idx - 1]
                break
    else:
        stripped = re.sub(r'\s+[A-Z]{2}\s*\d{0,5}(-\d{4})?\s*$', '', addr, flags=re.I)
        stripped = re.sub(r'\s+\d{5}(-\d{4})?\s*$', '', stripped).strip()
        words = stripped.split()
        city_segment = " ".join(words[-2:]) if len(words) >= 2 else stripped

    return city_segment.lower().strip()


def is_split_city(address):
    """Return True if the address city is in SPLIT_COUNTY_CITIES.

    Split cities straddle the King/Snohomish line, so a no-match result on
    the primary county warrants a retry on the other county.
    """
    city_lower = _extract_city_segment(address)
    for city in SPLIT_COUNTY_CITIES:
        if city_lower == city:
            return True
        if re.search(r'\b' + re.escape(city) + r'$', city_lower):
            return True
    return False


def detect_county(address):
    """Return 'king' if the address city matches a King County city, else 'snohomish'.

    As of v2.4.0 this receives the RAW address (with city/state/zip), so the
    city can be matched in its actual position instead of by a bare word
    search. The old bare-\\b search misrouted addresses whose STREET name
    happened to contain a King County city word (e.g. '100 Kent Pl, Everett').

    As of v2.4.2, cities that straddle the county line (SPLIT_COUNTY_CITIES,
    currently just Bothell) are NOT in KING_COUNTY_CITIES and so default to
    Snohomish here. Callers should use is_split_city() to decide whether to
    retry the other county on a no-match.
    """
    if not address:
        return "snohomish"

    city_lower = _extract_city_segment(address)

    # Exact match on the city segment, then a relaxed "ends with city" check
    # for the no-comma fallback (handles a one-word-city tail).
    for city in KING_COUNTY_CITIES:
        if city_lower == city:
            return "king"
    for city in KING_COUNTY_CITIES:
        if re.search(r'\b' + re.escape(city) + r'$', city_lower):
            return "king"
    return "snohomish"


# ═══════════════════════════════════════════════════════════════════════════════
#  KING COUNTY BULK DATA  (v2.3.0 - replaces KC web scraping in the routing path)
# ═══════════════════════════════════════════════════════════════════════════════
#
#  King County property data is read from five bulk ZIP files on disk, not
#  scraped. The loaders below are adapted from the standalone kc_lookup tool.
#  They are called once at startup by main(); the resulting indexes are then
#  reused for every King County address in the run.
#
#  Owner names: the public KC bulk file (EXTR_RPAcct_NoName.csv) has names
#  stripped by King County. KC owner names therefore come from the Market
#  Leader lead paste, not from the bulk data. A bare KC address with no lead
#  name attached produces blank owner fields and no Ownership Match.
# ═══════════════════════════════════════════════════════════════════════════════

def kc_normalize_address(addr):
    """Normalize an address for matching against the KC bulk address index.
    Returns uppercase, single-spaced, with directions and street types
    abbreviated and any trailing zip code removed."""
    if not addr:
        return ""

    # Uppercase and compress whitespace
    addr = " ".join(addr.upper().split())

    # Remove zip code from the end if present
    addr = re.sub(r'\s+\d{5}(-\d{4})?$', '', addr)

    # Standardize directional prefixes/suffixes
    addr = re.sub(r'\bNORTHWEST\b', 'NW', addr)
    addr = re.sub(r'\bNORTHEAST\b', 'NE', addr)
    addr = re.sub(r'\bSOUTHWEST\b', 'SW', addr)
    addr = re.sub(r'\bSOUTHEAST\b', 'SE', addr)
    addr = re.sub(r'\bNORTH\b', 'N', addr)
    addr = re.sub(r'\bSOUTH\b', 'S', addr)
    addr = re.sub(r'\bEAST\b', 'E', addr)
    addr = re.sub(r'\bWEST\b', 'W', addr)

    # Standardize street types
    addr = re.sub(r'\bSTREET\b', 'ST', addr)
    addr = re.sub(r'\bAVENUE\b', 'AVE', addr)
    addr = re.sub(r'\bDRIVE\b', 'DR', addr)
    addr = re.sub(r'\bROAD\b', 'RD', addr)
    addr = re.sub(r'\bLANE\b', 'LN', addr)
    addr = re.sub(r'\bCOURT\b', 'CT', addr)
    addr = re.sub(r'\bPLACE\b', 'PL', addr)
    addr = re.sub(r'\bBOULEVARD\b', 'BLVD', addr)
    addr = re.sub(r'\bCIRCLE\b', 'CIR', addr)
    addr = re.sub(r'\bTERRACE\b', 'TER', addr)
    addr = re.sub(r'\bPARKWAY\b', 'PKWY', addr)
    addr = re.sub(r'\bHIGHWAY\b', 'HWY', addr)
    addr = re.sub(r'\bWAY\b', 'WY', addr)

    return addr


def kc_extract_street_number(addr):
    """Extract the leading street number from an address for partial matching."""
    match = re.match(r'^(\d+)', addr)
    return match.group(1) if match else ""


def kc_collapse_spaces(text):
    """Collapse runs of whitespace into single spaces and strip.
    KC pads its address field with extra spaces."""
    if not text:
        return text
    return " ".join(text.split())


def kc_parse_city_from_billing(billing_addr):
    """Extract just the city out of a billing address string of the form
    'STREET, CITY STATE ZIP'. Returns the city in title case, or '' if it
    can't be parsed.

    v2.5.2: used as a fallback when Parcel.DistrictName is missing or set
    to 'King County' (the unincorporated marker). Many KC parcels have a
    blank or generic DistrictName even when the billing address clearly
    carries the municipal city, so we recover the city from there.

    Callers should only use this for owner-occupied parcels (billing
    street matches property street). For absentee owners, the billing
    city is the owner's mailing city, NOT the property's city.

    Examples:
        '32732 SE 44TH ST, FALL CITY WA 98024' -> 'Fall City'
        '17131 163RD AVE NE, WOODINVILLE  WA 98072' -> 'Woodinville'
        '23939 NE ADAIR RD, REDMOND WA 98053' -> 'Redmond'
        '' -> ''
        'PO BOX 12345, SEATTLE WA 98101' -> 'Seattle'
    """
    if not billing_addr or ',' not in billing_addr:
        return ''
    # Take everything after the first comma (the city/state/zip portion).
    tail = billing_addr.split(',', 1)[1].strip()
    if not tail:
        return ''
    # Strip a trailing zip (5 or 5-4) if present.
    tail = re.sub(r'\s+\d{5}(?:-\d{4})?\s*$', '', tail).strip()
    # Strip a trailing state code. Either preceded by whitespace (normal
    # case: "REDMOND WA") or as the entire remainder ("WA" alone, when the
    # city was missing from the original billing address).
    tail = re.sub(r'(?:^|\s+)[A-Z]{2}\s*$', '', tail).strip()
    if not tail:
        return ''
    return tail.title()


def kc_parse_zip_from_billing(billing_addr):
    """Extract the trailing 5-digit zip from a billing address string of the
    form 'STREET, CITY STATE ZIP'. Returns the zip as a 5-digit string, or
    '' if no zip is present.

    v2.5.3: companion to kc_parse_city_from_billing(). Used as a fallback
    when both Residential Building.ZipCode AND the address-string zip are
    blank (parcel 152407-9196 in the May 20 test data was the trigger).

    Like the city fallback, callers should only use this for owner-occupied
    parcels. For absentee owners the billing zip is the owner's mailing zip,
    not the property's.

    Examples:
        '32732 SE 44TH ST, FALL CITY WA 98024' -> '98024'
        '123 MAIN ST, SEATTLE WA 98101-1234' -> '98101'
        '123 MAIN ST, SEATTLE WA' -> ''
        '' -> ''
    """
    if not billing_addr:
        return ''
    m = re.search(r'(\d{5})(?:-\d{4})?\s*$', billing_addr)
    return m.group(1) if m else ''


def kc_read_csv_from_zip(zip_path, csv_name):
    """Read a CSV directly from a ZIP archive without extracting.
    Returns a list of dictionaries (one per row)."""
    rows = []
    with zipfile.ZipFile(zip_path, 'r') as zf:
        csv_files = [f for f in zf.namelist() if f.lower().endswith('.csv')]

        # Try exact match first (allowing for subfolders / casing differences)
        target = None
        for f in csv_files:
            if f.lower() == csv_name.lower() or f.lower().endswith('/' + csv_name.lower()):
                target = f
                break

        # Fall back to the first CSV found
        if not target and csv_files:
            target = csv_files[0]
            print(f"  Note: Using {target} (expected {csv_name})")

        if not target:
            print(f"  [WARNING] No CSV found in {zip_path}")
            return rows

        with zf.open(target) as f:
            reader = csv.DictReader(TextIOWrapper(f, encoding='utf-8', errors='replace'))
            for row in reader:
                rows.append(row)

    return rows


def kc_sanity_check_load(loader_name, loaded_count, total_rows):
    """Warn if fewer than KC_ROW_LOAD_WARN_THRESHOLD of rows were loaded.
    Catches silent schema breaks (e.g., a renamed column making every row
    get skipped). Returns True if healthy, False if suspicious."""
    if total_rows == 0:
        print(f"  [WARNING] {loader_name}: source CSV had 0 rows")
        return False
    ratio = loaded_count / total_rows
    if ratio < KC_ROW_LOAD_WARN_THRESHOLD:
        pct = ratio * 100
        print(f"  [WARNING] {loader_name}: only loaded {loaded_count:,} of "
              f"{total_rows:,} rows ({pct:.1f}%). Possible column-name change "
              f"in the source file. Check the KC schema.")
        return False
    return True


def kc_load_resbldg_data(script_dir):
    """Load residential building data and build the address index.
    Returns (address_index, pin_to_building) where:
        - address_index maps normalized_address -> PIN
        - pin_to_building maps PIN -> building details dict
    """
    zip_path = os.path.join(script_dir, KC_RESBLDG_ZIP)
    if not os.path.exists(zip_path):
        print(f"  [WARNING] {KC_RESBLDG_ZIP} not found")
        return {}, {}

    print(f"Loading residential building data from {KC_RESBLDG_ZIP}...")
    rows = kc_read_csv_from_zip(zip_path, KC_RESBLDG_CSV)

    address_index = {}    # normalized_address -> PIN
    pin_to_building = {}  # PIN -> building data
    processed = 0         # rows that parsed OK (for sanity check)

    for row in rows:
        major = row.get('Major', '').strip()
        minor = row.get('Minor', '').strip()

        if not major or not minor:
            continue

        # Build PIN (10 digits: 6-digit major + 4-digit minor)
        pin = f"{major.zfill(6)}{minor.zfill(4)}"
        processed += 1

        address = row.get('Address', '').strip()
        if address:
            norm_addr = kc_normalize_address(address)
            if norm_addr:
                address_index[norm_addr] = pin

        pin_to_building[pin] = {
            'Address': address,
            'YrBuilt': row.get('YrBuilt', ''),
            'Stories': row.get('Stories', ''),
            'SqFtTotLiving': row.get('SqFtTotLiving', ''),
            'Bedrooms': row.get('Bedrooms', ''),
            'BathFullCount': row.get('BathFullCount', ''),
            'Bath3qtrCount': row.get('Bath3qtrCount', ''),
            'BathHalfCount': row.get('BathHalfCount', ''),
            'BldgGrade': row.get('BldgGrade', ''),
            'Condition': row.get('Condition', ''),
            'HeatSystem': row.get('HeatSystem', ''),
            'SqFtGarageAttached': row.get('SqFtGarageAttached', ''),
            'SqFtTotBasement': row.get('SqFtTotBasement', ''),
            'SqFtFinBasement': row.get('SqFtFinBasement', ''),
            'ViewUtilization': row.get('ViewUtilization', ''),
            'ZipCode': row.get('ZipCode', ''),
        }

    print(f"  Loaded {len(rows)} buildings, {len(address_index)} unique addresses")
    kc_sanity_check_load("Residential Building", processed, len(rows))
    return address_index, pin_to_building


def kc_load_parcel_data(script_dir):
    """Load parcel data. Returns dict mapping PIN -> parcel details."""
    zip_path = os.path.join(script_dir, KC_PARCEL_ZIP)
    if not os.path.exists(zip_path):
        print(f"  [WARNING] {KC_PARCEL_ZIP} not found")
        return {}

    print(f"Loading parcel data from {KC_PARCEL_ZIP}...")
    rows = kc_read_csv_from_zip(zip_path, KC_PARCEL_CSV)

    pin_to_parcel = {}
    for row in rows:
        major = row.get('Major', '').strip()
        minor = row.get('Minor', '').strip()

        if not major or not minor:
            continue

        pin = f"{major.zfill(6)}{minor.zfill(4)}"

        # Combine view flags into a readable string
        views = []
        if row.get('MtRainier', '0') not in ('0', '', 'N'):
            views.append('Mt Rainier')
        if row.get('Olympics', '0') not in ('0', '', 'N'):
            views.append('Olympics')
        if row.get('Cascades', '0') not in ('0', '', 'N'):
            views.append('Cascades')
        if row.get('Territorial', '0') not in ('0', '', 'N'):
            views.append('Territorial')
        if row.get('SeattleSkyline', '0') not in ('0', '', 'N'):
            views.append('Seattle Skyline')
        if row.get('PugetSound', '0') not in ('0', '', 'N'):
            views.append('Puget Sound')
        if row.get('LakeWashington', '0') not in ('0', '', 'N'):
            views.append('Lake Washington')
        if row.get('LakeSammamish', '0') not in ('0', '', 'N'):
            views.append('Lake Sammamish')

        wfnt = row.get('WfntFootage', '0')
        waterfront = "Yes" if wfnt and wfnt != '0' else "No"

        pin_to_parcel[pin] = {
            'PlatName': row.get('PlatName', ''),
            'SqFtLot': row.get('SqFtLot', ''),
            'DistrictName': row.get('DistrictName', ''),
            'LevyCode': row.get('LevyCode', ''),
            'CurrentZoning': row.get('CurrentZoning', ''),
            'PresentUse': row.get('PresentUse', ''),
            'Views': ', '.join(views) if views else '',
            'Waterfront': waterfront,
            'WfntFootage': wfnt,
        }

    print(f"  Loaded {len(pin_to_parcel)} parcels")
    kc_sanity_check_load("Parcel", len(pin_to_parcel), len(rows))
    return pin_to_parcel


def kc_load_rpacct_data(script_dir):
    """Load real property account data. Returns dict mapping PIN -> account
    details. This is where the tax billing address comes from."""
    zip_path = os.path.join(script_dir, KC_RPACCT_ZIP)
    if not os.path.exists(zip_path):
        print(f"  [WARNING] {KC_RPACCT_ZIP} not found")
        return {}

    print(f"Loading account data from {KC_RPACCT_ZIP}...")
    rows = kc_read_csv_from_zip(zip_path, KC_RPACCT_CSV)

    pin_to_acct = {}
    for row in rows:
        major = row.get('Major', '').strip()
        minor = row.get('Minor', '').strip()

        if not major or not minor:
            continue

        pin = f"{major.zfill(6)}{minor.zfill(4)}"

        # Build the billing address
        addr_line = row.get('AddrLine', '').strip()
        city_state = row.get('CityState', '').strip()
        zip_code = row.get('ZipCode', '').strip()

        billing_addr = addr_line
        if city_state:
            billing_addr += f", {city_state}"
        if zip_code:
            billing_addr += f" {zip_code}"

        pin_to_acct[pin] = {
            'BillingAddress': billing_addr,
            'TaxStat': row.get('TaxStat', ''),
            'ApprLandVal': row.get('ApprLandVal', ''),
            'ApprImpsVal': row.get('ApprImpsVal', ''),
            'TaxableLandVal': row.get('TaxableLandVal', ''),
            'TaxableImpsVal': row.get('TaxableImpsVal', ''),
        }

    print(f"  Loaded {len(pin_to_acct)} accounts")
    kc_sanity_check_load("Real Property Account", len(pin_to_acct), len(rows))
    return pin_to_acct


def kc_load_value_data(script_dir):
    """Load value history data. Returns dict mapping PIN -> most recent values."""
    zip_path = os.path.join(script_dir, KC_VALUE_ZIP)
    if not os.path.exists(zip_path):
        print(f"  [WARNING] {KC_VALUE_ZIP} not found")
        return {}

    print(f"Loading value history from {KC_VALUE_ZIP}...")
    rows = kc_read_csv_from_zip(zip_path, KC_VALUE_CSV)

    pin_to_value = {}
    processed = 0  # rows that parsed OK (we intentionally keep only the
                   # latest year per PIN, so len() would understate this)
    for row in rows:
        major = row.get('Major', '').strip()
        minor = row.get('Minor', '').strip()
        # NOTE: the KC schema column is 'TaxYr', not 'TaxYear'. Do not "fix" this.
        tax_year = row.get('TaxYr', '').strip()

        if not major or not minor or not tax_year:
            continue

        pin = f"{major.zfill(6)}{minor.zfill(4)}"

        try:
            year = int(tax_year)
        except ValueError:
            continue

        processed += 1

        # Keep only the most recent year per PIN
        if pin in pin_to_value:
            if year <= pin_to_value[pin].get('_year', 0):
                continue

        pin_to_value[pin] = {
            '_year': year,
            'TaxYear': tax_year,
            'ApprTotVal': row.get('ApprLandVal', ''),  # recomputed below
            'ApprLandVal': row.get('ApprLandVal', ''),
            'ApprImpsVal': row.get('ApprImpsVal', ''),
            'TaxableTotVal': row.get('LandVal', ''),   # recomputed below
            'LandVal': row.get('LandVal', ''),
            'ImpsVal': row.get('ImpsVal', ''),
        }

    # KC's value history stores land and improvements separately with no
    # pre-computed total. Add them here.
    for pin, v in pin_to_value.items():
        try:
            appr_total = int(v.get('ApprLandVal') or 0) + int(v.get('ApprImpsVal') or 0)
            v['ApprTotVal'] = str(appr_total) if appr_total else ''
        except (ValueError, TypeError):
            v['ApprTotVal'] = ''
        try:
            tax_total = int(v.get('LandVal') or 0) + int(v.get('ImpsVal') or 0)
            v['TaxableTotVal'] = str(tax_total) if tax_total else ''
        except (ValueError, TypeError):
            v['TaxableTotVal'] = ''
        v.pop('_year', None)

    print(f"  Loaded values for {len(pin_to_value):,} parcels (from {processed:,} rows)")
    kc_sanity_check_load("Value History", processed, len(rows))
    return pin_to_value


def kc_load_sales_data(script_dir):
    """Load real property sales history. Returns dict mapping PIN -> list of
    sale records (sorted newest first)."""
    zip_path = os.path.join(script_dir, KC_SALES_ZIP)
    if not os.path.exists(zip_path):
        print(f"  [WARNING] {KC_SALES_ZIP} not found")
        return {}

    print(f"Loading sales history from {KC_SALES_ZIP}...")
    rows = kc_read_csv_from_zip(zip_path, KC_SALES_CSV)

    pin_to_sales = {}
    loaded = 0
    for row in rows:
        major = row.get('Major', '').strip()
        minor = row.get('Minor', '').strip()

        if not major or not minor:
            continue

        pin = f"{major.zfill(6)}{minor.zfill(4)}"

        sale = {
            'DocumentDate': row.get('DocumentDate', '').strip(),
            'SalePrice':    row.get('SalePrice', '').strip(),
            'SellerName':   row.get('SellerName', '').strip(),
            'BuyerName':    row.get('BuyerName', '').strip(),
            'SaleInstrument': row.get('SaleInstrument', '').strip(),
            'SaleReason':   row.get('SaleReason', '').strip(),
            'SaleWarning':  row.get('SaleWarning', '').strip(),
            'ExciseTaxNbr': row.get('ExciseTaxNbr', '').strip(),
        }

        pin_to_sales.setdefault(pin, []).append(sale)
        loaded += 1

    # Sort each PIN's sales newest first (DocumentDate is MM/DD/YYYY)
    def _date_key(s):
        d = s.get('DocumentDate', '')
        m = re.match(r'(\d{1,2})/(\d{1,2})/(\d{4})', d)
        if m:
            mm, dd, yyyy = m.groups()
            return f"{yyyy}{mm.zfill(2)}{dd.zfill(2)}"
        return "00000000"

    for pin in pin_to_sales:
        pin_to_sales[pin].sort(key=_date_key, reverse=True)

    print(f"  Loaded {loaded:,} sales across {len(pin_to_sales):,} parcels")
    kc_sanity_check_load("Real Property Sales", loaded, len(rows))
    return pin_to_sales


# v2.5.0: cache the loaded KC indexes to a pickle next to the ZIPs.
# First run after a fresh ZIP download builds the indexes and saves them.
# Every run after that loads the pickle (5-10s) instead of re-parsing 2.6M
# rows of CSV (3-8 minutes). Cache invalidates automatically when any ZIP's
# mtime is newer than the pickle. Bump KC_CACHE_VERSION whenever the
# structure returned by kc_load_all() changes, so old pickles are rejected
# on the next run instead of silently producing the wrong fields.
KC_CACHE_VERSION = 1
KC_CACHE_FILENAME = "kc_indexes_cache.pkl"
KC_ZIP_FILENAMES = [
    "Residential Building.zip",
    "Parcel.zip",
    "Real Property Account.zip",
    "Value History.zip",
    "Real Property Sales.zip",
]


def _kc_cache_path(script_dir):
    """Return the absolute path to the KC indexes cache file."""
    return os.path.join(script_dir, KC_CACHE_FILENAME)


def _kc_cache_is_fresh(script_dir):
    """Return True if the cache exists AND is newer than every KC ZIP that
    actually exists on disk. A missing ZIP is ignored (we can't compare its
    mtime), since the loaders themselves degrade gracefully on missing files.
    """
    cache_path = _kc_cache_path(script_dir)
    if not os.path.exists(cache_path):
        return False
    cache_mtime = os.path.getmtime(cache_path)
    for zname in KC_ZIP_FILENAMES:
        zpath = os.path.join(script_dir, zname)
        if os.path.exists(zpath):
            if os.path.getmtime(zpath) > cache_mtime:
                return False
    return True


def _kc_load_from_cache(script_dir):
    """Try to load the KC indexes from the cache file.
    Returns the indexes dict on success, or None on any failure (missing,
    corrupt, version mismatch, etc.). Callers should fall back to loading
    from the ZIPs.
    """
    import pickle
    cache_path = _kc_cache_path(script_dir)
    try:
        size_mb = os.path.getsize(cache_path) / (1024 * 1024)
        print(f"Loading King County indexes from cache "
              f"({KC_CACHE_FILENAME}, {size_mb:.0f} MB)...")
        with open(cache_path, "rb") as f:
            payload = pickle.load(f)
        if not isinstance(payload, dict) or payload.get("version") != KC_CACHE_VERSION:
            print(f"  [WARNING] Cache version mismatch "
                  f"(found {payload.get('version')!r}, "
                  f"expected {KC_CACHE_VERSION}). Rebuilding from ZIPs.")
            return None
        indexes = payload.get("indexes")
        if not isinstance(indexes, dict) or "address_index" not in indexes:
            print("  [WARNING] Cache structure is invalid. Rebuilding from ZIPs.")
            return None
        print(f"  Loaded {len(indexes.get('address_index', {})):,} addresses, "
              f"{len(indexes.get('pin_to_parcel', {})):,} parcels, "
              f"{len(indexes.get('pin_to_acct', {})):,} accounts, "
              f"{len(indexes.get('pin_to_value', {})):,} value records, "
              f"{len(indexes.get('pin_to_sales', {})):,} sales records.")
        return indexes
    except (pickle.PickleError, EOFError, OSError, AttributeError) as e:
        print(f"  [WARNING] Could not load cache ({e}). Rebuilding from ZIPs.")
        return None


def _kc_save_to_cache(script_dir, indexes):
    """Save the loaded KC indexes to the cache file. Best-effort: failures
    are warned and ignored, since a missing cache only costs a slow next run."""
    import pickle
    cache_path = _kc_cache_path(script_dir)
    tmp_path = cache_path + ".tmp"
    try:
        print(f"Saving King County indexes to cache for faster future runs...")
        payload = {"version": KC_CACHE_VERSION, "indexes": indexes}
        with open(tmp_path, "wb") as f:
            pickle.dump(payload, f, protocol=pickle.HIGHEST_PROTOCOL)
        # Atomic rename so a crash mid-write doesn't leave a half-written cache
        os.replace(tmp_path, cache_path)
        size_mb = os.path.getsize(cache_path) / (1024 * 1024)
        print(f"  Cache saved ({KC_CACHE_FILENAME}, {size_mb:.0f} MB). "
              f"Next run will load in seconds.")
    except (OSError, pickle.PickleError) as e:
        print(f"  [WARNING] Could not save cache ({e}). "
              f"Next run will rebuild from ZIPs.")
        # Clean up partial write
        try:
            if os.path.exists(tmp_path):
                os.remove(tmp_path)
        except OSError:
            pass


def kc_load_all(script_dir):
    """Load all five King County bulk data files once.
    Returns a dict of indexes, or None if the primary (address) index is empty.
    The address index is required; without it no KC address can be resolved.

    v2.5.0: caches the loaded indexes to disk. A fresh cache is used directly
    (5-10s load); a stale or missing cache triggers a full ZIP rebuild and the
    result is saved for next time.
    """
    # Try the cache first
    if _kc_cache_is_fresh(script_dir):
        cached = _kc_load_from_cache(script_dir)
        if cached is not None:
            return cached
        # If we get here, the cache existed but was unloadable. Fall through
        # to rebuild from ZIPs.

    # Full rebuild from ZIPs (also the path taken on first run)
    print("\nLoading King County bulk data from ZIPs...")
    print("(First run after a fresh ZIP download. This takes several minutes.")
    print(" Subsequent runs will use the cached indexes and load in seconds.)")
    address_index, pin_to_building = kc_load_resbldg_data(script_dir)
    pin_to_parcel = kc_load_parcel_data(script_dir)
    pin_to_acct   = kc_load_rpacct_data(script_dir)
    pin_to_value  = kc_load_value_data(script_dir)
    pin_to_sales  = kc_load_sales_data(script_dir)

    if not address_index:
        print("[WARNING] No King County address index. KC ZIP files are "
              "missing or unreadable. King County lookups will be skipped.")
        return None

    indexes = {
        "address_index":   address_index,
        "pin_to_building": pin_to_building,
        "pin_to_parcel":   pin_to_parcel,
        "pin_to_acct":     pin_to_acct,
        "pin_to_value":    pin_to_value,
        "pin_to_sales":    pin_to_sales,
    }

    # Save for next time (best-effort; failures don't break this run)
    _kc_save_to_cache(script_dir, indexes)

    return indexes


_KC_DIRECTIONS = ('N', 'S', 'E', 'W', 'NE', 'NW', 'SE', 'SW')
_KC_STREET_TYPES = ('ST', 'AVE', 'DR', 'RD', 'LN', 'CT', 'PL', 'BLVD',
                    'CIR', 'TER', 'PKWY', 'HWY', 'WY', 'WAY', 'LOOP',
                    'TRL', 'PT', 'RUN')


def kc_address_components(norm_addr):
    """Break a normalized KC address into matchable components.

    Returns a dict:
        number    - leading house number (str) or ''
        prefix    - leading directional (N/S/E/W/NE/...) or ''
        name      - core street-name token (the part that is neither a
                    directional nor a street type), or ''
        suffix    - trailing directional or ''
        sttype    - street type (ST/AVE/DR/...) or ''
        tokens    - the full ordered token list after the house number

    The point of this helper is to compare addresses by component instead of
    by substring. '30TH DR SE' and '330TH AVE NE' share no component: the
    name differs (30TH vs 330TH), the type differs (DR vs AVE), and the
    suffix differs (SE vs NE). The old substring test ('30TH' in '330TH...')
    matched them anyway. This is what produced the wrong-property KC row.
    """
    parts = norm_addr.split()
    comp = {'number': '', 'prefix': '', 'name': '', 'suffix': '',
            'sttype': '', 'tokens': []}
    if not parts:
        return comp

    idx = 0
    if parts[0].isdigit():
        comp['number'] = parts[0]
        idx = 1
    comp['tokens'] = parts[idx:]

    rest = list(parts[idx:])
    if not rest:
        return comp

    # Leading directional
    if rest[0] in _KC_DIRECTIONS:
        comp['prefix'] = rest.pop(0)

    # Trailing directional
    if rest and rest[-1] in _KC_DIRECTIONS:
        comp['suffix'] = rest.pop()

    # Trailing street type
    if rest and rest[-1] in _KC_STREET_TYPES:
        comp['sttype'] = rest.pop()

    # Whatever is left is the core street name (join in case of two-word names)
    comp['name'] = ' '.join(rest)
    return comp


def _kc_extract_zip(addr):
    """Pull a 5-digit zip off the end of an address, or '' if none."""
    if not addr:
        return ''
    m = re.search(r'(\d{5})(-\d{4})?\s*$', addr.strip())
    return m.group(1) if m else ''


def kc_components_match(search_c, cand_c):
    """True if a candidate address is a credible match for the search address,
    compared component by component. Strict where the search address gives us
    something to check, lenient where it doesn't.

    Rules:
      - house number must be identical
      - core street name must be identical
      - if the search has a directional (prefix or suffix), the candidate must
        share it. A search 'SE' will not match a candidate 'NE' or a candidate
        with no directional at all.
      - if BOTH have a street type, they must match. If the search omits the
        street type, the candidate's type is not held against it.
    """
    if not search_c['number'] or search_c['number'] != cand_c['number']:
        return False
    if not search_c['name'] or search_c['name'] != cand_c['name']:
        return False

    search_dir = search_c['prefix'] or search_c['suffix']
    cand_dir = cand_c['prefix'] or cand_c['suffix']
    if search_dir and search_dir != cand_dir:
        return False

    if search_c['sttype'] and cand_c['sttype'] \
            and search_c['sttype'] != cand_c['sttype']:
        return False

    return True


def kc_find_address(search_addr, address_index, pin_to_building=None,
                    raw_search_addr=None):
    """Find a King County property by address. Returns PIN or None.

    Tier 1: exact normalized match.
    Tier 2: strict component match. House number and street name must match.
            A directional, when the search has one, must match. A street
            type, when both sides have one, must match. When a zip is
            available on both sides, a candidate whose zip disagrees is
            rejected. If more than one credible candidate survives, the
            address is reported as ambiguous and skipped.

    v2.4.1: the old code had a substring-based third tier (`search_street in
    addr`) that matched '30TH' inside '330TH' and returned that lone hit with
    no checks: wrong street, wrong city, wrong zip, no warning. It also had a
    suffix-stripped tier that returned the FIRST candidate whose street type
    happened to be strippable to the same stem, so 'OAK ST' and 'OAK AVE'
    were interchangeable and whichever came first in the index won. Both
    loose tiers are gone. Everything below the exact match now goes through
    one strict, zip-guarded, ambiguity-aware path.

    pin_to_building, when supplied, lets the matcher read each candidate's
    ZipCode for the zip guard. raw_search_addr is the address BEFORE
    clean_address_for_search() stripped the city/zip; it is the only place
    the search-side zip still survives.
    """
    norm_search = kc_normalize_address(search_addr)

    # Tier 1: exact match
    if norm_search in address_index:
        return address_index[norm_search]

    # Tier 2: strict component match
    search_c = kc_address_components(norm_search)
    if not search_c['number'] or not search_c['name']:
        # No house number or no street name to match on. Refuse to guess.
        return None

    search_zip = _kc_extract_zip(raw_search_addr or search_addr)

    matches = []  # (addr, pin)
    for addr, pin in address_index.items():
        if not addr.startswith(search_c['number'] + ' '):
            continue
        cand_c = kc_address_components(addr)
        if not kc_components_match(search_c, cand_c):
            continue
        # Zip guard: if we know both zips and they disagree, this is not it.
        if search_zip and pin_to_building is not None:
            cand_zip = (pin_to_building.get(pin, {}) or {}).get('ZipCode', '')
            cand_zip = (cand_zip or '').strip()[:5]
            if cand_zip and cand_zip != search_zip:
                continue
        matches.append((addr, pin))

    if len(matches) == 1:
        return matches[0][1]
    if len(matches) > 1:
        print(f"  Multiple KC matches found for '{search_addr}':")
        for addr, pin in matches[:5]:
            print(f"    - {addr}")
        if len(matches) > 5:
            print(f"    ... and {len(matches) - 5} more")
        print(f"  Skipping (ambiguous): pick the exact address to disambiguate.")

    return None


def kc_build_data(pin, kc_indexes):
    """Combine the five KC bulk indexes into the master COLUMN_ORDER schema
    for one PIN. Returns a data dict, or None if the PIN has no building
    record. Owner fields are NOT set here; they come from the lead paste.

    The 15 columns KC bulk data cannot fill (Property Description, Taxpayer
    Name, Structure Description/Type, Neighborhood Code, Annual Tax Amount,
    Property Category, Status, Tax Code Area, Floor Details, Fireplace,
    Foundation, Exterior, Roof Type, Assessor Photo File) are left unset and
    will write as blank."""
    building = kc_indexes["pin_to_building"].get(pin, {})
    if not building:
        return None

    parcel = kc_indexes["pin_to_parcel"].get(pin, {})
    acct   = kc_indexes["pin_to_acct"].get(pin, {})
    value  = kc_indexes["pin_to_value"].get(pin, {})
    sales  = kc_indexes["pin_to_sales"].get(pin, [])

    data = {"County": "King"}

    # --- Identification ---
    data["Parcel Number"] = f"{pin[:6]}-{pin[6:]}"

    # --- Address (+ parsed components) ---
    # v2.5.1: KC's Residential Building.Address field has no city and
    # sometimes no zip (e.g. parcel 152407-9196, 32732 SE 44TH ST, comes
    # back as just "32732 SE 44TH ST"). Build a complete address using:
    #   - Street: from Residential Building.Address (zip stripped if present)
    #   - City:   from Parcel.DistrictName (the KC field for municipal
    #             jurisdiction; "King County" for unincorporated areas)
    #   - State:  always "WA" for KC properties
    #   - Zip:    from Residential Building.ZipCode (separate field, more
    #             reliable than the zip-in-Address that may be missing)
    raw_addr = kc_collapse_spaces(building.get('Address', ''))
    zip_code = (building.get('ZipCode', '') or '').strip()
    district = (parcel.get('DistrictName', '') or '').strip()

    # Strip a trailing zip from the address string, since we'll add a
    # canonical zip ourselves. Matches "12345" or "12345-6789" at end.
    street_only = raw_addr
    m = re.match(r'^(.+?)\s+(\d{5}(?:-\d{4})?)\s*$', raw_addr)
    if m:
        street_only = m.group(1).strip()
        # If ZipCode wasn't set on the building row, fall back to the
        # zip we just stripped off the address string.
        if not zip_code:
            zip_code = m.group(2).strip()

    # City: prefer DistrictName. "King County" is the unincorporated marker
    # in KC data; if that's what we got, leave the city blank rather than
    # writing "King County" into the City column. v2.5.1: parse_full_address
    # cannot fill the city for KC (the source string has none), so we set
    # it here directly from DistrictName.
    if district and district.upper() != "KING COUNTY":
        prop_city = district
    else:
        prop_city = ""

    # v2.5.2/v2.5.3: DistrictName is unreliable. Some parcels with a real
    # municipal city (e.g. Woodinville, Redmond, Fall City) come back with
    # a blank or "King County" DistrictName, and a few parcels also lack
    # both a Residential Building.ZipCode AND a trailing zip in their
    # Address field (parcel 152407-9196 in the May 20 data is the seen
    # case). In both cases the tax billing address carries the missing
    # piece. We only trust the billing fields when the billing STREET
    # matches the property street, since for absentee owners the billing
    # city/zip is the OWNER'S mailing info, not the property's.
    # Owner-occupancy is the gate; compute it once and reuse for city+zip.
    billing_match = False
    billing_addr = ''
    if (not prop_city) or (not zip_code):
        billing_addr = (acct.get('BillingAddress', '') or '').strip()
        if billing_addr:
            bill_street = billing_addr.split(',', 1)[0].strip()
            if bill_street and street_only and \
                    kc_normalize_address(bill_street) == \
                    kc_normalize_address(street_only):
                billing_match = True

    if not prop_city and billing_match:
        prop_city = kc_parse_city_from_billing(billing_addr)

    if not zip_code and billing_match:
        zip_code = kc_parse_zip_from_billing(billing_addr)

    # Build the canonical Property Address as "STREET, CITY WA ZIP".
    # When city is unknown we still emit "STREET WA ZIP" so the zip
    # survives into the parsed columns.
    parts = [street_only]
    if prop_city:
        parts.append(f"{prop_city} WA {zip_code}".strip())
    elif zip_code:
        parts.append(f"WA {zip_code}".strip())
    prop_addr = ", ".join(p for p in parts if p)

    data["Property Address"] = prop_addr
    data["Property Street"]  = street_only
    data["Property City"]    = prop_city
    data["Property State"]   = "WA" if (street_only or zip_code) else ""
    data["Property Zip"]     = zip_code

    # --- Subdivision ---
    data["Subdivision Name"] = parcel.get('PlatName', '')

    # --- Structure details ---
    data["Year Built"] = building.get('YrBuilt', '')
    data["Bedrooms"]   = building.get('Bedrooms', '')

    # KC stores full and 3/4 baths separately. The master schema's
    # "Full or 3/4 Baths" column is the sum of the two.
    full_b = building.get('BathFullCount', '') or '0'
    tq_b   = building.get('Bath3qtrCount', '') or '0'
    try:
        combined = int(full_b) + int(tq_b)
        data["Full or 3/4 Baths"] = str(combined) if combined else ''
    except ValueError:
        data["Full or 3/4 Baths"] = building.get('BathFullCount', '')
    data["Half Baths"] = building.get('BathHalfCount', '')

    data["Total Finished SF"] = building.get('SqFtTotLiving', '')

    # Lot size -> Size (gross) / Unit of Measure
    lot = parcel.get('SqFtLot', '')
    if lot:
        data["Size (gross)"]    = lot
        data["Unit of Measure"] = "Sq Ft"

    data["Use Code"]          = parcel.get('PresentUse', '')
    data["Property Grade"]    = building.get('BldgGrade', '')
    data["Property Condition"] = building.get('Condition', '')
    data["Heat"]              = building.get('HeatSystem', '')
    data["Garage SF"]         = building.get('SqFtGarageAttached', '')
    data["Views"]             = parcel.get('Views', '')
    data["Waterfront"]        = parcel.get('Waterfront', '')

    # --- Financial / value data ---
    if value:
        data["Latest Tax Year"]       = value.get('TaxYear', '')
        data["Market Total"]          = value.get('ApprTotVal', '')
        data["Market Land"]           = value.get('ApprLandVal', '')
        data["Market Improvement"]    = value.get('ApprImpsVal', '')
        data["Taxable Value Regular"] = value.get('TaxableTotVal', '')
        # KC bulk data has no separate assessed value; mirror the taxable total.
        data["Assessed Value"]        = value.get('TaxableTotVal', '')

    # --- Tax billing address (used for the Ownership Match check) ---
    data["Tax Address"] = acct.get('BillingAddress', '')

    # --- Sales history ---
    if sales:
        most_recent = sales[0]
        data["Most Recent Sale Date"]   = most_recent.get('DocumentDate', '')
        amt = most_recent.get('SalePrice', '')
        data["Most Recent Sale Amount"] = f"${amt}" if amt else ''
        data["Sales History"] = "; ".join(
            f"{s.get('DocumentDate','')} - ${s.get('SalePrice','')} "
            f"({s.get('SaleInstrument','')}) {s.get('SellerName','')} -> "
            f"{s.get('BuyerName','')}"
            for s in sales
        )

    return data


def compute_ownership_match_kc(data, lead_data):
    """King County ownership match. Lead-based, weaker than Snohomish's
    county-confirmed match.

    KC's public bulk file has no taxpayer names, so the lead name IS the
    owner name. The meaningful signal is the tax billing address: if it
    matches the property address the owner lives there; if not, the owner
    is absentee.

    Returns:
        CONFIRMED    - Lead name present AND tax billing address matches the
                       property address (owner-occupied)
        PARTIAL      - Lead name present but tax billing address differs from
                       the property address (absentee owner), or no billing
                       address is available to check
        NO LEAD DATA - No lead last name was provided
    """
    if not lead_data:
        return "NO LEAD DATA"
    if not lead_data.get("Lead Last Name", "").strip():
        return "NO LEAD DATA"

    prop_addr = data.get("Property Address", "")
    bill_addr = data.get("Tax Address", "")

    if not bill_addr:
        # No billing address to compare against. Lead exists, so not "NO LEAD
        # DATA", but we can't confirm owner-occupancy.
        return "PARTIAL"

    # Compare just the street portion of each address. The billing address
    # carries city/state/zip; the property address may or may not. Normalizing
    # and taking the leading street segment makes the comparison apples-to-apples.
    prop_norm = kc_normalize_address(prop_addr.split(",")[0])
    bill_norm = kc_normalize_address(bill_addr.split(",")[0])

    if prop_norm and bill_norm and prop_norm == bill_norm:
        return "CONFIRMED"
    return "PARTIAL"


def lookup_property_kc(address, kc_indexes, lead_data=None,
                       api_key=None, photo_folder="property_photos"):
    """Look up a King County property from the bulk data indexes.
    Mirrors the role scrape_property_kc() used to play, but reads from disk
    instead of the KC website.

    Owner names come from lead_data (the KC bulk file has none). Street View
    is the only photo source for KC rows.

    Returns a data dict or None on failure."""
    print(f"\n{'='*60}")
    print(f"  Looking up (King County, bulk data): {address}")
    print(f"{'='*60}")

    search_addr = clean_address_for_search(address)

    # v2.4.1: pass pin_to_building (for the candidate-zip guard) and the RAW
    # address (the only copy that still carries the search-side zip, since
    # clean_address_for_search strips the city and zip).
    pin = kc_find_address(search_addr, kc_indexes["address_index"],
                          kc_indexes.get("pin_to_building"),
                          raw_search_addr=address)
    if not pin:
        print(f"  [ERROR] No King County match for: {search_addr}")
        print(f"  Note: KC bulk address index is residential-only. Commercial, "
              f"vacant, and multifamily parcels cannot be found by address.")
        return None

    data = kc_build_data(pin, kc_indexes)
    if not data:
        print(f"  [ERROR] PIN {pin} has no residential building record")
        return None

    print(f"  -> Parcel: {data.get('Parcel Number', 'N/A')}")
    print(f"  -> Address: {data.get('Property Address', 'N/A')}")
    print(f"  -> Year Built: {data.get('Year Built', 'N/A')}")
    print(f"  -> Beds: {data.get('Bedrooms', 'N/A')} | "
          f"Baths: {data.get('Full or 3/4 Baths', 'N/A')} | "
          f"SF: {data.get('Total Finished SF', 'N/A')}")
    print(f"  -> Market Total: {data.get('Market Total', 'N/A')} "
          f"({data.get('Latest Tax Year', 'N/A')})")
    print(f"  -> Tax Billing Address: {data.get('Tax Address', 'N/A')}")

    # Owner names: from the lead paste only. KC public bulk data has none.
    if lead_data and lead_data.get("Lead Last Name", "").strip():
        data["Owner 1 First Name"] = lead_data.get("Lead First Name", "")
        data["Owner 1 Last Name"]  = lead_data.get("Lead Last Name", "")
        print(f"  -> Owner (from lead): {data['Owner 1 First Name']} "
              f"{data['Owner 1 Last Name']}")
    else:
        print(f"  -> No lead name attached. Owner fields left blank "
              f"(KC bulk data has no taxpayer names).")

    # Street View photo (same Google API as Snohomish). Only photo source for KC.
    if api_key and data.get("Property Address"):
        print(f"  -> Downloading Street View photo...")
        sv = download_street_view(
            data["Property Address"],
            data.get("Owner 1 Last Name", ""),
            api_key, photo_folder,
        )
        data["Photo File"] = sv
        if sv:
            print(f"  -> Street View photo saved: {sv}")
        else:
            print(f"  -> No Street View image available")
    else:
        data["Photo File"] = ""

    data = apply_title_case(data)
    data["County"] = "King"  # restore: apply_title_case would lowercase it
    return data


# ═══════════════════════════════════════════════════════════════════════════════
#  KING COUNTY WEB SCRAPER  (RETAINED, NOT ROUTED - see v2.3.0 notes above)
# ═══════════════════════════════════════════════════════════════════════════════
#  The functions below scraped the King County eReal Property website. As of
#  v2.3.0 they are NO LONGER CALLED. KC data now comes from the bulk loaders
#  above. This code is kept intact in case KC server access ever becomes
#  viable. Do not delete without a decision to retire it permanently.
# ═══════════════════════════════════════════════════════════════════════════════

def get_kc_session():
    """Open a session to the King County eReal Property site."""
    session = requests.Session()
    session.headers.update(HTTP_HEADERS)
    resp = session.get(KC_BASE_URL, allow_redirects=True)
    resp.raise_for_status()
    return session, resp


def search_address_kc(session, address, base_resp):
    """Submit an address search to King County eReal and return the response."""
    fields = _extract_asp_fields(base_resp.text)
    # King County field names use the long ASP.NET naming convention
    form = {
        **fields,
        "kingcounty_gov$cphContent$checkbox_acknowledge": "on",  # Terms acceptance checkbox
        "kingcounty_gov$cphContent$hf_accept": "1",
        "kingcounty_gov$cphContent$txtAddress": address,
        "kingcounty_gov$cphContent$txtCity":    "",
        "kingcounty_gov$cphContent$txtZip":     "",
        "kingcounty_gov$cphContent$btn_SearchAddress": "Search",
        "__SCROLLPOSITIONX": "0",
        "__SCROLLPOSITIONY": "0",
        "__VIEWSTATEENCRYPTED": "",
    }
    resp = session.post(KC_BASE_URL, data=form, allow_redirects=True)
    resp.raise_for_status()
    return resp


def parse_kc_search_results(resp):
    """Extract parcel links from KC search results page.
    Returns list of (parcel_nbr, address, dashboard_url) or [('DIRECT', resp)]."""
    # If we landed directly on a Dashboard page
    if "Dashboard.aspx" in resp.url:
        return [("DIRECT", resp)]

    soup = BeautifulSoup(resp.text, "html.parser")
    results = []
    for link in soup.find_all("a", href=True):
        href = link["href"]
        if "Dashboard.aspx" in href and "ParcelNbr=" in href:
            m = re.search(r'ParcelNbr=(\d+)', href)
            parcel = m.group(1) if m else link.text.strip()
            row = link.find_parent("tr")
            addr = ""
            if row:
                cells = row.find_all("td")
                if len(cells) >= 2:
                    addr = cells[1].text.strip()
            full_url = href if href.startswith("http") else urljoin(KC_BASE_URL, href)
            results.append((parcel, addr, full_url))
    return results


def parse_kc_dashboard(html, url):
    """Parse the King County eReal Property Dashboard page.
    Returns a data dict with all available fields."""
    soup = BeautifulSoup(html, "html.parser")
    data = {"County": "King"}

    # --- Parcel header block (GridViewStyle table with Parcel Number / Name / Site Address / Legal) ---
    header_table = soup.find("table", id="cphContent_DetailsViewDashboardHeader")
    if header_table:
        for row in header_table.find_all("tr"):
            cells = row.find_all("td")
            if len(cells) == 2:
                lbl, val = cells[0].text.strip(), cells[1].text.strip()
                if lbl == "Parcel Number":
                    # KC format: "295440-0020" - store without hyphen for consistency
                    data["Parcel Number"] = val.replace("-", "")
                elif lbl == "Name":
                    data.update(split_owner_names(val.strip()))
                elif lbl == "Site Address":
                    # KC gives "7214 204TH DR NE 98053" - no city. We'll clean it up.
                    data["Property Address"] = val.strip()
                elif lbl == "Legal":
                    data["Property Description"] = val.strip()
                    # Subdivision name is the legal description directly for KC
                    data["Subdivision Name"] = val.strip()

    # --- Parse full address into components ---
    if data.get("Property Address"):
        data.update(parse_full_address(data["Property Address"]))

    # --- Building characteristics ---
    building_table = soup.find("table", id="cphContent_DetailsViewPropTypeR")
    if building_table:
        for row in building_table.find_all("tr"):
            cells = row.find_all("td")
            if len(cells) == 2:
                lbl, val = cells[0].text.strip(), cells[1].text.strip()
                if lbl == "Year Built":
                    data["Year Built"] = val
                elif lbl == "Total Square Footage":
                    data["Total Finished SF"] = val
                elif lbl == "Number Of Bedrooms":
                    data["Bedrooms"] = val
                elif lbl == "Number Of Baths":
                    # KC gives decimal baths e.g. 3.75
                    # Store in Full or 3/4 Baths as-is, leave Half Baths blank
                    data["Full or 3/4 Baths"] = val
                elif lbl == "Grade":
                    data["Property Grade"] = val
                elif lbl == "Condition":
                    data["Property Condition"] = val
                elif lbl == "Lot Size":
                    data["Size (gross)"] = val
                    data["Unit of Measure"] = "Sq Ft"
                elif lbl == "Views":
                    data["Views"] = val
                elif lbl == "Waterfront":
                    data["Waterfront"] = val.strip() if val.strip() else "No"

    # --- Tax roll history - grab most recent row ---
    tax_table = soup.find("table", id="cphContent_GridViewDBTaxRoll")
    if tax_table:
        rows = tax_table.find_all("tr")
        # First data row (index 1) is the most recent
        if len(rows) >= 2:
            cells = rows[1].find_all("td")
            if len(cells) >= 9:
                data["Latest Tax Year"]      = cells[1].text.strip()
                data["Market Land"]          = cells[2].text.strip()
                data["Market Improvement"]   = cells[3].text.strip()
                data["Market Total"]         = cells[4].text.strip()
                data["Taxable Value Regular"] = cells[8].text.strip()
                # KC doesn't separate Assessed Value from Market Total at this level
                data["Assessed Value"]       = cells[8].text.strip()

    # --- Levy code (Tax Code Area) ---
    levy_form = soup.find("table", id="cphContent_FormViewLevyDist")
    if levy_form:
        levy_label = levy_form.find("span", id="cphContent_FormViewLevyDist_Label1")
        if levy_label:
            data["Tax Code Area"] = levy_label.text.strip()

    # --- Assessor photo ---
    photo_img = soup.find("img", id="cphContent_FormViewPictCurr_CurrentImage")
    if photo_img and photo_img.get("src"):
        src = photo_img["src"]
        # Build absolute URL
        data["_kc_photo_url"] = src if src.startswith("http") else urljoin(url, src)

    return data


def scrape_property_kc(session, base_resp, address, api_key=None, photo_folder="property_photos"):
    """Search King County eReal for an address, parse the Dashboard, download photos.
    Returns a data dict or None on failure."""
    print(f"\n{'='*60}")
    print(f"  Searching (King County): {address}")
    print(f"{'='*60}")

    # Strip city/state for the search box - KC wants just street number + name
    search_addr = clean_address_for_search(address)

    try:
        results_resp = search_address_kc(session, search_addr, base_resp)
    except Exception as e:
        print(f"  [ERROR] KC search failed: {e}")
        return None

    results = parse_kc_search_results(results_resp)

    if not results:
        print(f"  [ERROR] No KC results found for: {search_addr}")
        return None

    # Fetch Dashboard page
    if results[0][0] == "DIRECT":
        dash_html = results[0][1].text
        dash_url  = results[0][1].url
    else:
        parcel_nbr, parcel_addr, dash_href = results[0]
        print(f"  -> Found: {parcel_nbr} | {parcel_addr}")
        if len(results) > 1:
            print(f"  -> ({len(results)} total results, using first match)")
        try:
            r = session.get(dash_href, timeout=15)
            r.raise_for_status()
            dash_html, dash_url = r.text, r.url
        except Exception as e:
            print(f"  [ERROR] Could not fetch KC Dashboard: {e}")
            return None

    data = parse_kc_dashboard(dash_html, dash_url)

    print(f"  -> Parcel: {data.get('Parcel Number', 'N/A')}")
    print(f"  -> Owner: {data.get('Owner 1 First Name', '')} {data.get('Owner 1 Last Name', '')}")
    print(f"  -> Assessed: {data.get('Assessed Value', 'N/A')}")
    print(f"  -> Year Built: {data.get('Year Built', 'N/A')}")
    print(f"  -> Beds: {data.get('Bedrooms', 'N/A')} | Baths: {data.get('Full or 3/4 Baths', 'N/A')} | SF: {data.get('Total Finished SF', 'N/A')}")
    print(f"  -> Grade: {data.get('Property Grade', 'N/A')} | Condition: {data.get('Property Condition', 'N/A')}")

    # Assessor photo
    kc_photo_url = data.pop("_kc_photo_url", "")
    if kc_photo_url:
        print(f"  -> Downloading assessor photo...")
        ap = download_assessor_photo(
            kc_photo_url,
            data.get("Property Address", ""),
            data.get("Owner 1 Last Name", ""),
            photo_folder,
            session=session,
        )
        data["Assessor Photo File"] = ap
        if ap:
            print(f"  -> Assessor photo saved: {ap}")

    # Street View photo (same Google API as Snohomish)
    if api_key and data.get("Property Address"):
        print(f"  -> Downloading Street View photo...")
        sv = download_street_view(
            data["Property Address"],
            data.get("Owner 1 Last Name", ""),
            api_key, photo_folder,
        )
        data["Photo File"] = sv
        if sv:
            print(f"  -> Street View photo saved: {sv}")
        else:
            print(f"  -> No Street View image available")
    else:
        data["Photo File"] = ""

    data = apply_title_case(data)
    return data


# ═══════════════════════════════════════════════════════════════════════════════
#  PHOTO DOWNLOADS
# ═══════════════════════════════════════════════════════════════════════════════

def _photo_filename(address, owner_last, suffix):
    """Build '9714 51st Ave NE (Potter) - Street View.jpg'."""
    street = smart_title_case(address.split(",")[0].strip())
    street = re.sub(r'[<>:"/\\|?*]', '', street).strip()
    if owner_last:
        owner = re.sub(r'[<>:"/\\|?*]', '', smart_title_case(owner_last)).strip()
        return f"{street} ({owner}) - {suffix}.jpg"
    return f"{street} - {suffix}.jpg"


def download_street_view(address, owner_last, api_key, folder):
    """Download Google Street View photo. Returns filepath or ''."""
    if not api_key:
        return ""
    os.makedirs(folder, exist_ok=True)
    path = os.path.join(folder, _photo_filename(address, owner_last, "Street View"))
    try:
        resp = requests.get("https://maps.googleapis.com/maps/api/streetview",
                            params={"size": "2048x2048", "location": address,
                                    "fov": "60", "pitch": "5", "key": api_key},
                            timeout=15)
        if resp.status_code == 200 and "image" in resp.headers.get("Content-Type", ""):
            with open(path, "wb") as f:
                f.write(resp.content)
            return path
    except Exception as e:
        print(f"  [WARNING] Street View download failed: {e}")
    return ""


def download_assessor_photo(photo_url, address, owner_last, folder, session=None):
    """Download the assessor's property photo. Returns filepath or ''.
    Pass session for King County (MediaHandler requires same session cookies)."""
    if not photo_url:
        return ""
    os.makedirs(folder, exist_ok=True)
    path = os.path.join(folder, _photo_filename(address, owner_last, "Assessor"))
    try:
        fetcher = session.get if session else requests.get
        resp = fetcher(photo_url, timeout=15)
        if resp.status_code == 200 and len(resp.content) > 1000:
            with open(path, "wb") as f:
                f.write(resp.content)
            return path
    except Exception as e:
        print(f"  [WARNING] Assessor photo download failed: {e}")
    return ""


# ═══════════════════════════════════════════════════════════════════════════════
#  SCRAPE ONE PROPERTY (Snohomish County orchestrator)
# ═══════════════════════════════════════════════════════════════════════════════

def _street_signature(addr):
    """Reduce an address to a comparable (house_number, street_core) signature
    for wildcard-result matching. Drops unit suffixes, abbreviates directions
    and street types, and ignores city/state/zip.

    Returns (number, set_of_street_words) or (None, set()) if no house number.
    """
    if not addr:
        return None, set()
    # Cut at the first comma (city/state/zip) and strip a trailing zip.
    core = addr.split(",")[0].strip()
    core = re.sub(r'\s+\d{5}(-\d{4})?$', '', core).strip()
    # Drop unit/apt suffixes: "#2", "APT 3", "UNIT B", "STE 100".
    core = re.sub(r'\s+(#|APT|UNIT|STE|SUITE)\s*\S*$', '', core, flags=re.I)
    core = re.sub(r'\s+#\S*$', '', core)
    # Normalize directions and street types to the abbreviated forms.
    for full, abbr in DIRECTION_MAP.items():
        core = re.sub(r'\b' + full + r'\b', abbr, core, flags=re.I)
    for full, abbr in STREET_TYPE_MAP.items():
        core = re.sub(r'\b' + full + r'\b', abbr, core, flags=re.I)
    parts = core.upper().split()
    if not parts or not parts[0].isdigit():
        return None, set()
    number = parts[0]
    street_words = set(parts[1:])
    if not street_words:
        # A bare house number with no street (e.g. a truncated lead '19815').
        # Treat as unparseable so callers refuse to guess a match.
        return None, set()
    return number, street_words


def _filter_wildcard_results(results, input_address):
    """Bug 2 fix. Given wildcard search results and the original input
    address, return only the results whose street actually matches the
    input street. Prevents the scraper from silently grabbing the wrong
    property (e.g. '212 Ave C' matching '212 86TH AVE SE').

    A result matches when the house number is identical AND the result's
    street words are a subset/superset overlap of the input's street words
    (every input street word appears in the result, allowing for the
    assessor's fuller formatting).

    If the input has no parseable house number and street (e.g. a truncated
    lead like '19815'), nothing matches and the caller skips the address.
    """
    in_num, in_words = _street_signature(input_address)
    if in_num is None or not in_words:
        # Truncated or unparseable input: refuse to guess.
        return []
    matched = []
    for r in results:
        if r[0] == "DIRECT":
            matched.append(r)
            continue
        # parse_search_results tuple: (parcel, addr, url)
        result_addr = r[1] if len(r) > 1 else ""
        res_num, res_words = _street_signature(result_addr)
        if res_num is None:
            continue
        if res_num != in_num:
            continue
        # Every input street word must appear in the result's street words.
        if in_words.issubset(res_words):
            matched.append(r)
    return matched


def scrape_property(session, address, base_resp, api_key=None, photo_folder="property_photos"):
    """Search for an address, parse the parcel page, download photos.
    Returns a data dict or None on failure.

    As of v2.4.0 this receives the RAW address (with city/state/zip). It
    cleans the address itself, the same way lookup_property_kc() does. This
    keeps county routing (detect_county, which needs the city) separate from
    the assessor search (which needs the city stripped)."""
    search_address_str = clean_address_for_search(address)

    print(f"\n{'='*60}")
    print(f"  Searching: {search_address_str}")
    print(f"{'='*60}")

    # --- Search ---
    try:
        results_resp = search_address(session, search_address_str, base_resp)
    except Exception as e:
        print(f"  [ERROR] Search failed: {e}")
        return None

    results = parse_search_results(results_resp)

    # Wildcard retry if no results
    if not results:
        m = re.match(r'^(\d+)\s+(.+)$', search_address_str)
        if m:
            wildcard = f"{m.group(1)}%{m.group(2).split()[0]}"
            print(f"  -> No exact match, retrying with wildcard: {wildcard}")
            try:
                s2, b2 = get_session()
                results_resp = search_address(s2, wildcard, b2)
                wild_results = parse_search_results(results_resp)
                if wild_results and wild_results[0][0] != "DIRECT":
                    # Bug 2 fix: do NOT blindly take the first wildcard result.
                    # Require the result's street name to match the input
                    # street name, or skip rather than guess a wrong property.
                    results = _filter_wildcard_results(wild_results, search_address_str)
                    if results:
                        session = s2
                    else:
                        print(f"  [ERROR] Wildcard retry returned "
                              f"{len(wild_results)} result(s), none matching the "
                              f"input street. Skipping rather than guessing.")
                elif wild_results and wild_results[0][0] == "DIRECT":
                    results = wild_results
                    session = s2
            except Exception:
                pass

    if not results:
        print(f"  [ERROR] No results found for: {search_address_str}")
        return None

    # --- Fetch parcel page ---
    if results[0][0] == "DIRECT":
        parcel_html = results[0][1].text
        parcel_url  = results[0][1].url
    else:
        parcel_num, parcel_addr, parcel_href = results[0]
        print(f"  -> Found: {parcel_num} | {parcel_addr}")
        if len(results) > 1:
            print(f"  -> ({len(results)} total results, using first match)")
        try:
            r = session.get(parcel_href, timeout=15)
            r.raise_for_status()
            parcel_html, parcel_url = r.text, r.url
        except Exception as e:
            print(f"  [ERROR] Could not fetch parcel page: {e}")
            return None

    # --- Parse parcel page ---
    data, struct_link = parse_parcel_page(parcel_html, parcel_url)

    print(f"  -> Parcel: {data.get('Parcel Number', 'N/A')}")
    print(f"  -> Owner: {data.get('Owner 1 First Name', '')} {data.get('Owner 1 Last Name', '')}")
    print(f"  -> Assessed: {data.get('Assessed Value', 'N/A')}")
    print(f"  -> Year Built: {data.get('Year Built', 'N/A')}")
    print(f"  -> Last Sale: {data.get('Most Recent Sale Date', 'N/A')} for {data.get('Most Recent Sale Amount', 'N/A')}")
    print(f"  -> Annual Tax: {data.get('Annual Tax Amount', 'N/A')}")
    print(f"  -> Neighborhood: {data.get('Neighborhood Code', 'N/A')}")

    # --- Structure details ---
    if struct_link:
        print(f"  -> Fetching structure details...")
        struct = parse_structure_page(session, struct_link)
        assessor_url = struct.pop("_assessor_photo_url", "")
        data.update(struct)
        print(f"  -> Beds: {struct.get('Bedrooms', 'N/A')} | Baths: {struct.get('Full or 3/4 Baths', 'N/A')} | SF: {struct.get('Total Finished SF', 'N/A')}")

        if assessor_url:
            print(f"  -> Downloading assessor photo...")
            ap = download_assessor_photo(assessor_url, data.get("Property Address", ""),
                                          data.get("Owner 1 Last Name", ""), photo_folder)
            data["Assessor Photo File"] = ap
            if ap:
                print(f"  -> Assessor photo saved: {ap}")
    else:
        print(f"  -> No structure detail link found")

    # --- Street View photo ---
    if api_key and data.get("Property Address"):
        print(f"  -> Downloading Street View photo...")
        sv = download_street_view(data["Property Address"],
                                   data.get("Owner 1 Last Name", ""),
                                   api_key, photo_folder)
        data["Photo File"] = sv
        if sv:
            print(f"  -> Street View photo saved: {sv}")
        else:
            print(f"  -> No Street View image available")
    else:
        data["Photo File"] = ""

    # --- Title case ---
    data = apply_title_case(data)
    data["County"] = "Snohomish"
    return data


# ═══════════════════════════════════════════════════════════════════════════════
#  OUTPUT: CSV + EXCEL
# ═══════════════════════════════════════════════════════════════════════════════

def get_next_row_number(script_dir):
    """Determine next Row # by reading existing CSV."""
    csv_path = os.path.join(script_dir, OUTPUT_CSV)
    if not os.path.exists(csv_path):
        return 1
    try:
        with open(csv_path, "r", encoding="utf-8") as f:
            reader = csv.DictReader(f)
            max_row = 0
            for row in reader:
                try:
                    max_row = max(max_row, int(row.get("Row #", 0)))
                except (ValueError, TypeError):
                    pass
            return max_row + 1
    except Exception:
        return 1


def write_outputs(all_data, script_dir):
    """Write/append data to CSV and rebuild Excel from CSV."""
    if not all_data:
        return

    csv_path  = os.path.join(script_dir, OUTPUT_CSV)
    xlsx_path = os.path.join(script_dir, OUTPUT_XLSX)

    # Merge fieldnames: priority order + any new fields
    existing_fields = []
    if os.path.exists(csv_path):
        try:
            with open(csv_path, "r", encoding="utf-8") as f:
                reader = csv.DictReader(f)
                existing_fields = list(reader.fieldnames or [])
        except Exception:
            pass

    all_keys = set()
    for d in all_data:
        all_keys.update(d.keys())
    if existing_fields:
        all_keys.update(existing_fields)

    fieldnames = [f for f in COLUMN_ORDER if f in all_keys]
    for f in sorted(all_keys):
        if f not in fieldnames:
            fieldnames.append(f)

    # Read existing CSV rows
    existing_rows = []
    if os.path.exists(csv_path):
        try:
            with open(csv_path, "r", encoding="utf-8") as f:
                existing_rows = list(csv.DictReader(f))
        except Exception:
            pass

    # Append new data
    all_rows = existing_rows + all_data
    new_count = len(all_data)

    # Write complete CSV
    with open(csv_path, "w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames, extrasaction="ignore")
        writer.writeheader()
        for row in all_rows:
            writer.writerow(row)

    print(f">> CSV updated: {csv_path}")
    print(f"   Total rows: {len(all_rows)} ({new_count} new)")

    # Build Excel from CSV
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Property Data"

        header_font = Font(bold=True, color="FFFFFF", size=11)
        header_fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")

        # Headers
        for col, name in enumerate(fieldnames, 1):
            cell = ws.cell(row=1, column=col, value=name)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")

        # Data rows
        for r_idx, row in enumerate(all_rows, 2):
            for c_idx, field in enumerate(fieldnames, 1):
                ws.cell(row=r_idx, column=c_idx, value=row.get(field, ""))

        # Auto-width (capped at 40)
        for col in ws.columns:
            max_len = 0
            col_letter = col[0].column_letter
            for cell in col:
                if cell.value:
                    max_len = max(max_len, len(str(cell.value)))
            ws.column_dimensions[col_letter].width = min(max_len + 2, 40)

        # Freeze header row
        ws.freeze_panes = "A2"

        wb.save(xlsx_path)
        print(f">> Excel updated: {xlsx_path}")
        print(f"   Total rows: {len(all_rows)} ({new_count} new)")

    except ImportError:
        print("  [WARNING] openpyxl not installed - Excel file not created")
    except PermissionError:
        print(f"  [ERROR] Cannot save Excel - close {OUTPUT_XLSX} and re-run")
    except Exception as e:
        print(f"  [WARNING] Excel error: {e}")


# ═══════════════════════════════════════════════════════════════════════════════
#  OUTPUT: SQLite DATABASE  (v2.6.0 - system of record)
# ═══════════════════════════════════════════════════════════════════════════════

def save_to_database(all_data, script_dir):
    """Ingest this run's records into propintel.db and refresh the spreadsheet
    view. Each record dict uses the same COLUMN_ORDER keys the CSV used, so the
    shared propintel_db.ingest_record() path handles dedup, repeat-lead
    flagging, and latest-wins property updates.

    The whole batch is one transaction. The spreadsheet (propintel_export.xlsx)
    is regenerated from the DB afterward as a convenience view; it is not the
    store, so it can be deleted and rebuilt at any time."""
    try:
        import propintel_db_v0_1_0 as pdb
    except ImportError as e:
        print(f"\n[ERROR] Could not import propintel_db_v0_1_0 ({e}).")
        print(f"        Records were NOT saved. Make sure "
              f"propintel_db_v0_1_0.py is in {script_dir}.")
        return

    conn = pdb.connect()
    pdb.init_db(conn)
    when = pdb._now()

    prop_new = prop_upd = le_new = le_collapsed = 0
    for rec in all_data:
        out = pdb.ingest_record(conn, rec, when)
        if out["property"] == "new":
            prop_new += 1
        else:
            prop_upd += 1
        if out["lead_event"] == "new":
            le_new += 1
        elif out["lead_event"] == "collapsed":
            le_collapsed += 1
    conn.commit()

    print(f"\n>> Saved to {pdb.DEFAULT_DB_FILENAME}: "
          f"{prop_new} new / {prop_upd} updated propert{'y' if (prop_new + prop_upd) == 1 else 'ies'}, "
          f"{le_new} lead event(s)"
          + (f", {le_collapsed} same-day duplicate(s) skipped" if le_collapsed else ""))

    # Regenerate the spreadsheet view from the DB (export, not the store).
    export_path = os.path.join(script_dir, "propintel_export.xlsx")
    try:
        n = pdb.export_xlsx(conn, export_path)
        print(f">> Spreadsheet view refreshed: {export_path} ({n} rows)")
    except ImportError:
        print(f"  [NOTE] openpyxl not installed; DB saved but no Excel view. "
              f"Export later with: python propintel_db_v0_1_0.py export propintel_export.xlsx")
    except PermissionError:
        print(f"  [WARNING] {export_path} is open in Excel - could not refresh "
              f"the view. The DB was saved. Close it and re-export with: "
              f"python propintel_db_v0_1_0.py export propintel_export.xlsx")
    except Exception as e:
        print(f"  [WARNING] Excel view refresh skipped: {e}")

    conn.close()


# ═══════════════════════════════════════════════════════════════════════════════
#  MAIN
# ═══════════════════════════════════════════════════════════════════════════════

def load_api_key(script_dir):
    """Load Google Maps API key from config.txt."""
    path = os.path.join(script_dir, "config.txt")
    if os.path.exists(path):
        with open(path, "r") as f:
            key = f.read().strip()
            if key:
                return key
    return None


def detect_input_format(content):
    """Detect whether file content is Market Leader, tab-separated, or plain addresses."""
    if any(kw in content for kw in MARKET_LEADER_KEYWORDS):
        return "market_leader"
    if "\t" in content and any(
        re.match(r'^[A-Za-z].*\t\d+\s', line.strip())
        for line in content.split("\n") if line.strip()
    ):
        return "tab_separated"
    return "plain_addresses"


def main():
    script_dir = os.path.dirname(os.path.abspath(__file__))
    lookups = []  # list of (search_address, lead_data_or_None)

    # --- Parse input ---
    if len(sys.argv) < 2:
        print("Usage: python snoco_scraper.py <address or file.txt>")
        sys.exit(1)

    if len(sys.argv) == 2 and sys.argv[1].endswith(".txt"):
        filepath = sys.argv[1]
        if not os.path.isabs(filepath):
            filepath = os.path.join(script_dir, filepath)

        with open(filepath, "r", encoding="utf-8") as f:
            content = f.read()

        fmt = detect_input_format(content)

        if fmt == "market_leader":
            leads = parse_lead_file(filepath)
            if not leads:
                print("[ERROR] Could not parse any leads from file")
                sys.exit(1)
            print(f"Detected Market Leader lead file: {len(leads)} lead(s) found")
            for lead in leads:
                addr = lead.get("address", "")
                if addr:
                    lead_data = build_lead_data(lead)
                    # Store the RAW address (with city). detect_county() must
                    # see the city to route correctly. scrape_property() and
                    # lookup_property_kc() clean the address themselves.
                    lookups.append((addr, lead_data))
                    print(f"  Lead: {lead.get('first_name', '')} {lead.get('last_name', '')} | {addr}")
                else:
                    print(f"  [WARNING] No address for lead: {lead.get('name', 'unknown')}")

        elif fmt == "tab_separated":
            print("Detected Name + Address format (tab-separated)")
            lookups = parse_tab_separated(content)

        else:
            addresses = [line.strip() for line in content.split("\n") if line.strip()]
            for addr in addresses:
                lookups.append((addr, None))
    else:
        # Command-line addresses
        for addr in sys.argv[1:]:
            lookups.append((addr, None))

    if not lookups:
        print("[ERROR] No addresses to look up")
        sys.exit(1)

    # --- Setup ---
    print(f"\nProperty Tax Scraper v2.6.0 (Snohomish web + King County bulk data, cached)")
    print(f"{'='*50}")
    print(f"Properties to look up: {len(lookups)}")

    api_key = load_api_key(script_dir)
    photo_folder = os.path.join(script_dir, "property_photos")
    print(f"Google API key loaded - Street View photos enabled" if api_key
          else "No config.txt found - Street View photos disabled")

    # Which counties does this run actually need?
    counties_needed = {detect_county(addr) for addr, _ in lookups}

    # v2.4.2: if any address is in a split-county city (e.g. Bothell), we may
    # need to retry the other county on a no-match, so make sure both data
    # sources are loaded for the run.
    if any(is_split_city(addr) for addr, _ in lookups):
        counties_needed.update({"snohomish", "king"})

    # Snohomish: live web scrape (only connect if a Snohomish address is present)
    session, base_resp = None, None
    if "snohomish" in counties_needed:
        print("\nConnecting to snoco.org...")
        try:
            session, base_resp = get_session()
            print("Connected!")
        except Exception as e:
            print(f"[WARNING] Could not connect to Snohomish County: {e}")
            session, base_resp = None, None

    # King County: load the bulk data ZIPs from disk (only if a KC address is present)
    kc_indexes = None
    if "king" in counties_needed:
        kc_indexes = kc_load_all(script_dir)

    if "snohomish" in counties_needed and not session \
            and "king" in counties_needed and not kc_indexes:
        print("[ERROR] Could not reach Snohomish County and could not load "
              "King County bulk data. Nothing to do.")
        sys.exit(1)
    if counties_needed == {"snohomish"} and not session:
        print("[ERROR] Could not connect to Snohomish County. Check your "
              "internet connection.")
        sys.exit(1)
    if counties_needed == {"king"} and not kc_indexes:
        print("[ERROR] Could not load King County bulk data. Check that the "
              "five ZIP files are in the script folder.")
        sys.exit(1)

    print()

    next_row = get_next_row_number(script_dir)
    today = datetime.now().strftime("%m/%d/%Y")

    # --- Look up each property ---
    all_data = []
    success = 0
    for i, (address, lead_data) in enumerate(lookups):
        county = detect_county(address)
        data = None

        if county == "king":
            if kc_indexes:
                # KC bulk lookup. Owner names are set inside lookup_property_kc()
                # from lead_data, since the KC public bulk file has no names.
                data = lookup_property_kc(address, kc_indexes, lead_data,
                                          api_key, photo_folder)
            else:
                print(f"\n[SKIP] {address} - King County bulk data unavailable")
        else:
            if session:
                data = scrape_property(session, address, base_resp, api_key, photo_folder)
            else:
                print(f"\n[SKIP] {address} - Snohomish County website unavailable")

        # v2.4.2: split-county-city fallback. If a primary lookup returned
        # nothing AND the city straddles the line (Bothell), try the other
        # county before giving up. This catches the v2.4.1 case where a
        # Snohomish-side Bothell parcel routed to King County and failed the
        # residential-only bulk index.
        if not data and is_split_city(address):
            other = "snohomish" if county == "king" else "king"
            print(f"  -> Split-county city, retrying as {other.title()} County...")
            if other == "king" and kc_indexes:
                data = lookup_property_kc(address, kc_indexes, lead_data,
                                          api_key, photo_folder)
                if data:
                    county = "king"
            elif other == "snohomish" and session:
                data = scrape_property(session, address, base_resp, api_key, photo_folder)
                if data:
                    county = "snohomish"

        if data:
            if county == "king":
                # KC: owner names already set from the lead inside
                # lookup_property_kc(). Merge the rest of the lead fields and
                # use the KC (lead-based) ownership match.
                if lead_data:
                    for k, v in lead_data.items():
                        # don't let blank lead fields clobber owner names
                        if k not in ("Owner 1 First Name", "Owner 1 Last Name"):
                            data[k] = v
                    data["Ownership Match"] = compute_ownership_match_kc(data, lead_data)
                    print(f"  -> Ownership Match: {data['Ownership Match']}")
                else:
                    data["Ownership Match"] = "NO LEAD DATA"
                    data["Lead Type"] = ""
            else:
                # Snohomish: county-confirmed owner names, full match logic.
                if lead_data:
                    data.update(lead_data)
                    data = fix_owner_names_with_lead(data, lead_data)
                    data["Ownership Match"] = compute_ownership_match(data, lead_data)
                    print(f"  -> Ownership Match: {data['Ownership Match']}")
                else:
                    data["Ownership Match"] = "NO LEAD DATA"
                    data["Lead Type"] = ""

            data["Row #"] = str(next_row)
            data["Date Retrieved"] = today
            next_row += 1
            all_data.append(data)
            success += 1

        # Refresh the Snohomish session every 10 properties (viewstate staleness).
        # King County needs no refresh: the bulk data is loaded once and held.
        if (i + 1) % 10 == 0 and (i + 1) < len(lookups):
            if session:
                print(f"\n  [Refreshing Snohomish session...]")
                try:
                    session, base_resp = get_session()
                except Exception:
                    pass
                time.sleep(1)

    # --- Write output ---
    # v2.6.0: the SQLite database (propintel.db) is the system of record. Each
    # scraped record is ingested with dedup + repeat-lead flagging via the
    # shared propintel_db.ingest_record() path. The CSV append model is retired;
    # a spreadsheet VIEW is regenerated from the DB after the run (export, not
    # the store), so re-running a lead no longer creates duplicate rows and the
    # old "close the xlsx and re-run" lock failure is gone.
    if all_data:
        save_to_database(all_data, script_dir)

    print(f"\n{'='*50}")
    print(f"DONE! {success}/{len(lookups)} properties scraped successfully.")


if __name__ == "__main__":
    main()
